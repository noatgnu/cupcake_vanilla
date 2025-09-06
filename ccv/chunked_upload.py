"""
Chunked upload implementation for CUPCAKE Vanilla metadata files.
"""

import os
from typing import Any, Dict, List

from drf_chunked_upload.serializers import ChunkedUploadSerializer
from drf_chunked_upload.views import ChunkedUploadView
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from ccc.chunked_upload import BaseChunkedUpload

from .models import MetadataColumn, MetadataTable, SamplePool
from .utils import detect_pooled_samples


class MetadataFileUpload(BaseChunkedUpload):
    """Custom chunked upload model for metadata files."""

    class Meta:
        app_label = "ccv"

    def get_allowed_extensions(self) -> List[str]:
        """Get list of allowed file extensions for metadata files."""
        return [".xlsx", ".xls", ".tsv", ".txt", ".sdrf"]

    def get_allowed_mime_types(self) -> List[str]:
        """Get list of allowed MIME types for metadata files."""
        return [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
            "application/vnd.ms-excel",  # .xls
            "text/tab-separated-values",  # .tsv
            "text/plain",  # .txt
            "application/octet-stream",  # Generic binary (fallback)
        ]


class MetadataFileUploadSerializer(ChunkedUploadSerializer):
    """Custom serializer for metadata file uploads."""

    class Meta:
        model = MetadataFileUpload
        fields = ("id", "file", "filename", "offset", "created_at", "status", "completed_at")
        read_only_fields = ("id", "created_at", "status", "completed_at")


class MetadataChunkedUploadView(ChunkedUploadView):
    """
    Chunked upload view for metadata files (SDRF, Excel, TSV).
    Supports large file uploads with progress tracking.
    """

    model = MetadataFileUpload
    serializer_class = MetadataFileUploadSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def on_completion(self, uploaded_file, request):
        """Handle completion of file upload and metadata processing."""
        try:
            # Get processing parameters from request
            metadata_table_id = request.data.get("metadata_table_id")
            create_pools = request.data.get("create_pools", True)
            replace_existing = request.data.get("replace_existing", False)

            if metadata_table_id:
                # Get target metadata table
                try:
                    metadata_table = MetadataTable.objects.get(id=metadata_table_id)

                    # Check permissions
                    if not metadata_table.can_edit(request.user):
                        return Response(
                            {"error": "Permission denied: cannot edit this metadata table"},
                            status=status.HTTP_403_FORBIDDEN,
                        )

                    # Process file based on type
                    file_path = uploaded_file.file.path
                    filename = uploaded_file.filename or uploaded_file.file.name
                    file_ext = os.path.splitext(filename.lower())[1]

                    if file_ext in [".xlsx", ".xls"]:
                        result = self._process_excel_file(
                            file_path,
                            metadata_table,
                            create_pools,
                            replace_existing,
                            request.user,
                        )
                    elif file_ext in [".tsv", ".txt", ".sdrf"]:
                        result = self._process_text_file(
                            file_path,
                            metadata_table,
                            create_pools,
                            replace_existing,
                            request.user,
                        )
                    else:
                        result = {"error": f"Unsupported file type: {file_ext}"}
                        return Response(result, status=status.HTTP_400_BAD_REQUEST)

                    # Clean up uploaded file
                    uploaded_file.delete()

                    return Response(result, status=status.HTTP_200_OK)

                except MetadataTable.DoesNotExist:
                    result = {"error": "Invalid metadata_table_id"}
                    return Response(result, status=status.HTTP_400_BAD_REQUEST)
            else:
                # No metadata table specified - just return success
                result = {"message": "File uploaded successfully", "filename": uploaded_file.filename}
                return Response(result, status=status.HTTP_200_OK)

        except Exception as e:
            # Log error but don't fail the upload completion
            import logging

            logger = logging.getLogger(__name__)
            logger.error(f"Failed to process metadata file: {str(e)}")
            result = {"warning": f"File uploaded but processing failed: {str(e)}"}
            return Response(result, status=status.HTTP_200_OK)

    def _process_excel_file(
        self,
        file_path: str,
        metadata_table: MetadataTable,
        create_pools: bool,
        replace_existing: bool,
        user: Any,
    ) -> Dict[str, Any]:
        """Process Excel file and extract metadata."""

        # Load Excel workbook
        wb = load_workbook(file_path, data_only=True)

        # Process main worksheet
        main_ws = wb["main"] if "main" in wb.sheetnames else wb.active

        # Convert to list of lists
        data = []
        for row in main_ws.iter_rows(values_only=True):
            if row and any(cell is not None for cell in row):  # Skip empty rows
                data.append([str(cell) if cell is not None else "" for cell in row])

        if not data:
            return {"error": "No data found in Excel file"}

        headers = data[0]
        data_rows = data[1:]

        return self._create_metadata_from_data(headers, data_rows, metadata_table, create_pools, replace_existing, user)

    def _process_text_file(
        self,
        file_path: str,
        metadata_table: MetadataTable,
        create_pools: bool,
        replace_existing: bool,
        user: Any,
    ) -> Dict[str, Any]:
        """Process text/TSV/SDRF file and extract metadata."""

        # Read file content
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read().strip()

        if not content:
            return {"error": "Empty file"}

        # Split into lines and parse
        lines = content.split("\n")

        # Detect delimiter (tab is preferred for SDRF)
        first_line = lines[0]
        delimiter = "\t" if "\t" in first_line else "," if "," in first_line else "\t"

        # Parse data
        headers = lines[0].split(delimiter)
        data_rows = [line.split(delimiter) for line in lines[1:] if line.strip()]

        return self._create_metadata_from_data(headers, data_rows, metadata_table, create_pools, replace_existing, user)

    def _create_metadata_from_data(
        self,
        headers: List[str],
        data_rows: List[List[str]],
        metadata_table: MetadataTable,
        create_pools: bool,
        replace_existing: bool,
        user: Any,
    ) -> Dict[str, Any]:
        """Create metadata columns and pools from parsed data."""

        created_columns = []
        created_pools = []

        # Clear existing columns if replace_existing is True
        if replace_existing:
            metadata_table.columns.all().delete()
            metadata_table.sample_pools.all().delete()

        # Update sample count based on data
        if data_rows:
            metadata_table.sample_count = len(data_rows)
            metadata_table.save(update_fields=["sample_count"])

        # Create metadata columns
        for i, header in enumerate(headers):
            # Parse header format: name[type] or just name
            if "[" in header and "]" in header:
                name = header.split("[")[0].strip()
                metadata_type = header.split("[")[1].rstrip("]").strip()
            else:
                name = header.strip()
                metadata_type = "characteristics"  # Default type

            if not name:  # Skip empty headers
                continue

            # Determine default value from first few non-empty rows
            default_value = ""
            for row in data_rows[:5]:  # Check first 5 rows
                if i < len(row) and row[i].strip():
                    default_value = row[i].strip()
                    break

            # Create metadata column first
            metadata_column = MetadataColumn.objects.create(
                metadata_table=metadata_table,
                name=name,
                type=metadata_type,
                column_position=i,
                value=default_value,
            )

            # Apply ontology mapping if needed
            from ccv.utils import apply_ontology_mapping_to_column

            apply_ontology_mapping_to_column(metadata_column)

            # Convert value using SDRF conventions via MetadataColumn instance method
            if default_value:
                converted_value = metadata_column.convert_sdrf_to_metadata(default_value)
                if converted_value is not None:
                    metadata_column.value = converted_value
                    metadata_column.save(update_fields=["value"])

            created_columns.append(metadata_column)

        # Detect and create pools if requested
        if create_pools:
            pooled_column_index, sn_rows, pooled_rows = detect_pooled_samples(data_rows, headers)

            if pooled_column_index is not None:
                # Create pools from detected data
                import_pools_data = []

                # Process SN= rows (reference pools)
                for sn_row_idx in sn_rows:
                    if sn_row_idx < len(data_rows):
                        row = data_rows[sn_row_idx]
                        if pooled_column_index < len(row):
                            sdrf_value = row[pooled_column_index].strip()

                            if sdrf_value.startswith("SN="):
                                source_names = sdrf_value[3:].split(",")
                                source_names = [name.strip() for name in source_names]

                                # Find source name column
                                source_name_column_index = None
                                for j, h in enumerate(headers):
                                    if "source name" in h.lower():
                                        source_name_column_index = j
                                        break

                                pool_name = (
                                    row[source_name_column_index]
                                    if source_name_column_index and source_name_column_index < len(row)
                                    else f"Pool {len(import_pools_data) + 1}"
                                )

                                # Find matching samples
                                pooled_only_samples = []
                                for sample_idx, sample_row in enumerate(data_rows):
                                    if (
                                        source_name_column_index
                                        and source_name_column_index < len(sample_row)
                                        and sample_row[source_name_column_index].strip() in source_names
                                    ):
                                        pooled_only_samples.append(sample_idx + 1)

                                import_pools_data.append(
                                    {
                                        "pool_name": pool_name,
                                        "pooled_only_samples": pooled_only_samples,
                                        "pooled_and_independent_samples": [],
                                        "is_reference": True,
                                        "sdrf_value": sdrf_value,
                                    }
                                )

                # Create SamplePool objects
                for pool_data in import_pools_data:
                    sample_pool = SamplePool.objects.create(
                        metadata_table=metadata_table,
                        pool_name=pool_data["pool_name"],
                        pooled_only_samples=pool_data["pooled_only_samples"],
                        pooled_and_independent_samples=pool_data["pooled_and_independent_samples"],
                        is_reference=pool_data["is_reference"],
                        created_by=metadata_table.creator,
                    )
                    created_pools.append(sample_pool)

        return {
            "message": "File processed successfully",
            "created_columns": len(created_columns),
            "created_pools": len(created_pools),
            "sample_rows": len(data_rows),
            "metadata_columns": [
                {
                    "id": col.id,
                    "name": col.name,
                    "type": col.type,
                    "position": col.column_position,
                }
                for col in created_columns
            ],
            "sample_pools": [
                {
                    "id": pool.id,
                    "name": pool.pool_name,
                    "total_samples": pool.get_total_samples(),
                    "is_reference": pool.is_reference,
                }
                for pool in created_pools
            ],
        }
