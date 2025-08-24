"""
Django REST Framework ViewSets for CUPCAKE Vanilla metadata management.
"""

import io
import json

from django.contrib.auth.models import User
from django.db import models
from django.db.models import Q
from django.http import HttpResponse

from django_filters.rest_framework import DjangoFilterBackend
from django_filters.views import FilterMixin
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.filters import SearchFilter
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from rest_framework.response import Response

from ccc.models import LabGroup

from .models import (
    CellOntology,
    ChEBICompound,
    FavouriteMetadataOption,
    HumanDisease,
    MetadataColumn,
    MetadataColumnTemplate,
    MetadataColumnTemplateShare,
    MetadataTable,
    MetadataTableTemplate,
    MondoDisease,
    MSUniqueVocabularies,
    NCBITaxonomy,
    PSIMSOntology,
    SamplePool,
    Schema,
    Species,
    SubcellularLocation,
    Tissue,
    UberonAnatomy,
    Unimod,
)
from .serializers import (
    CellOntologySerializer,
    ChEBICompoundSerializer,
    FavouriteMetadataOptionSerializer,
    HumanDiseaseSerializer,
    MetadataCollectionSerializer,
    MetadataColumnSerializer,
    MetadataColumnTemplateSerializer,
    MetadataColumnTemplateShareSerializer,
    MetadataExportSerializer,
    MetadataImportSerializer,
    MetadataTableSerializer,
    MetadataTableTemplateSerializer,
    MondoDiseaseSerializer,
    MSUniqueVocabulariesSerializer,
    NCBITaxonomySerializer,
    OntologySuggestionSerializer,
    PSIMSOntologySerializer,
    SamplePoolSerializer,
    SchemaSerializer,
    SpeciesSerializer,
    SubcellularLocationSerializer,
    TissueSerializer,
    UberonAnatomySerializer,
    UnimodSerializer,
)
from .utils import (
    apply_ontology_mapping_to_column,
    convert_sdrf_to_metadata,
    detect_ontology_type,
    sort_metadata,
    synchronize_pools_with_import_data,
    validate_sdrf,
    validate_sdrf_data_against_ontologies,
)


class MetadataTableViewSet(FilterMixin, viewsets.ModelViewSet):
    """ViewSet for managing MetadataTable objects."""

    queryset = MetadataTable.objects.all()
    serializer_class = MetadataTableSerializer
    permission_classes = [IsAuthenticated]
    search_fields = ["name", "description", "owner__username"]
    filterset_fields = ["owner", "lab_group", "is_published", "is_locked"]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Filter by owner
        owner_id = self.request.query_params.get("owner_id")
        if owner_id:
            queryset = queryset.filter(owner_id=owner_id)

        # Filter by lab group
        lab_group_id = self.request.query_params.get("lab_group_id")
        if lab_group_id:
            queryset = queryset.filter(lab_group_id=lab_group_id)

        # Filter by publication status
        is_published = self.request.query_params.get("is_published")
        if is_published is not None:
            queryset = queryset.filter(is_published=is_published.lower() == "true")

        # Filter by locked status
        is_locked = self.request.query_params.get("is_locked")
        if is_locked is not None:
            queryset = queryset.filter(is_locked=is_locked.lower() == "true")

        return queryset.order_by("-created_at", "name")

    @action(detail=True, methods=["post"])
    def add_column(self, request, pk=None):
        """Add a new column to this metadata table."""
        table = self.get_object()

        # Check permissions
        if not table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        column_data = request.data.get("column_data", {})
        position = request.data.get("position")

        try:
            column = table.add_column(column_data, position)
            serializer = MetadataColumnSerializer(column)
            return Response(
                {
                    "message": "Column added successfully",
                    "column": serializer.data,
                }
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to add column: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def remove_column(self, request, pk=None):
        """Remove a column from this metadata table."""
        table = self.get_object()

        # Check permissions
        if not table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        column_id = request.data.get("column_id")

        if not column_id:
            return Response(
                {"error": "column_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        success = table.remove_column(column_id)

        if success:
            return Response({"message": "Column removed successfully"})
        else:
            return Response(
                {"error": "Column not found or could not be removed"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def reorder_column(self, request, pk=None):
        """Reorder a column within this metadata table."""
        table = self.get_object()

        # Check permissions
        if not table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        column_id = request.data.get("column_id")
        new_position = request.data.get("new_position")

        if column_id is None or new_position is None:
            return Response(
                {"error": "column_id and new_position are required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        success = table.reorder_column(column_id, new_position)

        if success:
            return Response({"message": "Column reordered successfully"})
        else:
            return Response(
                {"error": "Column not found or could not be reordered"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def normalize_column_positions(self, request, pk=None):
        """Normalize column positions to be sequential starting from 0."""
        table = self.get_object()

        # Check permissions
        if not table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        table.normalize_column_positions()

        return Response({"message": "Column positions normalized successfully"})


class MetadataColumnViewSet(FilterMixin, viewsets.ModelViewSet):
    """ViewSet for managing MetadataColumn objects."""

    queryset = MetadataColumn.objects.all()
    serializer_class = MetadataColumnSerializer
    permission_classes = [IsAuthenticated]
    search_fields = ["name", "type", "value"]
    filterset_fields = ["metadata_table", "type", "mandatory", "hidden"]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Filter by metadata table
        metadata_table_id = self.request.query_params.get("metadata_table_id")
        if metadata_table_id:
            queryset = queryset.filter(metadata_table_id=metadata_table_id)

        # Filter by metadata type
        metadata_type = self.request.query_params.get("type")
        if metadata_type:
            queryset = queryset.filter(type__icontains=metadata_type)

        # Filter by name
        name = self.request.query_params.get("name")
        if name:
            queryset = queryset.filter(name__icontains=name)

        # Filter by hidden status
        hidden = self.request.query_params.get("hidden")
        if hidden is not None:
            queryset = queryset.filter(hidden=hidden.lower() == "true")

        return queryset.order_by("column_position", "name")

    @action(detail=False, methods=["post"])
    def bulk_create(self, request):
        """Create multiple metadata columns at once."""
        serializer = self.get_serializer(data=request.data, many=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["post"])
    def validate_sdrf_data(self, request):
        """Validate SDRF data format."""
        metadata_ids = request.data.get("metadata_ids", [])
        sample_number = request.data.get("sample_number", 1)

        metadata_columns = MetadataColumn.objects.filter(id__in=metadata_ids)
        result_data, _ = sort_metadata(list(metadata_columns), sample_number)

        errors = validate_sdrf(result_data)

        return Response(
            {
                "valid": len(errors) == 0,
                "errors": errors,
                "sample_count": len(result_data) - 1 if result_data else 0,
            }
        )

    @action(detail=False, methods=["get"])
    def ontology_suggestions(self, request):
        """
        Get ontology suggestions for a metadata column with enhanced search capabilities.

        Query Parameters:
        - column_id: ID of the metadata column
        - search: Search term to filter results
        - limit: Maximum number of results (default: 20, max: 100)
        - search_type: Type of search - 'icontains', 'istartswith', or 'exact' (default: 'icontains')
        """
        # Get column ID from query params
        column_id = request.query_params.get("column_id")
        if not column_id:
            return Response(
                {"error": "column_id parameter is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            column = MetadataColumn.objects.get(pk=column_id)
        except MetadataColumn.DoesNotExist:
            return Response(
                {"error": "Column not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        search_term = request.query_params.get("search", "")
        limit = min(int(request.query_params.get("limit", 20)), 100)  # Cap at 100
        search_type = request.query_params.get("search_type", "icontains")

        # Validate search_type
        valid_search_types = ["icontains", "istartswith", "exact"]
        if search_type not in valid_search_types:
            return Response(
                {"error": f"Invalid search_type. Must be one of: {', '.join(valid_search_types)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            suggestions = column.get_ontology_suggestions(search_term, limit, search_type)
        except Exception as e:
            return Response(
                {"error": f"Error retrieving suggestions: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # Serialize suggestions with proper field mappings
        serializer = OntologySuggestionSerializer(
            suggestions, many=True, context={"ontology_type": column.ontology_type}
        )

        return Response(
            {
                "ontology_type": column.ontology_type,
                "suggestions": serializer.data,
                "search_term": search_term,
                "search_type": search_type,
                "limit": limit,
                "count": len(suggestions),
                "custom_filters": column.custom_ontology_filters,
                "has_more": len(suggestions) >= limit,  # Indicate if there might be more results
            }
        )

    @action(detail=True, methods=["post"])
    def validate_value(self, request, pk=None):
        """Validate a value against the column's ontology."""
        column = self.get_object()
        value = request.data.get("value", "")

        is_valid = column.validate_value_against_ontology(value)

        response_data = {
            "valid": is_valid,
            "value": value,
            "ontology_type": column.ontology_type,
        }

        if not is_valid and column.ontology_type:
            # Provide suggestions for invalid values using enhanced search
            suggestions = column.get_ontology_suggestions(value, limit=5, search_type="icontains")
            response_data["suggestions"] = suggestions

        return Response(response_data)

    @action(detail=True, methods=["post"])
    def update_column_value(self, request, pk=None):
        """
        Update column value (default or sample-specific) with automatic modifier calculation.

        Request body:
        {
            "value": "new_value",
            "sample_indices": [1, 2, 3] or null,  # null means update default value
            "value_type": "default" | "sample_specific" | "replace_all"
        }

        Response:
        {
            "message": "Column value updated successfully",
            "column": {...},  # Updated column data
            "changes": {
                "old_default": "old_value",
                "new_default": "new_value",
                "old_modifiers": [...],
                "new_modifiers": [...],
                "updated_samples": [1, 2, 3]
            },
            "value_type": "sample_specific"
        }
        """
        column = self.get_object()

        # Check if user can edit this column's metadata table
        if not column.metadata_table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        value = request.data.get("value", "")
        sample_indices = request.data.get("sample_indices")  # List of 1-based sample numbers or null
        value_type = request.data.get("value_type", "default")  # "default", "sample_specific", "replace_all"

        # Validate value_type
        valid_types = ["default", "sample_specific", "replace_all"]
        if value_type not in valid_types:
            return Response(
                {"error": f"Invalid value_type. Must be one of: {', '.join(valid_types)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate sample_indices for sample_specific updates
        if value_type == "sample_specific":
            if not sample_indices or not isinstance(sample_indices, list):
                return Response(
                    {"error": "sample_indices must be a non-empty list for sample_specific updates"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Validate that sample indices are within valid range
            max_samples = column.metadata_table.sample_count
            for idx in sample_indices:
                if not isinstance(idx, int) or idx < 1 or idx > max_samples:
                    return Response(
                        {"error": f"Invalid sample index {idx}. Must be between 1 and {max_samples}"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

        try:
            # Use the smart update method from the model
            changes = column.update_column_value_smart(
                value=value, sample_indices=sample_indices, value_type=value_type
            )

            # Save the column
            column.save()

            # Return updated column data
            serializer = MetadataColumnSerializer(column)

            return Response(
                {
                    "message": "Column value updated successfully",
                    "column": serializer.data,
                    "changes": changes,
                    "value_type": value_type,
                }
            )

        except Exception as e:
            return Response(
                {"error": f"Error updating column value: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=["post"])
    def apply_ontology_mapping(self, request, pk=None):
        """Apply automatic ontology mapping to a column."""
        column = self.get_object()

        # Check if user can edit this column's metadata table
        if not column.metadata_table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        was_applied = apply_ontology_mapping_to_column(column)

        return Response(
            {
                "applied": was_applied,
                "ontology_type": column.ontology_type,
                "message": (f"Ontology mapping {'applied' if was_applied else 'already present or not detected'}"),
            }
        )

    @action(detail=False, methods=["get"])
    def detect_ontology_type(self, request):
        """Detect ontology type for given column name and type."""

        column_name = request.query_params.get("name", "")
        column_type = request.query_params.get("type", "")

        if not column_name:
            return Response(
                {"error": "column name is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        detected_type = detect_ontology_type(column_name, column_type)

        return Response(
            {
                "column_name": column_name,
                "column_type": column_type,
                "detected_ontology_type": detected_type,
            }
        )


class SamplePoolViewSet(FilterMixin, viewsets.ModelViewSet):
    """ViewSet for managing SamplePool objects."""

    queryset = SamplePool.objects.all()
    serializer_class = SamplePoolSerializer
    permission_classes = [IsAuthenticated]
    search_fields = ["name", "description"]
    filterset_fields = ["owner", "lab_group"]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Filter by metadata table
        metadata_table_id = self.request.query_params.get("metadata_table_id")
        if metadata_table_id:
            queryset = queryset.filter(metadata_table_id=metadata_table_id)

        # Filter by reference status
        is_reference = self.request.query_params.get("is_reference")
        if is_reference is not None:
            queryset = queryset.filter(is_reference=is_reference.lower() == "true")

        return queryset.order_by("pool_name")

    @action(detail=True, methods=["get"])
    def metadata_columns(self, request, pk=None):
        """Get metadata columns associated with this pool."""
        pool = self.get_object()

        # Get metadata columns for the same metadata table as the pool
        metadata_columns = MetadataColumn.objects.filter(metadata_table=pool.metadata_table)

        serializer = MetadataColumnSerializer(metadata_columns, many=True)
        return Response(serializer.data)


class MetadataTableTemplateViewSet(FilterMixin, viewsets.ModelViewSet):
    """ViewSet for managing MetadataTableTemplate objects."""

    queryset = MetadataTableTemplate.objects.all()
    serializer_class = MetadataTableTemplateSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ["name", "description"]
    filterset_fields = ["lab_group", "visibility", "is_default", "owner"]

    def get_queryset(self):
        """Filter queryset based on query parameters and user access."""
        queryset = super().get_queryset()

        # Filter by templates the user has access to:
        # 1. Templates created by the user
        # 2. Public templates
        # 3. Templates from lab groups the user is a member of
        user_groups = self.request.user.lab_groups.all()
        accessible_queryset = queryset.filter(
            Q(owner=self.request.user)
            | Q(visibility="public")  # User's own templates
            | Q(lab_group__in=user_groups)  # Public templates  # Lab group templates
        ).distinct()

        # Apply additional query parameter filters
        owner_id = self.request.query_params.get("owner_id")
        if owner_id:
            accessible_queryset = accessible_queryset.filter(owner_id=owner_id)

        # Filter by lab group
        lab_group_id = self.request.query_params.get("lab_group_id")
        if lab_group_id:
            accessible_queryset = accessible_queryset.filter(lab_group_id=lab_group_id)

        # Filter by public status
        is_public = self.request.query_params.get("is_public")
        if is_public is not None:
            if is_public.lower() == "true":
                accessible_queryset = accessible_queryset.filter(visibility="public")
            else:
                accessible_queryset = accessible_queryset.exclude(visibility="public")

        # Filter by default status
        is_default = self.request.query_params.get("is_default")
        if is_default is not None:
            accessible_queryset = accessible_queryset.filter(is_default=is_default.lower() == "true")

        return accessible_queryset.order_by("-is_default", "name")

    @action(detail=True, methods=["post"])
    def add_column(self, request, pk=None):
        """Add a new column to this template."""
        template = self.get_object()

        # Check permissions
        if not template.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        column_data = request.data.get("column_data", {})
        position = request.data.get("position")

        try:
            column = template.add_column_to_template(column_data, position)
            serializer = MetadataColumnSerializer(column)
            return Response(
                {
                    "message": "Column added successfully",
                    "column": serializer.data,
                }
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to add column: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def remove_column(self, request, pk=None):
        """Remove a column from this template."""
        template = self.get_object()

        # Check permissions
        if not template.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        column_id = request.data.get("column_id")

        if not column_id:
            return Response(
                {"error": "column_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        success = template.remove_column_from_template(column_id)

        if success:
            return Response({"message": "Column removed successfully"})
        else:
            return Response(
                {"error": "Column not found or could not be removed"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def reorder_column(self, request, pk=None):
        """Reorder a column within this template."""
        template = self.get_object()

        # Check permissions
        if not template.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        column_id = request.data.get("column_id")
        new_position = request.data.get("new_position")

        if column_id is None or new_position is None:
            return Response(
                {"error": "column_id and new_position are required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        success = template.reorder_template_column(column_id, new_position)

        if success:
            return Response({"message": "Column reordered successfully"})
        else:
            return Response(
                {"error": "Column not found or could not be reordered"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def duplicate_column(self, request, pk=None):
        """Duplicate a column within this template."""
        template = self.get_object()

        # Check permissions
        if not template.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        column_id = request.data.get("column_id")
        new_name = request.data.get("new_name")

        if not column_id:
            return Response(
                {"error": "column_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        duplicated_column = template.duplicate_column_in_template(column_id, new_name)

        if duplicated_column:
            serializer = MetadataColumnSerializer(duplicated_column)
            return Response(
                {
                    "message": "Column duplicated successfully",
                    "column": serializer.data,
                }
            )
        else:
            return Response(
                {"error": "Column not found or could not be duplicated"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def normalize_column_positions(self, request, pk=None):
        """Normalize column positions to be sequential starting from 0."""
        template = self.get_object()

        # Check permissions
        if not template.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        template.normalize_template_column_positions()

        return Response({"message": "Column positions normalized successfully"})

    @action(detail=True, methods=["post"])
    def apply_to_metadata_table(self, request, pk=None):
        """Apply this template to a metadata table."""
        template = self.get_object()
        metadata_table_id = request.data.get("metadata_table_id")

        if not metadata_table_id:
            return Response(
                {"error": "metadata_table_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            metadata_table = MetadataTable.objects.get(id=metadata_table_id)
        except MetadataTable.DoesNotExist:
            return Response(
                {"error": "Invalid metadata_table_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check permissions
        if not metadata_table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Create metadata columns for the metadata table based on template
        created_columns = []
        for template_column in template.user_columns.all():
            new_column = MetadataColumn.objects.create(
                name=template_column.name,
                type=template_column.type,
                column_position=template_column.column_position,
                value=template_column.value,
                mandatory=template_column.mandatory,
                hidden=template_column.hidden,
                auto_generated=template_column.auto_generated,
                readonly=template_column.readonly,
                modifiers=template_column.modifiers,
                metadata_table=metadata_table,
            )
            created_columns.append(new_column)

        serializer = MetadataColumnSerializer(created_columns, many=True)
        return Response(
            {
                "message": f"Template applied successfully. Created {len(created_columns)} metadata columns.",
                "created_columns": serializer.data,
            }
        )

    @action(detail=False, methods=["get"])
    def available_schemas(self, request):
        """Get list of available schemas for creating templates."""

        # Get schemas available to the user
        schemas = Schema.get_available_schemas(user=request.user)

        schema_info = []
        for schema in schemas:
            schema_info.append(
                {
                    "id": schema.id,
                    "name": schema.name,
                    "display_name": schema.display_name,
                    "description": schema.description,
                    "is_builtin": schema.is_builtin,
                    "tags": schema.tags,
                    "usage_count": schema.usage_count,
                    "file_size": schema.file_size,
                    "created_at": schema.created_at.isoformat() if schema.created_at else None,
                }
            )

        return Response(schema_info)

    @action(detail=False, methods=["post"])
    def create_from_schema(self, request):
        """Create a new template from schema definitions."""

        # Validate required fields
        name = request.data.get("name")
        schema_ids = request.data.get("schema_ids", [])

        # Support legacy 'schemas' parameter for backward compatibility
        if not schema_ids and "schemas" in request.data:
            schema_names = request.data.get("schemas", [])
            # Convert names to IDs
            available_schemas = Schema.get_available_schemas(user=request.user)
            schema_name_to_id = {s.name: s.id for s in available_schemas}
            schema_ids = [schema_name_to_id.get(name) for name in schema_names if name in schema_name_to_id]
            schema_ids = [sid for sid in schema_ids if sid is not None]

        if not name:
            return Response({"error": "Template name is required"}, status=status.HTTP_400_BAD_REQUEST)

        if not schema_ids:
            # Default to minimum schema
            try:
                minimum_schema = Schema.objects.get(name="minimum", is_active=True)
                schema_ids = [minimum_schema.id]
            except Schema.DoesNotExist:
                return Response(
                    {"error": "No default schema available. Please specify schema_ids."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Validate schemas exist and user has access
        available_schemas = Schema.get_available_schemas(user=request.user)
        available_ids = {s.id for s in available_schemas}
        invalid_schema_ids = [sid for sid in schema_ids if sid not in available_ids]
        if invalid_schema_ids:
            return Response(
                {"error": f'Invalid or inaccessible schema IDs: {", ".join(map(str, invalid_schema_ids))}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Get optional parameters
            description = request.data.get("description")
            lab_group_id = request.data.get("lab_group_id")
            is_public = request.data.get("is_public", False)
            is_default = request.data.get("is_default", False)

            # Convert is_public to visibility
            visibility = "public" if is_public else "private"

            # Get lab group if specified
            lab_group = None
            if lab_group_id:
                try:
                    lab_group = LabGroup.objects.get(id=lab_group_id)
                    # Check if user is member of this lab group
                    if not lab_group.members.filter(id=request.user.id).exists():
                        return Response(
                            {"error": "You are not a member of the specified lab group"},
                            status=status.HTTP_403_FORBIDDEN,
                        )
                except LabGroup.DoesNotExist:
                    return Response({"error": "Lab group not found"}, status=status.HTTP_404_NOT_FOUND)

            # Create template from schemas
            template = MetadataTableTemplate.create_from_schemas(
                name=name,
                schema_ids=schema_ids,
                owner=request.user,
                lab_group=lab_group,
                description=description,
                visibility=visibility,
                is_default=is_default,
            )

            # Serialize and return the created template
            serializer = self.get_serializer(template)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {"error": f"Failed to create template: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=["post"])
    def create_table_from_template(self, request):
        """Create a new metadata table from an existing template."""
        # Validate required fields
        name = request.data.get("name")
        template_id = request.data.get("template_id")

        if not name:
            return Response({"error": "Table name is required"}, status=status.HTTP_400_BAD_REQUEST)

        if not template_id:
            return Response({"error": "Template ID is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Get the template
            template = MetadataTableTemplate.objects.get(id=template_id)

            # Check if user has access to this template
            if not template.can_view(request.user):
                return Response({"error": "You do not have access to this template"}, status=status.HTTP_403_FORBIDDEN)

            # Get optional parameters
            description = request.data.get("description")
            sample_count = request.data.get("sample_count", 1)
            lab_group_id = request.data.get("lab_group_id")

            # Get lab group if specified
            lab_group = None
            if lab_group_id:
                try:
                    lab_group = LabGroup.objects.get(id=lab_group_id)
                    # Check if user is member of this lab group
                    if not lab_group.members.filter(id=request.user.id).exists():
                        return Response(
                            {"error": "You are not a member of the specified lab group"},
                            status=status.HTTP_403_FORBIDDEN,
                        )
                except LabGroup.DoesNotExist:
                    return Response({"error": "Lab group not found"}, status=status.HTTP_404_NOT_FOUND)

            # Create metadata table from template
            metadata_table = template.create_table_from_template(
                table_name=name,
                owner=request.user,
                sample_count=sample_count,
                description=description,
                lab_group=lab_group,
            )

            # Serialize and return the created table
            serializer = MetadataTableSerializer(metadata_table)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        except MetadataTableTemplate.DoesNotExist:
            return Response({"error": "Template not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response(
                {"error": f"Failed to create table from template: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["post"])
    def create_table_from_schemas(self, request):
        """Create a new metadata table directly from schema definitions."""

        # Validate required fields
        name = request.data.get("name")
        schema_ids = request.data.get("schema_ids", [])

        # Support legacy 'schemas' parameter for backward compatibility
        if not schema_ids and "schemas" in request.data:
            schema_names = request.data.get("schemas", [])
            # Convert names to IDs
            available_schemas = Schema.get_available_schemas(user=request.user)
            schema_name_to_id = {s.name: s.id for s in available_schemas}
            schema_ids = [schema_name_to_id.get(name) for name in schema_names if name in schema_name_to_id]
            schema_ids = [sid for sid in schema_ids if sid is not None]

        if not name:
            return Response({"error": "Table name is required"}, status=status.HTTP_400_BAD_REQUEST)

        if not schema_ids:
            # Default to minimum schema
            try:
                minimum_schema = Schema.objects.get(name="minimum", is_active=True)
                schema_ids = [minimum_schema.id]
            except Schema.DoesNotExist:
                return Response(
                    {"error": "No default schema available. Please specify schema_ids."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Validate schemas exist and user has access
        available_schemas = Schema.get_available_schemas(user=request.user)
        available_ids = {s.id for s in available_schemas}
        invalid_schema_ids = [sid for sid in schema_ids if sid not in available_ids]
        if invalid_schema_ids:
            return Response(
                {"error": f'Invalid or inaccessible schema IDs: {", ".join(map(str, invalid_schema_ids))}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Get optional parameters
            description = request.data.get("description")
            sample_count = request.data.get("sample_count", 1)
            lab_group_id = request.data.get("lab_group_id")

            # Get lab group if specified
            lab_group = None
            if lab_group_id:
                try:
                    lab_group = LabGroup.objects.get(id=lab_group_id)
                    # Check if user is member of this lab group
                    if not lab_group.members.filter(id=request.user.id).exists():
                        return Response(
                            {"error": "You are not a member of the specified lab group"},
                            status=status.HTTP_403_FORBIDDEN,
                        )
                except LabGroup.DoesNotExist:
                    return Response({"error": "Lab group not found"}, status=status.HTTP_404_NOT_FOUND)

            # Create a temporary template to leverage the schema functionality
            temp_template = MetadataTableTemplate(
                name=f"temp_template_for_{name}",
                owner=request.user,
                lab_group=lab_group,
                visibility="private",  # Set default visibility
            )

            # Create metadata table from schemas using the template method
            metadata_table = temp_template.create_table_from_schemas(
                table_name=name,
                schema_ids=schema_ids,
                owner=request.user,
                sample_count=sample_count,
                description=description,
                lab_group=lab_group,
            )

            # Serialize and return the created table
            serializer = MetadataTableSerializer(metadata_table)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {"error": f"Failed to create table from schemas: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class FavouriteMetadataOptionViewSet(FilterMixin, viewsets.ModelViewSet):
    """ViewSet for managing FavouriteMetadataOption objects."""

    queryset = FavouriteMetadataOption.objects.all()
    serializer_class = FavouriteMetadataOptionSerializer
    permission_classes = [IsAuthenticated]
    search_fields = ["name", "value", "display_value"]
    filterset_fields = ["is_global", "user", "lab_group"]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Filter by user
        user_id = self.request.query_params.get("user_id")
        if user_id:
            queryset = queryset.filter(user_id=user_id)

        # Filter by lab group
        lab_group_id = self.request.query_params.get("lab_group_id")
        if lab_group_id:
            queryset = queryset.filter(lab_group_id=lab_group_id)

        # Filter by global status
        is_global = self.request.query_params.get("is_global")
        if is_global is not None:
            queryset = queryset.filter(is_global=is_global.lower() == "true")

        # Filter by metadata name
        name = self.request.query_params.get("name")
        if name:
            queryset = queryset.filter(name__icontains=name)

        # Filter by metadata type
        metadata_type = self.request.query_params.get("type")
        if metadata_type:
            queryset = queryset.filter(type__icontains=metadata_type)

        return queryset.order_by("name", "type")


class MetadataManagementViewSet(viewsets.GenericViewSet):
    """ViewSet for metadata management operations like import/export."""

    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=["post"])
    def export_excel_template(self, request):
        """Export metadata as Excel template (matching original CUPCAKE)."""
        serializer = MetadataExportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data

        # Get metadata table
        try:
            metadata_table = MetadataTable.objects.get(id=data["metadata_table_id"])
        except (MetadataTable.DoesNotExist, KeyError):
            return Response(
                {"error": "metadata_table_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check permissions
        if not metadata_table.can_view(request.user):
            return Response(
                {"error": "Permission denied: cannot view this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Get metadata columns (user can specify specific columns or get all)
        if data.get("metadata_column_ids"):
            metadata_columns = metadata_table.columns.filter(id__in=data["metadata_column_ids"])
        else:
            metadata_columns = metadata_table.columns.all()

        # Separate main and hidden metadata (original CUPCAKE logic)
        main_metadata = [m for m in metadata_columns if not m.hidden]
        hidden_metadata = [m for m in metadata_columns if m.hidden]

        # Sort metadata and get structured output (original CUPCAKE approach)
        result_main, id_map_main = sort_metadata(main_metadata, metadata_table.sample_count)
        result_hidden = []
        id_map_hidden = {}
        if hidden_metadata:
            result_hidden, id_map_hidden = sort_metadata(hidden_metadata, metadata_table.sample_count)

        # Get pools and prepare pool data (original CUPCAKE logic)
        pools = list(metadata_table.sample_pools.all())
        has_pools = len(pools) > 0
        pool_id_map_main, pool_id_map_hidden = {}, {}

        if has_pools and data.get("include_pools", True):
            # Note: Pools in ccv don't have separate metadata like original CUPCAKE
            # We'll use the same metadata structure but organize by pool
            pass

        # Get favourites for each metadata column (original CUPCAKE logic)
        favourites = {}

        # User-specific favourites
        user_favourites = FavouriteMetadataOption.objects.filter(
            user=request.user, service_lab_group__isnull=True, lab_group__isnull=True
        )
        for fav in user_favourites:
            if fav.name.lower() not in favourites:
                favourites[fav.name.lower()] = []
            favourites[fav.name.lower()].append(f"{fav.display_value}[*]")

        # Lab group favourites
        if data.get("lab_group_id"):
            try:
                lab_group = LabGroup.objects.get(id=data["lab_group_id"])
                lab_favourites = FavouriteMetadataOption.objects.filter(lab_group=lab_group)
                for fav in lab_favourites:
                    if fav.name.lower() not in favourites:
                        favourites[fav.name.lower()] = []
                    favourites[fav.name.lower()].append(f"{fav.display_value}[**]")
                    # Add "not applicable" for required metadata
                    if fav.name.lower() == "tissue" or fav.name.lower() == "organism part":
                        favourites[fav.name.lower()].append("not applicable")
            except LabGroup.DoesNotExist:
                pass

        # Global recommendations
        global_favourites = FavouriteMetadataOption.objects.filter(is_global=True)
        for fav in global_favourites:
            if fav.name.lower() not in favourites:
                favourites[fav.name.lower()] = []
            favourites[fav.name.lower()].append(f"{fav.display_value}[***]")

        # Create Excel workbook with multiple sheets (original CUPCAKE structure)
        wb = Workbook()
        main_ws = wb.active
        main_ws.title = "main"
        hidden_ws = wb.create_sheet(title="hidden")
        id_metadata_column_map_ws = wb.create_sheet(title="id_metadata_column_map")

        # Create pool sheets if pools exist
        pool_id_metadata_column_map_ws = None
        pool_object_map_ws = None

        if has_pools and data.get("include_pools", True):
            # Note: Pool worksheets created but not fully utilized in simplified ccv
            wb.create_sheet(title="pool_main")
            wb.create_sheet(title="pool_hidden")
            pool_id_metadata_column_map_ws = wb.create_sheet(title="pool_id_metadata_column_map")
            pool_object_map_ws = wb.create_sheet(title="pool_object_map")

        # Fill ID metadata column mapping (original CUPCAKE format)
        id_metadata_column_map_ws.append(["id", "column", "name", "type", "hidden"])
        for k, v in id_map_main.items():
            id_metadata_column_map_ws.append([k, v["column"], v["name"], v["type"], v["hidden"]])
        for k, v in id_map_hidden.items():
            id_metadata_column_map_ws.append([k, v["column"], v["name"], v["type"], v["hidden"]])

        # Fill pool ID mapping if pools exist
        if has_pools and data.get("include_pools", True) and pool_id_metadata_column_map_ws:
            pool_id_metadata_column_map_ws.append(["id", "column", "name", "type", "hidden"])
            for k, v in pool_id_map_main.items():
                pool_id_metadata_column_map_ws.append([k, v["column"], v["name"], v["type"], v["hidden"]])
            for k, v in pool_id_map_hidden.items():
                pool_id_metadata_column_map_ws.append([k, v["column"], v["name"], v["type"], v["hidden"]])

            # Fill pool object mapping sheet
            pool_object_map_ws.append(
                [
                    "pool_name",
                    "pooled_only_samples",
                    "pooled_and_independent_samples",
                    "is_reference",
                    "sdrf_value",
                ]
            )
            for pool in pools:
                pool_object_map_ws.append(
                    [
                        pool.pool_name,
                        (json.dumps(pool.pooled_only_samples) if pool.pooled_only_samples else "[]"),
                        (
                            json.dumps(pool.pooled_and_independent_samples)
                            if pool.pooled_and_independent_samples
                            else "[]"
                        ),
                        pool.is_reference,
                        pool.sdrf_value or "",
                    ]
                )

        # Excel styling (original CUPCAKE formatting)
        fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Populate main worksheet (original CUPCAKE logic)
        if result_main:
            main_ws.append(result_main[0])
            main_work_area = f"A1:{get_column_letter(len(result_main[0]))}{metadata_table.sample_count + 1}"

            for row in result_main[1:]:
                main_ws.append(row)

            # Apply styling
            for row in main_ws[main_work_area]:
                for cell in row:
                    cell.fill = fill
                    cell.border = thin_border

            # Auto-adjust column widths
            for col in main_ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = max_length + 2
                main_ws.column_dimensions[column].width = adjusted_width

            # Add informational notes (original CUPCAKE notes)
            note_texts = [
                "Note: Cells that are empty will automatically be filled with 'not applicable' or "
                "'not available' depending on the column when submitted.",
                "[*] User-specific favourite options.",
                "[**] Lab group-recommended options.",
                "[***] Global recommendations.",
            ]

            start_row = metadata_table.sample_count + 2
            for i, note_text in enumerate(note_texts):
                main_ws.merge_cells(
                    start_row=start_row + i,
                    start_column=1,
                    end_row=start_row + i,
                    end_column=len(result_main[0]),
                )
                note_cell = main_ws.cell(row=start_row + i, column=1)
                note_cell.value = note_text
                note_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Populate hidden worksheet (original CUPCAKE logic)
        if result_hidden:
            hidden_work_area = f"A1:{get_column_letter(len(result_hidden[0]))}{metadata_table.sample_count + 1}"
            hidden_ws.append(result_hidden[0])
            for row in result_hidden[1:]:
                hidden_ws.append(row)

            # Apply styling
            for row in hidden_ws[hidden_work_area]:
                for cell in row:
                    cell.fill = fill
                    cell.border = thin_border

            # Auto-adjust column widths
            for col in hidden_ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = max_length + 2
                hidden_ws.column_dimensions[column].width = adjusted_width

        # Add data validation dropdowns for main worksheet (original CUPCAKE logic)
        if result_main:
            for i, header in enumerate(result_main[0]):
                name_parts = header.split("[")
                name = name_parts[1].replace("]", "") if len(name_parts) > 1 else name_parts[0]

                # Build option list
                option_list = []
                # Required columns get "not applicable", others get "not available"
                if name.lower() in [
                    "organism",
                    "disease",
                    "organism part",
                    "tissue",
                    "biological replicate",
                ]:
                    option_list.append("not applicable")
                else:
                    option_list.append("not available")

                # Add favourites if available
                if name.lower() in favourites:
                    option_list.extend(favourites[name.lower()])

                # Create data validation
                if option_list:
                    dv = DataValidation(
                        type="list",
                        formula1=f'"{",".join(option_list)}"',
                        showDropDown=False,
                    )
                    col_letter = get_column_letter(i + 1)
                    main_ws.add_data_validation(dv)
                    dv.add(f"{col_letter}2:{col_letter}{metadata_table.sample_count + 1}")

        # Add data validation dropdowns for hidden worksheet (original CUPCAKE logic)
        if result_hidden:
            for i, header in enumerate(result_hidden[0]):
                name_parts = header.split("[")
                name = name_parts[1].replace("]", "") if len(name_parts) > 1 else name_parts[0]

                # Build option list
                option_list = []
                # Required columns get "not applicable", others get "not available"
                if name.lower() in [
                    "organism",
                    "disease",
                    "organism part",
                    "tissue",
                    "biological replicate",
                ]:
                    option_list.append("not applicable")
                else:
                    option_list.append("not available")

                # Add favourites if available
                if name.lower() in favourites:
                    option_list.extend(favourites[name.lower()])

                # Create data validation
                if option_list:
                    dv = DataValidation(
                        type="list",
                        formula1=f'"{",".join(option_list)}"',
                        showDropDown=False,
                    )
                    col_letter = get_column_letter(i + 1)
                    hidden_ws.add_data_validation(dv)
                    dv.add(f"{col_letter}2:{col_letter}{metadata_table.sample_count + 1}")

        # Return Excel file as HTTP response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            content=output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="metadata_template.xlsx"'
        return response

    def _find_matching_column_template(self, column_name, metadata_table=None, templates=None):
        """
        Intelligent column template matching for SDRF import.

        Args:
            column_name: Name of the column from SDRF
            metadata_table: Target metadata table to search in first
            templates: List of available templates (from table's associated templates)

        Returns:
            MetadataColumnTemplate or None
        """
        # Normalize column name for matching
        normalized_name = column_name.lower().strip()

        # Step 1: Look in current table's existing columns
        if metadata_table:
            existing_column = metadata_table.columns.filter(name__iexact=column_name).first()
            if existing_column and existing_column.template:
                return existing_column.template

        # Step 2: Look in table's associated templates (from schemas)
        if templates:
            # Try exact match first
            exact_match = None
            for template in templates:
                if template.column_name.lower() == normalized_name:
                    exact_match = template
                    break

            if exact_match:
                return exact_match

            # Try partial match
            partial_matches = []
            for template in templates:
                template_name_lower = template.column_name.lower()
                # Check if column name contains template name or vice versa
                if normalized_name in template_name_lower or template_name_lower in normalized_name:
                    partial_matches.append(template)

            # Return the best partial match (shortest name = more specific)
            if partial_matches:
                return min(partial_matches, key=lambda t: len(t.column_name))

        # Step 3: Look in global/system templates
        # Try exact match in system templates
        system_template = MetadataColumnTemplate.objects.filter(
            column_name__iexact=column_name, is_system_template=True, is_active=True
        ).first()

        if system_template:
            return system_template

        # Try partial match in system templates
        system_partial = MetadataColumnTemplate.objects.filter(
            column_name__icontains=normalized_name, is_system_template=True, is_active=True
        ).first()

        return system_partial

    def _get_column_type_from_name(self, column_name):
        """Determine column type based on SDRF naming conventions."""
        name_lower = column_name.lower()

        if name_lower.startswith("characteristics[") or "characteristics" in name_lower:
            return "characteristics"
        elif name_lower.startswith("factor value[") or "factor value" in name_lower:
            return "factor_value"
        elif name_lower.startswith("comment[") or "comment" in name_lower:
            return "comment"
        elif name_lower in ["source name", "source_name"]:
            return "source_name"
        elif name_lower in ["assay name", "assay_name"]:
            return "assay_name"
        elif name_lower in ["technology name", "technology_name"]:
            return "technology_name"
        else:
            return "characteristics"  # Default fallback

    def _organize_columns_by_schema(self, metadata_table, created_columns):
        """
        Reorder columns based on schemas found in the table's templates.

        Organizing by sections and putting unmatched columns last.
        """
        # Get all schemas associated with the table's templates
        table_templates = []

        # Get templates from table's lab group if it exists
        if metadata_table.lab_group:
            table_templates.extend(
                MetadataColumnTemplate.objects.filter(lab_group=metadata_table.lab_group, is_active=True)
            )

        # Get templates from the table owner
        if metadata_table.owner:
            table_templates.extend(MetadataColumnTemplate.objects.filter(owner=metadata_table.owner, is_active=True))

        # Get system templates
        table_templates.extend(MetadataColumnTemplate.objects.filter(is_system_template=True, is_active=True))

        # Get unique schemas from these templates
        schema_ids = set()
        for template in table_templates:
            if hasattr(template, "source_schema") and template.source_schema:
                try:
                    schema = Schema.objects.get(name=template.source_schema, is_active=True)
                    schema_ids.add(schema.id)
                except Schema.DoesNotExist:
                    continue

        # If we have schemas, use the table's reorder method
        if schema_ids:
            metadata_table.reorder_columns_by_schema(schema_ids=list(schema_ids))
        else:
            # Fallback: basic section-based ordering
            self._basic_column_reordering(metadata_table, created_columns)

        # Apply the same reordering logic to sample pool columns
        created_pools = list(metadata_table.sample_pools.all())
        if created_pools:
            for pool in created_pools:
                if pool.metadata_columns.exists():
                    if schema_ids:
                        try:
                            pool.reorder_pool_columns_by_schema(schema_ids=list(schema_ids))
                        except Exception as e:
                            # Log error but don't fail the import, fall back to basic reordering
                            print(f"Warning: Failed to reorder pool '{pool.pool_name}' columns by schema: {e}")
                            pool.basic_pool_column_reordering()
                    else:
                        # No schemas found, use basic section-based ordering for pools
                        pool.basic_pool_column_reordering()

    def _basic_column_reordering(self, metadata_table, created_columns):
        """Basic column reordering by section when no schemas are available."""
        sections = {
            "source_name": [],
            "characteristics": [],
            "assay_name": [],
            "technology_name": [],
            "comment": [],
            "factor_value": [],
            "other": [],
        }

        # Group columns by type
        for column in metadata_table.columns.all():
            col_type = column.type.lower()
            if col_type in sections:
                sections[col_type].append(column)
            else:
                sections["other"].append(column)

        # Assign positions
        position = 0
        section_order = [
            "source_name",
            "characteristics",
            "assay_name",
            "technology_name",
            "comment",
            "factor_value",
            "other",
        ]

        for section in section_order:
            for column in sections[section]:
                column.column_position = position
                column.save(update_fields=["column_position"])
                position += 1

    def _get_table_template(self, metadata_table):
        """Get the table template that existing columns originate from."""

        # Find table template by looking at existing columns' template references
        existing_columns = list(metadata_table.columns.all())
        if existing_columns:
            # Look for columns that have template references
            template_column_templates = set()
            for col in existing_columns:
                if col.template:
                    template_column_templates.add(col.template)

            if template_column_templates:
                # Find which table template contains the most of these column templates
                best_template = None
                best_match_count = 0

                for table_template in MetadataTableTemplate.objects.all():
                    # Get all template columns in this table template
                    table_template_columns = set()
                    for template_col in table_template.user_columns.all():
                        if hasattr(template_col, "template") and template_col.template:
                            table_template_columns.add(template_col.template)

                    # Count matches
                    match_count = len(template_column_templates.intersection(table_template_columns))
                    if match_count > best_match_count:
                        best_match_count = match_count
                        best_template = table_template

                if best_template and best_match_count > 0:
                    return best_template

        # If no template found from existing columns, try to infer from table metadata
        # This could be enhanced to look at table creation history or naming patterns
        return None

    def _find_or_create_matching_column(
        self, clean_name, metadata_type, metadata_table, table_template, column_position, occurrence_number
    ):
        """Find existing column or create new one using template properties."""

        # Get all existing columns with the same name (case insensitive)
        existing_columns = list(metadata_table.columns.filter(name__iexact=clean_name).order_by("column_position"))

        if occurrence_number <= len(existing_columns):
            # Replace content of existing column at this occurrence
            existing_column = existing_columns[occurrence_number - 1]
            existing_column.column_position = column_position
            existing_column.save(update_fields=["column_position"])
            return existing_column
        else:
            # Create new column (occurrence_number > existing columns count)
            # Look for template column to copy properties from
            template_column = None
            if table_template:
                # Try to find exact match in template
                template_column = table_template.user_columns.filter(name__iexact=clean_name).first()

            # Create new column with template properties if available
            if template_column:
                # Create column using template properties (copy, don't reference existing template column)
                # Get the template column's type safely
                template_type = getattr(template_column, "type", None) or getattr(template_column, "column_type", "")

                metadata_column = MetadataColumn.objects.create(
                    name=clean_name,
                    type=metadata_type or template_type,
                    column_position=column_position,
                    metadata_table=metadata_table,
                    template=getattr(template_column, "template", None),  # Use the original template reference
                    ontology_type=getattr(template_column, "ontology_type", None),
                    mandatory=getattr(template_column, "mandatory", False),
                    hidden=getattr(template_column, "hidden", False),
                    readonly=getattr(template_column, "readonly", False),
                    auto_generated=getattr(template_column, "auto_generated", False),
                )
            else:
                # Create basic column without template
                metadata_column = MetadataColumn.objects.create(
                    name=clean_name,
                    type=metadata_type,
                    column_position=column_position,
                    metadata_table=metadata_table,
                )

                # Apply automatic ontology mapping
                apply_ontology_mapping_to_column(metadata_column)

            return metadata_column

    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser])
    def import_sdrf_file(self, request):
        """Import metadata from SDRF file with intelligent column matching and schema organization."""
        serializer = MetadataImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data

        try:
            # Read and parse SDRF file
            file_content = data["file"].read().decode("utf-8")
            lines = file_content.strip().split("\n")

            if not lines:
                return Response({"error": "Empty file"}, status=status.HTTP_400_BAD_REQUEST)

            # Parse headers and data
            headers = lines[0].split("\t")
            data_rows = [line.split("\t") for line in lines[1:]]

            # Check for pooled sample column and identify pool data (original CUPCAKE logic)
            pooled_column_index = None
            pooled_rows = []
            sn_rows = []

            for i, header in enumerate(headers):
                header_lower = header.lower()
                if "pooled sample" in header_lower or "pooled_sample" in header_lower:
                    pooled_column_index = i
                    break

            # If we found a pooled sample column, process the data
            if pooled_column_index is not None:
                for row_index, row in enumerate(data_rows):
                    if pooled_column_index < len(row):
                        pooled_value = row[pooled_column_index].strip()
                        if pooled_value.startswith("SN="):
                            sn_rows.append(row_index)
                        elif pooled_value.lower() == "pooled":
                            pooled_rows.append(row_index)

            # Get target metadata table
            try:
                metadata_table = MetadataTable.objects.get(id=data["metadata_table_id"])
            except MetadataTable.DoesNotExist:
                return Response(
                    {"error": "Invalid metadata_table_id"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Check permissions
            if not metadata_table.can_edit(request.user):
                return Response(
                    {"error": "Permission denied: cannot edit this metadata table"},
                    status=status.HTTP_403_FORBIDDEN,
                )

            # Clear existing columns if replace_existing is True
            if data.get("replace_existing"):
                metadata_table.columns.all().delete()
                metadata_table.sample_pools.all().delete()

            # Remove SN= rows from data before general processing (original CUPCAKE logic)
            sn_data = []
            if pooled_column_index is not None and sn_rows:
                # Store SN= rows for later pool creation
                for row_index in sorted(sn_rows, reverse=True):  # Reverse order to maintain indices
                    if row_index < len(data_rows):
                        sn_data.append(data_rows[row_index])
                        del data_rows[row_index]

            # Update sample count and extend/truncate data if needed (original CUPCAKE logic)
            expected_sample_count = metadata_table.sample_count or len(data_rows)
            if len(data_rows) != expected_sample_count:
                # Extend the number of samples to match expected count and fill with empty strings
                if len(data_rows) < expected_sample_count:
                    data_rows.extend(
                        [["" for i in range(len(headers))] for j in range(expected_sample_count - len(data_rows))]
                    )
                else:
                    data_rows = data_rows[:expected_sample_count]

                    # If we removed SN= rows and truncated data, add SN= rows back for pool creation
                    if pooled_column_index is not None and sn_rows and sn_data:
                        # Add SN= rows back to the end of the truncated data
                        data_rows.extend(sn_data)
                        # Update sn_rows indices to reflect their new positions at the end
                        sn_rows = list(range(expected_sample_count, len(data_rows)))

            # Update metadata table sample count
            metadata_table.sample_count = len(data_rows)
            metadata_table.save(update_fields=["sample_count"])

            # Enhanced SDRF import with intelligent column matching and schema organization
            created_columns = []
            column_name_usage = {}  # Track how many times each name is used

            # Get the metadata table template that existing columns come from
            table_template = self._get_table_template(metadata_table)

            # Process headers with intelligent template matching
            for i, header in enumerate(headers):
                # Parse header format: type[name] (original CUPCAKE format)
                header_lower = header.lower()
                if "[" in header_lower and "]" in header_lower:
                    metadata_type = header_lower.split("[")[0].strip()
                    name = header_lower.strip()
                else:
                    metadata_type = "special"
                    name = header_lower.strip()
                name = name.replace("_", " ")
                # Count usage of this column name
                if name not in column_name_usage:
                    column_name_usage[name] = 0
                column_name_usage[name] += 1

                # Try to find matching column in existing table or template
                metadata_column = self._find_or_create_matching_column(
                    name, metadata_type, metadata_table, table_template, i, column_name_usage[name]
                )

                created_columns.append(metadata_column)

            # retrieve metadata columns after reordering
            for i, metadata_column in enumerate(created_columns):
                metadata_value_map = {}

                # Process each data row for this column
                for j, row in enumerate(data_rows):
                    if i < len(row) and row[i]:
                        cell_value = row[i].strip()

                        if cell_value == "":
                            continue
                        if cell_value == "not applicable":
                            metadata_column.not_applicable = True
                            continue

                        value = metadata_column.convert_sdrf_to_metadata(cell_value)

                        if value not in metadata_value_map:
                            metadata_value_map[value] = []
                        metadata_value_map[value].append(j)

                # Set the most common value as default
                max_count = 0
                max_value = None
                for value in metadata_value_map:
                    if len(metadata_value_map[value]) > max_count:
                        max_count = len(metadata_value_map[value])
                        max_value = value

                if max_value:
                    metadata_column.value = max_value

                # Calculate modifiers for other values (exact original CUPCAKE logic)
                modifiers = []
                for value in metadata_value_map:
                    if value != max_value:
                        modifier = {"samples": [], "value": value}
                        # Sort from lowest to highest. Add samples index. For continuous samples, add range
                        samples = metadata_value_map[value]
                        samples.sort()
                        start = samples[0]
                        end = samples[0]
                        for i2 in range(1, len(samples)):
                            if samples[i2] == end + 1:
                                end = samples[i2]
                            else:
                                if start == end:
                                    modifier["samples"].append(str(start + 1))
                                else:
                                    modifier["samples"].append(f"{start + 1}-{end + 1}")
                                start = samples[i2]
                                end = samples[i2]
                        if start == end:
                            modifier["samples"].append(str(start + 1))
                        else:
                            modifier["samples"].append(f"{start + 1}-{end + 1}")
                        if len(modifier["samples"]) == 1:
                            modifier["samples"] = modifier["samples"][0]
                        else:
                            modifier["samples"] = ",".join(modifier["samples"])
                        modifiers.append(modifier)

                if modifiers:
                    # Store modifiers as native JSON object
                    metadata_column.modifiers = modifiers

                metadata_column.save()

            created_pools = []
            if data.get("create_pools") and pooled_column_index is not None:
                # Find source name column
                source_name_column_index = None
                for idx, header in enumerate(headers):
                    header_lower = header.lower()
                    if "source name" in header_lower or "source_name" in header_lower:
                        source_name_column_index = idx
                        break

                # Pool synchronization: track pools from import data (exact original CUPCAKE logic)
                import_pools_data = []

                if sn_rows and sn_data:
                    # Case 1: There are rows with SN= values - create pools from them
                    for pool_index, row in enumerate(sn_data):
                        if pooled_column_index < len(row):
                            sdrf_value = row[pooled_column_index].strip()

                            # Extract source names from SN= value
                            if sdrf_value.startswith("SN="):
                                source_names = sdrf_value[3:].split(",")
                                source_names = [name.strip() for name in source_names]

                                # Get pool name from source name column or use default
                                pool_name = (
                                    row[source_name_column_index]
                                    if source_name_column_index is not None and source_name_column_index < len(row)
                                    else f"Pool {pool_index + 1}"
                                )

                                # Find sample indices that match these source names
                                pooled_only_samples = []
                                pooled_and_independent_samples = []

                                for sample_index, sample_row in enumerate(data_rows):
                                    if source_name_column_index is not None and source_name_column_index < len(
                                        sample_row
                                    ):
                                        sample_source_name = sample_row[source_name_column_index].strip()
                                        if sample_source_name in source_names:
                                            # Check if this sample is also marked as "not pooled" or independent
                                            sample_pooled_value = ""
                                            if pooled_column_index < len(sample_row):
                                                sample_pooled_value = sample_row[pooled_column_index].strip().lower()

                                            if (
                                                sample_pooled_value == "not pooled"
                                                or sample_pooled_value == ""
                                                or sample_pooled_value == "independent"
                                            ):
                                                # Sample exists both in pool and as independent
                                                pooled_and_independent_samples.append(sample_index + 1)
                                            else:
                                                # Sample is only in pool
                                                pooled_only_samples.append(sample_index + 1)

                                # Store pool data for synchronization
                                import_pools_data.append(
                                    {
                                        "pool_name": pool_name,
                                        "pooled_only_samples": pooled_only_samples,
                                        "pooled_and_independent_samples": pooled_and_independent_samples,
                                        "is_reference": True,  # SN= pools are reference pools
                                        "metadata_row": row,
                                        "sdrf_value": sdrf_value,
                                    }
                                )

                elif pooled_rows:
                    # Case 2: No SN= rows but there are "pooled" rows - create a pool from them
                    # Get source names of all pooled samples
                    pooled_source_names = []
                    pooled_only_samples = []

                    for row_index in pooled_rows:
                        if (
                            source_name_column_index is not None
                            and row_index < len(data_rows)
                            and source_name_column_index < len(data_rows[row_index])
                        ):
                            source_name = data_rows[row_index][source_name_column_index].strip()
                            if source_name:
                                pooled_source_names.append(source_name)
                                pooled_only_samples.append(row_index + 1)

                    if pooled_source_names:
                        # Create SN= value from source names
                        sdrf_value = "SN=" + ",".join(pooled_source_names)
                        pool_name = "Pool 1"
                        template_row = data_rows[pooled_rows[0]]

                        # Store pool data for synchronization
                        import_pools_data.append(
                            {
                                "pool_name": pool_name,
                                "pooled_only_samples": pooled_only_samples,
                                "pooled_and_independent_samples": [],
                                "is_reference": False,  # Pooled rows are not reference pools by default
                                "metadata_row": template_row,
                                "sdrf_value": sdrf_value,
                            }
                        )

                # Synchronize pools with sophisticated logic (matching original CUPCAKE)
                if import_pools_data:
                    synchronize_pools_with_import_data(metadata_table, import_pools_data, created_columns, request.user)
                    # Get the updated pools list for response
                    created_pools = list(metadata_table.sample_pools.all())

            # Perform ontology validation on imported data
            validation_results = validate_sdrf_data_against_ontologies(headers, data_rows, created_columns)

            # Reorganize columns based on schema order from template
            if table_template and hasattr(table_template, "user_columns"):
                # Collect schemas from the template's columns
                schema_ids = set()
                for template_column in table_template.user_columns.all():
                    if (
                        hasattr(template_column, "template")
                        and template_column.template
                        and hasattr(template_column.template, "schema")
                        and template_column.template.schema
                    ):
                        schema_ids.add(template_column.template.schema.id)

                if schema_ids:
                    try:
                        metadata_table.reorder_columns_by_schema(schema_ids=list(schema_ids))
                    except Exception as e:
                        # Log error but don't fail the import
                        print(f"Warning: Failed to reorder columns by schema: {e}")
                        # Fall back to normalizing positions
                        metadata_table.normalize_column_positions()
                else:
                    # No schemas found, just normalize positions
                    metadata_table.normalize_column_positions()

                # Apply the same reordering logic to sample pool columns
                if created_pools:
                    for pool in created_pools:
                        if pool.metadata_columns.exists():
                            if schema_ids:
                                try:
                                    pool.reorder_pool_columns_by_schema(schema_ids=list(schema_ids))
                                except Exception as e:
                                    # Log error but don't fail the import, fall back to basic reordering
                                    print(f"Warning: Failed to reorder pool '{pool.pool_name}' columns by schema: {e}")
                                    pool.basic_pool_column_reordering()
                            else:
                                # No schemas found, use basic section-based ordering for pools
                                pool.basic_pool_column_reordering()
            else:
                # No template available, just normalize positions
                metadata_table.normalize_column_positions()

                # Apply basic reordering to sample pool columns when no template available
                if created_pools:
                    for pool in created_pools:
                        if pool.metadata_columns.exists():
                            pool.basic_pool_column_reordering()

            # Include validation results in response
            response_data = {
                "message": "SDRF file imported successfully",
                "created_columns": len(created_columns),
                "created_pools": len(created_pools),
                "pools_detected": pooled_column_index is not None,
                "sn_rows_count": len(sn_rows),
                "pooled_rows_count": len(pooled_rows),
                "sample_rows": len(data_rows),
                "ontology_validation": {
                    "errors_count": len(validation_results["errors"]),
                    "warnings_count": len(validation_results["warnings"]),
                    "errors": validation_results["errors"][:10],  # Limit to first 10 errors
                    "suggestions_available": len(validation_results["suggestions"]) > 0,
                },
            }

            # Add warning if there are ontology validation errors
            if validation_results["errors"]:
                response_data[
                    "message"
                ] += f" (Warning: {len(validation_results['errors'])} ontology validation errors found)"

            return Response(response_data)

        except Exception as e:
            return Response(
                {"error": f"Import failed: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser])
    def import_excel_file(self, request):
        """Import metadata from Excel file (matching original CUPCAKE)."""
        serializer = MetadataImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data

        try:
            # Read Excel workbook
            file_content = data["file"].read()
            wb = load_workbook(io.BytesIO(file_content))

            # Get main worksheet
            if "main" not in wb.sheetnames:
                return Response(
                    {"error": "Excel file must contain a 'main' worksheet"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            main_ws = wb["main"]
            main_headers = [cell.value for cell in main_ws[1]]
            main_data = [list(row) for row in main_ws.iter_rows(min_row=2, values_only=True)]

            # Get hidden worksheet if exists
            hidden_headers = []
            hidden_data = []
            if "hidden" in wb.sheetnames:
                hidden_ws = wb["hidden"]
                if hidden_ws.max_row > 1:
                    hidden_headers = [cell.value for cell in hidden_ws[1]]
                    hidden_data = [list(row) for row in hidden_ws.iter_rows(min_row=2, values_only=True)]

            # Get ID metadata column mapping
            if "id_metadata_column_map" not in wb.sheetnames:
                return Response(
                    {"error": "Excel file must contain an 'id_metadata_column_map' worksheet"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            id_metadata_column_map_ws = wb["id_metadata_column_map"]
            id_metadata_column_map_list = [
                list(row) for row in id_metadata_column_map_ws.iter_rows(min_row=2, values_only=True)
            ]
            id_metadata_column_map = {}
            for row in id_metadata_column_map_list:
                if row[0] is not None:
                    id_metadata_column_map[int(row[0])] = {
                        "column": row[1],
                        "name": row[2],
                        "type": row[3],
                        "hidden": row[4],
                    }

            # Check for pool sheets (original CUPCAKE logic)
            # Pool data extraction - simplified in ccv
            # pool_main_headers, pool_main_data = [], []
            # pool_hidden_headers, pool_hidden_data = [], []
            pool_id_metadata_column_map = {}
            pool_object_map_data = []

            if "pool_main" in wb.sheetnames:
                pool_main_ws = wb["pool_main"]
                if pool_main_ws.max_row > 1:
                    # pool_main_headers = [cell.value for cell in pool_main_ws[1]]
                    # pool_main_data = [
                    #     list(row)
                    #     for row in pool_main_ws.iter_rows(min_row=2, values_only=True)
                    # ]
                    pass

            if "pool_hidden" in wb.sheetnames:
                pool_hidden_ws = wb["pool_hidden"]
                if pool_hidden_ws.max_row > 1:
                    # pool_hidden_headers = [cell.value for cell in pool_hidden_ws[1]]
                    # pool_hidden_data = [
                    #     list(row)
                    #     for row in pool_hidden_ws.iter_rows(min_row=2, values_only=True)
                    # ]
                    pass

            if "pool_id_metadata_column_map" in wb.sheetnames:
                pool_id_metadata_column_map_ws = wb["pool_id_metadata_column_map"]
                pool_id_metadata_column_map_list = [
                    list(row) for row in pool_id_metadata_column_map_ws.iter_rows(min_row=2, values_only=True)
                ]
                for row in pool_id_metadata_column_map_list:
                    if row[0] is not None:
                        pool_id_metadata_column_map[int(row[0])] = {
                            "column": row[1],
                            "name": row[2],
                            "type": row[3],
                            "hidden": row[4],
                        }

            if "pool_object_map" in wb.sheetnames:
                pool_object_map_ws = wb["pool_object_map"]
                if pool_object_map_ws.max_row > 1:
                    pool_object_map_data = [
                        list(row) for row in pool_object_map_ws.iter_rows(min_row=2, values_only=True)
                    ]

            # Get target metadata table
            try:
                metadata_table = MetadataTable.objects.get(id=data["metadata_table_id"])
            except MetadataTable.DoesNotExist:
                return Response(
                    {"error": "Invalid metadata_table_id"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Check permissions
            if not metadata_table.can_edit(request.user):
                return Response(
                    {"error": "Permission denied: cannot edit this metadata table"},
                    status=status.HTTP_403_FORBIDDEN,
                )

            # Clear existing columns if replace_existing is True
            if data.get("replace_existing"):
                metadata_table.columns.all().delete()
                metadata_table.sample_pools.all().delete()

            # Combine main and hidden data (original CUPCAKE logic)
            if hidden_data:
                headers = main_headers + hidden_headers
                combined_data = [main_row + hidden_row for main_row, hidden_row in zip(main_data, hidden_data)]
            else:
                headers = main_headers
                combined_data = main_data

            # Update sample count and extend/truncate data if needed (original CUPCAKE logic)
            expected_sample_count = metadata_table.sample_count or len(combined_data)
            if len(combined_data) != expected_sample_count:
                if len(combined_data) < expected_sample_count:
                    combined_data.extend(
                        [["" for i in range(len(headers))] for j in range(expected_sample_count - len(combined_data))]
                    )
                else:
                    combined_data = combined_data[:expected_sample_count]

            # Update metadata table sample count
            metadata_table.sample_count = len(combined_data)
            metadata_table.save(update_fields=["sample_count"])

            # Create metadata columns from main headers (original CUPCAKE logic)
            created_columns = []
            for n, header in enumerate(main_headers):
                # Find metadata column ID from mapping
                id_from_map = 0
                for map_id, map_data in id_metadata_column_map.items():
                    if map_data["column"] == n and not map_data["hidden"]:
                        id_from_map = map_id
                        break

                # Try to get existing column or create new one
                if id_from_map > 0:
                    try:
                        metadata_column = MetadataColumn.objects.get(id=id_from_map)
                    except MetadataColumn.DoesNotExist:
                        metadata_column = self._create_column_from_header(header, False, metadata_table, n)
                else:
                    metadata_column = self._create_column_from_header(header, False, metadata_table, n)

                # Apply automatic ontology mapping
                apply_ontology_mapping_to_column(metadata_column)
                created_columns.append(metadata_column)

            # Create metadata columns from hidden headers (original CUPCAKE logic)
            for n, header in enumerate(hidden_headers):
                # Find metadata column ID from mapping
                id_from_map = 0
                for map_id, map_data in id_metadata_column_map.items():
                    if map_data["column"] == n and map_data["hidden"]:
                        id_from_map = map_id
                        break

                # Try to get existing column or create new one
                if id_from_map > 0:
                    try:
                        metadata_column = MetadataColumn.objects.get(id=id_from_map)
                    except MetadataColumn.DoesNotExist:
                        metadata_column = self._create_column_from_header(
                            header, True, metadata_table, len(main_headers) + n
                        )
                else:
                    metadata_column = self._create_column_from_header(
                        header, True, metadata_table, len(main_headers) + n
                    )

                # Apply automatic ontology mapping
                apply_ontology_mapping_to_column(metadata_column)
                created_columns.append(metadata_column)

            # Process data to populate column values and modifiers (exact original CUPCAKE logic)
            for i, metadata_column in enumerate(created_columns):
                metadata_value_map = {}

                # Process each data row for this column
                for j, row in enumerate(combined_data):
                    if i < len(row) and row[i]:
                        cell_value = str(row[i]).strip()

                        if cell_value == "":
                            continue
                        if cell_value == "not applicable":
                            metadata_column.not_applicable = True
                            continue

                        # Convert value using original CUPCAKE conversion
                        value = convert_sdrf_to_metadata(metadata_column.name.lower(), cell_value)

                        if value not in metadata_value_map:
                            metadata_value_map[value] = []
                        metadata_value_map[value].append(j)

                # Set the most common value as default
                max_count = 0
                max_value = None
                for value in metadata_value_map:
                    if len(metadata_value_map[value]) > max_count:
                        max_count = len(metadata_value_map[value])
                        max_value = value

                if max_value:
                    metadata_column.value = max_value

                # Calculate modifiers for other values (exact original CUPCAKE logic)
                modifiers = []
                for value in metadata_value_map:
                    if value != max_value:
                        modifier = {"samples": [], "value": value}
                        # Sort from lowest to highest. Add samples index. For continuous samples, add range
                        samples = metadata_value_map[value]
                        samples.sort()
                        start = samples[0]
                        end = samples[0]
                        for i2 in range(1, len(samples)):
                            if samples[i2] == end + 1:
                                end = samples[i2]
                            else:
                                if start == end:
                                    modifier["samples"].append(str(start + 1))
                                else:
                                    modifier["samples"].append(f"{start + 1}-{end + 1}")
                                start = samples[i2]
                                end = samples[i2]
                        if start == end:
                            modifier["samples"].append(str(start + 1))
                        else:
                            modifier["samples"].append(f"{start + 1}-{end + 1}")
                        if len(modifier["samples"]) == 1:
                            modifier["samples"] = modifier["samples"][0]
                        else:
                            modifier["samples"] = ",".join(modifier["samples"])
                        modifiers.append(modifier)

                if modifiers:
                    # Store modifiers as native JSON object
                    metadata_column.modifiers = modifiers

                metadata_column.save()

            # Handle pool data if present (original CUPCAKE approach)
            created_pools = []
            if pool_object_map_data and data.get("create_pools"):
                for pool_row in pool_object_map_data:
                    pool_name = pool_row[0] if pool_row[0] else f"Pool {len(created_pools) + 1}"
                    pooled_only_samples = json.loads(pool_row[1]) if pool_row[1] else []
                    pooled_and_independent_samples = json.loads(pool_row[2]) if pool_row[2] else []
                    is_reference = pool_row[3] if len(pool_row) > 3 else False
                    # sdrf_value = pool_row[4] if len(pool_row) > 4 else ""  # Not used in ccv

                    # Create pool
                    sample_pool = SamplePool.objects.create(
                        metadata_table=metadata_table,
                        pool_name=pool_name,
                        pooled_only_samples=pooled_only_samples,
                        pooled_and_independent_samples=pooled_and_independent_samples,
                        is_reference=is_reference,
                        created_by=request.user,
                    )
                    created_pools.append(sample_pool)

            # Perform ontology validation on imported data
            validation_results = validate_sdrf_data_against_ontologies(headers, combined_data, created_columns)

            # Include validation results in response
            response_data = {
                "message": "Excel file imported successfully",
                "created_columns": len(created_columns),
                "created_pools": len(created_pools),
                "sample_rows": len(combined_data),
                "has_hidden_data": len(hidden_data) > 0,
                "has_pool_data": len(pool_object_map_data) > 0,
                "ontology_validation": {
                    "errors_count": len(validation_results["errors"]),
                    "warnings_count": len(validation_results["warnings"]),
                    "errors": validation_results["errors"][:10],  # Limit to first 10 errors
                    "suggestions_available": len(validation_results["suggestions"]) > 0,
                },
            }

            # Add warning if there are ontology validation errors
            if validation_results["errors"]:
                response_data[
                    "message"
                ] += f" (Warning: {len(validation_results['errors'])} ontology validation errors found)"

            return Response(response_data)

        except Exception as e:
            return Response(
                {"error": f"Excel import failed: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    def _create_column_from_header(self, header, is_hidden, metadata_table, position):
        """Helper method to create MetadataColumn from header (original CUPCAKE logic)."""
        header = header.lower() if header else ""

        # Parse header format: type[name] (original CUPCAKE format)
        if "[" in header:
            metadata_type = header.split("[")[0]
            name = header.split("[")[1].replace("]", "")
        else:
            metadata_type = ""
            name = header

        # Create metadata column with proper naming (match original CUPCAKE)
        metadata_column = MetadataColumn.objects.create(
            name=name.capitalize().replace("Ms1", "MS1").replace("Ms2", "MS2"),
            type=metadata_type.capitalize(),
            column_position=position,
            metadata_table=metadata_table,
            hidden=is_hidden,
            readonly=False,
        )

        return metadata_column

    @action(detail=False, methods=["get"])
    def collect_metadata(self, request):
        """Collect metadata from metadata table."""
        serializer = MetadataCollectionSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data

        try:
            metadata_table = MetadataTable.objects.get(id=data["metadata_table_id"])
        except MetadataTable.DoesNotExist:
            return Response(
                {"error": "Invalid metadata_table_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Get metadata columns for the specified table
        metadata_columns = MetadataColumn.objects.filter(metadata_table=metadata_table)

        # Filter by metadata names if specified
        if data.get("metadata_names"):
            metadata_columns = metadata_columns.filter(name__in=data["metadata_names"])

        if data.get("unique_values_only"):
            # Return unique values grouped by metadata name
            unique_metadata = {}
            for column in metadata_columns:
                name = column.name.lower()
                if name not in unique_metadata:
                    unique_metadata[name] = {
                        "name": column.name,
                        "types": {},
                        "unique_values": set(),
                        "value_count": 0,
                    }

                if column.type not in unique_metadata[name]["types"]:
                    unique_metadata[name]["types"][column.type] = 0
                unique_metadata[name]["types"][column.type] += 1

                if column.value:
                    unique_metadata[name]["unique_values"].add(column.value)
                    unique_metadata[name]["value_count"] += 1

            # Convert sets to lists for JSON serialization
            for name_data in unique_metadata.values():
                name_data["unique_values"] = sorted(list(name_data["unique_values"]))

            return Response(
                {
                    "metadata_table_id": data["metadata_table_id"],
                    "metadata_table_name": metadata_table.name,
                    "unique_values_only": True,
                    "metadata_columns": unique_metadata,
                    "total_unique_values": sum(len(md["unique_values"]) for md in unique_metadata.values()),
                }
            )
        else:
            # Return full metadata column details
            serializer = MetadataColumnSerializer(metadata_columns, many=True)
            return Response(
                {
                    "metadata_table_id": data["metadata_table_id"],
                    "metadata_table_name": metadata_table.name,
                    "unique_values_only": False,
                    "metadata_columns": serializer.data,
                    "total_columns": metadata_columns.count(),
                }
            )

    @action(detail=False, methods=["post"])
    def export_sdrf_file(self, request):
        """Export metadata as SDRF file (matching original CUPCAKE)."""
        serializer = MetadataExportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data

        # Get metadata table
        try:
            metadata_table = MetadataTable.objects.get(id=data["metadata_table_id"])
        except (MetadataTable.DoesNotExist, KeyError):
            return Response(
                {"error": "metadata_table_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check permissions
        if not metadata_table.can_view(request.user):
            return Response(
                {"error": "Permission denied: cannot view this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Get metadata columns (user can specify specific columns or get all)
        if data.get("metadata_column_ids"):
            metadata_columns = metadata_table.columns.filter(id__in=data["metadata_column_ids"])
        else:
            metadata_columns = metadata_table.columns.all()

        # Filter out hidden columns for SDRF export (SDRF standard doesn't support hidden columns)
        visible_metadata = metadata_columns.filter(hidden=False)

        # Sort metadata and get structured output for SDRF format
        result_data, _ = sort_metadata(list(visible_metadata), metadata_table.sample_count)

        # Add pool data if requested and pools exist (original CUPCAKE logic)
        if data.get("include_pools", True):
            pools = list(metadata_table.sample_pools.all())
            if pools:
                # Add SN= rows for reference pools (original CUPCAKE approach)
                for pool in pools:
                    if pool.is_reference and pool.sdrf_value:
                        # Create a pool row following original CUPCAKE format
                        pool_row = ["" for _ in result_data[0]]  # Initialize with empty values

                        # Find the pooled sample column index
                        pooled_column_index = None
                        for i, header in enumerate(result_data[0]):
                            header_lower = header.lower()
                            if "pooled sample" in header_lower or "pooled_sample" in header_lower:
                                pooled_column_index = i
                                break

                        # Find the source name column index
                        source_name_column_index = None
                        for i, header in enumerate(result_data[0]):
                            header_lower = header.lower()
                            if "source name" in header_lower or "source_name" in header_lower:
                                source_name_column_index = i
                                break

                        # Set pool data in the row
                        if pooled_column_index is not None:
                            pool_row[pooled_column_index] = pool.sdrf_value

                        if source_name_column_index is not None:
                            pool_row[source_name_column_index] = pool.pool_name

                        # Fill other columns with pool-specific default values
                        for i, metadata_column in enumerate(visible_metadata):
                            if i < len(pool_row) and not pool_row[i]:  # Only fill empty cells
                                # Use column default value or "not applicable" for required fields
                                if metadata_column.value:
                                    pool_row[i] = metadata_column.value
                                elif metadata_column.name.lower() in [
                                    "organism",
                                    "disease",
                                    "organism part",
                                    "tissue",
                                ]:
                                    pool_row[i] = "not applicable"
                                else:
                                    pool_row[i] = "not available"

                        # Add pool row to results
                        result_data.append(pool_row)

        # Convert to tab-separated format (SDRF standard)
        sdrf_content = []
        for row in result_data:
            # Convert all values to strings and handle None values
            str_row = [str(cell) if cell is not None else "" for cell in row]
            sdrf_content.append("\t".join(str_row))

        sdrf_text = "\n".join(sdrf_content)

        # Validate SDRF format if validation was requested
        validation_results = None
        if data.get("validate_sdrf", False):
            try:
                validation_results = validate_sdrf(result_data)
            except Exception as e:
                validation_results = {
                    "errors": [f"SDRF validation failed: {str(e)}"],
                    "warnings": [],
                    "suggestions": [],
                }

        # Return as downloadable file
        response = HttpResponse(sdrf_text, content_type="text/tab-separated-values")

        # Create filename based on metadata table name
        safe_filename = "".join(c for c in metadata_table.name if c.isalnum() or c in (" ", "-", "_")).rstrip()
        safe_filename = safe_filename.replace(" ", "_")
        response["Content-Disposition"] = f'attachment; filename="{safe_filename}_metadata.sdrf"'

        # Add validation results as headers if available
        if validation_results:
            response["X-SDRF-Validation-Errors"] = str(len(validation_results.get("errors", [])))
            response["X-SDRF-Validation-Warnings"] = str(len(validation_results.get("warnings", [])))
            if validation_results.get("errors"):
                # Include first few errors in header (truncated for header length limits)
                error_summary = "; ".join(validation_results["errors"][:3])
                if len(error_summary) > 200:
                    error_summary = error_summary[:197] + "..."
                response["X-SDRF-Validation-Error-Summary"] = error_summary

        return response


# ===================================================================
# ONTOLOGY AND CONTROLLED VOCABULARY VIEWSETS
# ===================================================================


class SpeciesViewSet(FilterMixin, viewsets.ReadOnlyModelViewSet):
    """ViewSet for Species controlled vocabulary (read-only)."""

    queryset = Species.objects.all()
    serializer_class = SpeciesSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ["official_name", "common_name", "synonym"]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Filter by species code
        code = self.request.query_params.get("code")
        if code:
            queryset = queryset.filter(code__iexact=code)

        # Filter by taxon
        taxon = self.request.query_params.get("taxon")
        if taxon:
            queryset = queryset.filter(taxon=taxon)

        return queryset.order_by("official_name")


class TissueViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for Tissue controlled vocabulary (read-only)."""

    queryset = Tissue.objects.all()
    serializer_class = TissueSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Search by identifier or accession
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                models.Q(identifier__icontains=search)
                | models.Q(accession__icontains=search)
                | models.Q(synonyms__icontains=search)
            )

        return queryset.order_by("identifier")


class HumanDiseaseViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for HumanDisease controlled vocabulary (read-only)."""

    queryset = HumanDisease.objects.all()
    serializer_class = HumanDiseaseSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Search by identifier, accession, or definition
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                models.Q(identifier__icontains=search)
                | models.Q(accession__icontains=search)
                | models.Q(definition__icontains=search)
                | models.Q(synonyms__icontains=search)
                | models.Q(acronym__icontains=search)
            )

        return queryset.order_by("identifier")


class SubcellularLocationViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for SubcellularLocation controlled vocabulary (read-only)."""

    queryset = SubcellularLocation.objects.all()
    serializer_class = SubcellularLocationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Search by various fields
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                models.Q(accession__icontains=search)
                | models.Q(location_identifier__icontains=search)
                | models.Q(definition__icontains=search)
                | models.Q(synonyms__icontains=search)
            )

        return queryset.order_by("accession")


class MSUniqueVocabulariesViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for MSUniqueVocabularies controlled vocabulary (read-only)."""

    queryset = MSUniqueVocabularies.objects.all()
    serializer_class = MSUniqueVocabulariesSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Filter by term type
        term_type = self.request.query_params.get("term_type")
        if term_type:
            queryset = queryset.filter(term_type__iexact=term_type)

        # Search by accession, name, or definition
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                models.Q(accession__icontains=search)
                | models.Q(name__icontains=search)
                | models.Q(definition__icontains=search)
            )

        return queryset.order_by("accession")


class UnimodViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for Unimod controlled vocabulary (read-only)."""

    queryset = Unimod.objects.all()
    serializer_class = UnimodSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Search by accession, name, or definition
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                models.Q(accession__icontains=search)
                | models.Q(name__icontains=search)
                | models.Q(definition__icontains=search)
            )

        return queryset.order_by("accession")


# ===================================================================
# METADATA COLUMN TEMPLATE VIEWSETS
# ===================================================================


class MetadataColumnTemplateViewSet(FilterMixin, viewsets.ModelViewSet):
    """ViewSet for managing MetadataColumnTemplate objects."""

    queryset = MetadataColumnTemplate.objects.all()
    serializer_class = MetadataColumnTemplateSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ["name", "description", "column_name", "tags"]
    filterset_fields = ["visibility", "owner", "lab_group", "is_active"]

    def get_queryset(self):
        """Filter queryset based on visibility and user permissions."""
        user = self.request.user
        queryset = super().get_queryset()

        # Filter by visibility and permissions
        queryset = queryset.filter(
            models.Q(visibility="global")
            | models.Q(visibility="public")
            | models.Q(visibility="private", owner=user)
            | models.Q(visibility="lab_group", lab_group__members=user)
            | models.Q(shared_with_users=user)
        ).distinct()

        # Apply additional filters
        visibility = self.request.query_params.get("visibility")
        if visibility:
            queryset = queryset.filter(visibility=visibility)

        owner_id = self.request.query_params.get("owner_id")
        if owner_id:
            queryset = queryset.filter(owner_id=owner_id)

        lab_group_id = self.request.query_params.get("lab_group_id")
        if lab_group_id:
            queryset = queryset.filter(lab_group_id=lab_group_id)

        ontology_type = self.request.query_params.get("ontology_type")
        if ontology_type:
            queryset = queryset.filter(ontology_type=ontology_type)

        category = self.request.query_params.get("category")
        if category:
            queryset = queryset.filter(category__icontains=category)

        is_active = self.request.query_params.get("is_active")
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == "true")

        return queryset.order_by("-usage_count", "name")

    def perform_create(self, serializer):
        """Ensure owner is set when creating templates."""
        serializer.save(owner=self.request.user)

    def perform_update(self, serializer):
        """Check permissions before updating."""
        template = self.get_object()
        if not template.can_edit(self.request.user):
            raise PermissionDenied("You don't have permission to edit this template")
        serializer.save()

    def perform_destroy(self, instance):
        """Check permissions before deleting."""
        if not instance.can_delete(self.request.user):
            raise PermissionDenied("You don't have permission to delete this template")
        super().perform_destroy(instance)

    @action(detail=True, methods=["post"])
    def create_metadata_column(self, request, pk=None):
        """Create a MetadataColumn from this template."""
        template = self.get_object()

        # Check if user can view this template
        if not template.can_view(request.user):
            return Response(
                {"error": "You don't have permission to use this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        metadata_table_id = request.data.get("metadata_table_id")
        position = request.data.get("position")

        if not metadata_table_id:
            return Response(
                {"error": "metadata_table_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            metadata_table = MetadataTable.objects.get(id=metadata_table_id)
        except MetadataTable.DoesNotExist:
            return Response(
                {"error": "Invalid metadata_table_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check if user can edit the metadata table
        if not metadata_table.can_edit(request.user):
            return Response(
                {"error": "You don't have permission to edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Create the metadata column
        try:
            column = template.create_metadata_column(metadata_table, position)

            return Response(
                {
                    "message": "Metadata column created successfully from template",
                    "column_id": column.id,
                    "template_usage_count": template.usage_count,
                },
                status=status.HTTP_201_CREATED,
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to create column: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def share_template(self, request, pk=None):
        """Share this template with another user."""
        template = self.get_object()

        # Check if user can edit this template (only creators/staff can share)
        if not template.can_edit(request.user):
            return Response(
                {"error": "You don't have permission to share this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        user_id = request.data.get("user_id")
        permission_level = request.data.get("permission_level", "use")

        if not user_id:
            return Response(
                {"error": "user_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            target_user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response(
                {"error": "Invalid user_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Create or update share
        share, created = MetadataColumnTemplateShare.objects.update_or_create(
            template=template,
            user=target_user,
            defaults={
                "shared_by": request.user,
                "permission_level": permission_level,
            },
        )

        action = "created" if created else "updated"
        return Response(
            {
                "message": f"Template share {action} successfully",
                "share_id": share.id,
                "user": target_user.username,
                "permission_level": permission_level,
            }
        )

    @action(detail=True, methods=["delete"])
    def unshare_template(self, request, pk=None):
        """Remove template sharing for a user."""
        template = self.get_object()

        # Check if user can edit this template
        if not template.can_edit(request.user):
            return Response(
                {"error": "You don't have permission to manage sharing for this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        user_id = request.data.get("user_id")
        if not user_id:
            return Response(
                {"error": "user_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            share = MetadataColumnTemplateShare.objects.get(template=template, user_id=user_id)
            share.delete()
            return Response({"message": "Template sharing removed successfully"})
        except MetadataColumnTemplateShare.DoesNotExist:
            return Response(
                {"error": "Template share not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

    @action(detail=False, methods=["get"])
    def my_templates(self, request):
        """Get templates created by the current user."""
        templates = self.get_queryset().filter(owner=request.user)

        # Apply search if provided
        search = request.query_params.get("search")
        if search:
            templates = templates.filter(
                models.Q(name__icontains=search)
                | models.Q(description__icontains=search)
                | models.Q(column_name__icontains=search)
            )

        serializer = self.get_serializer(templates, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def popular_templates(self, request):
        """Get most popular public templates."""
        templates = (
            self.get_queryset()
            .filter(visibility__in=["public", "global"], is_active=True)
            .order_by("-usage_count")[:20]
        )

        serializer = self.get_serializer(templates, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def ontology_suggestions(self, request):
        """Get ontology suggestions for this column template."""
        # Import serializer locally to avoid circular import issues

        # Get template ID from query params
        template_id = request.query_params.get("template_id")
        if not template_id:
            return Response(
                {"error": "template_id parameter is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Get template directly without the viewset's filtered queryset
        try:
            template = MetadataColumnTemplate.objects.get(pk=template_id)
            # Check if user can view this template
            if not template.can_view(request.user):
                return Response(
                    {"error": "You don't have permission to access this template"},
                    status=status.HTTP_403_FORBIDDEN,
                )
        except MetadataColumnTemplate.DoesNotExist:
            return Response(
                {"error": "Template not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

        search_term = request.query_params.get("search", "")
        limit = min(int(request.query_params.get("limit", 20)), 100)  # Cap at 100
        search_type = request.query_params.get("search_type", "icontains")

        # Validate search_type
        valid_search_types = ["icontains", "istartswith", "exact"]
        if search_type not in valid_search_types:
            return Response(
                {"error": f"Invalid search_type. Must be one of: {', '.join(valid_search_types)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not template.ontology_type:
            return Response(
                {
                    "suggestions": [],
                    "has_more": False,
                    "total_count": 0,
                    "message": "No ontology type specified for this template",
                }
            )

        try:
            # Use the template's own method for getting suggestions
            suggestions = template.get_ontology_suggestions(search_term, limit, search_type)
            total_count = len(suggestions)

            # Serialize suggestions with enhanced data
            serializer = OntologySuggestionSerializer(
                suggestions, many=True, context={"ontology_type": template.ontology_type}
            )

            return Response(
                {
                    "suggestions": serializer.data,
                    "has_more": len(suggestions) == limit,  # Template method doesn't provide total count
                    "total_count": total_count,
                    "ontology_type": template.ontology_type,
                    "search_term": search_term,
                    "search_type": search_type,
                }
            )

        except Exception as e:
            return Response(
                {"error": f"Failed to fetch ontology suggestions: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class MetadataColumnTemplateShareViewSet(FilterMixin, viewsets.ModelViewSet):
    """ViewSet for managing MetadataColumnTemplateShare objects."""

    queryset = MetadataColumnTemplateShare.objects.all()
    serializer_class = MetadataColumnTemplateShareSerializer
    permission_classes = [IsAuthenticated]
    search_fields = ["template__name", "template__description"]
    filterset_fields = ["template", "shared_by", "shared_with", "permission_level"]

    def get_queryset(self):
        """Filter queryset based on user permissions."""
        user = self.request.user
        queryset = super().get_queryset()

        # Users can only see shares they created, received, or are staff
        if not user.is_staff:
            queryset = queryset.filter(models.Q(shared_by=user) | models.Q(user=user))

        # Filter by template
        template_id = self.request.query_params.get("template_id")
        if template_id:
            queryset = queryset.filter(template_id=template_id)

        # Filter by permission level
        permission_level = self.request.query_params.get("permission_level")
        if permission_level:
            queryset = queryset.filter(permission_level=permission_level)

        return queryset.order_by("-shared_at")

    def perform_create(self, serializer):
        """Ensure shared_by is set when creating shares."""
        template = serializer.validated_data["template"]

        # Check if user can edit the template
        if not template.can_edit(self.request.user):
            raise PermissionDenied("You don't have permission to share this template")

        serializer.save(shared_by=self.request.user)

    def perform_update(self, serializer):
        """Check permissions before updating shares."""
        share = self.get_object()
        if share.shared_by != self.request.user and not self.request.user.is_staff:
            raise PermissionDenied("You can only update shares you created")
        serializer.save()

    def perform_destroy(self, instance):
        """Check permissions before deleting shares."""
        if (
            instance.shared_by != self.request.user
            and instance.user != self.request.user
            and not self.request.user.is_staff
        ):
            raise PermissionDenied("You can only delete shares you created or received")
        super().perform_destroy(instance)


# ===================================================================
# COMPREHENSIVE ONTOLOGY VIEWSETS FOR SDRF VALIDATION
# ===================================================================


class MondoDiseaseViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for MONDO Disease Ontology."""

    queryset = MondoDisease.objects.filter(obsolete=False)
    serializer_class = MondoDiseaseSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["name", "identifier"]
    search_fields = ["name", "definition", "synonyms", "identifier"]
    ordering_fields = ["name", "identifier", "created_at"]
    ordering = ["name"]

    @action(detail=False, methods=["get"])
    def search_suggestions(self, request):
        """Get disease suggestions for SDRF validation."""
        query = request.GET.get("q", "").strip()
        limit = int(request.GET.get("limit", 20))

        if not query or len(query) < 2:
            return Response({"results": []})

        # Search by name and synonyms
        queryset = self.get_queryset().filter(models.Q(name__icontains=query) | models.Q(synonyms__icontains=query))[
            :limit
        ]

        suggestions = []
        for disease in queryset:
            suggestions.append(
                {
                    "identifier": disease.identifier,
                    "name": disease.name,
                    "definition": disease.definition or "",
                    "synonyms": [s.strip() for s in disease.synonyms.split(";") if s.strip()]
                    if disease.synonyms
                    else [],
                    "source": "mondo",
                    "match_type": "fuzzy",
                }
            )

        return Response({"results": suggestions})


class UberonAnatomyViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for UBERON Anatomy Ontology."""

    queryset = UberonAnatomy.objects.filter(obsolete=False)
    serializer_class = UberonAnatomySerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["name", "identifier"]
    search_fields = ["name", "definition", "synonyms", "identifier"]
    ordering_fields = ["name", "identifier", "created_at"]
    ordering = ["name"]

    @action(detail=False, methods=["get"])
    def search_suggestions(self, request):
        """Get anatomy/tissue suggestions for SDRF validation."""
        query = request.GET.get("q", "").strip()
        limit = int(request.GET.get("limit", 20))

        if not query or len(query) < 2:
            return Response({"results": []})

        # Search by name and synonyms
        queryset = self.get_queryset().filter(models.Q(name__icontains=query) | models.Q(synonyms__icontains=query))[
            :limit
        ]

        suggestions = []
        for anatomy in queryset:
            suggestions.append(
                {
                    "identifier": anatomy.identifier,
                    "name": anatomy.name,
                    "definition": anatomy.definition or "",
                    "synonyms": [s.strip() for s in anatomy.synonyms.split(";") if s.strip()]
                    if anatomy.synonyms
                    else [],
                    "source": "uberon",
                    "match_type": "fuzzy",
                }
            )

        return Response({"results": suggestions})


class NCBITaxonomyViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for NCBI Taxonomy."""

    queryset = NCBITaxonomy.objects.all()
    serializer_class = NCBITaxonomySerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["scientific_name", "common_name", "rank", "tax_id"]
    search_fields = ["scientific_name", "common_name", "synonyms"]
    ordering_fields = ["scientific_name", "tax_id", "rank"]
    ordering = ["scientific_name"]

    @action(detail=False, methods=["get"])
    def search_suggestions(self, request):
        """Get organism suggestions for SDRF validation."""
        query = request.GET.get("q", "").strip()
        limit = int(request.GET.get("limit", 20))

        if not query or len(query) < 2:
            return Response({"results": []})

        # Search by scientific name, common name, and synonyms
        queryset = self.get_queryset().filter(
            models.Q(scientific_name__icontains=query)
            | models.Q(common_name__icontains=query)
            | models.Q(synonyms__icontains=query)
        )[:limit]

        suggestions = []
        for taxonomy in queryset:
            suggestions.append(
                {
                    "identifier": str(taxonomy.tax_id),
                    "name": taxonomy.scientific_name,
                    "definition": f"Rank: {taxonomy.rank}"
                    + (f", Common: {taxonomy.common_name}" if taxonomy.common_name else ""),
                    "synonyms": [s.strip() for s in taxonomy.synonyms.split(";") if s.strip()]
                    if taxonomy.synonyms
                    else [],
                    "source": "ncbitaxon",
                    "match_type": "fuzzy",
                }
            )

        return Response({"results": suggestions})


class ChEBICompoundViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for ChEBI Chemical Compounds."""

    queryset = ChEBICompound.objects.filter(obsolete=False)
    serializer_class = ChEBICompoundSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["name", "identifier", "formula"]
    search_fields = ["name", "definition", "synonyms", "identifier", "formula"]
    ordering_fields = ["name", "identifier", "mass"]
    ordering = ["name"]

    @action(detail=False, methods=["get"])
    def search_suggestions(self, request):
        """Get chemical compound suggestions for SDRF validation."""
        query = request.GET.get("q", "").strip()
        limit = int(request.GET.get("limit", 20))

        if not query or len(query) < 2:
            return Response({"results": []})

        # Search by name, synonyms, and formula
        queryset = self.get_queryset().filter(
            models.Q(name__icontains=query) | models.Q(synonyms__icontains=query) | models.Q(formula__icontains=query)
        )[:limit]

        suggestions = []
        for compound in queryset:
            suggestions.append(
                {
                    "identifier": compound.identifier,
                    "name": compound.name,
                    "definition": compound.definition or "",
                    "synonyms": [s.strip() for s in compound.synonyms.split(";") if s.strip()]
                    if compound.synonyms
                    else [],
                    "source": "chebi",
                    "match_type": "fuzzy",
                }
            )

        return Response({"results": suggestions})


class PSIMSOntologyViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for PSI-MS Ontology."""

    queryset = PSIMSOntology.objects.filter(obsolete=False)
    serializer_class = PSIMSOntologySerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["name", "identifier", "category"]
    search_fields = ["name", "definition", "synonyms", "identifier"]
    ordering_fields = ["name", "identifier", "category"]
    ordering = ["name"]

    @action(detail=False, methods=["get"])
    def search_suggestions(self, request):
        """Get mass spectrometry term suggestions for SDRF validation."""
        query = request.GET.get("q", "").strip()
        category = request.GET.get("category", "").strip()
        limit = int(request.GET.get("limit", 20))

        if not query or len(query) < 2:
            return Response({"results": []})

        # Search by name and synonyms
        queryset = self.get_queryset().filter(models.Q(name__icontains=query) | models.Q(synonyms__icontains=query))

        # Filter by category if provided
        if category:
            queryset = queryset.filter(category=category)

        queryset = queryset[:limit]

        suggestions = []
        for term in queryset:
            suggestions.append(
                {
                    "identifier": term.identifier,
                    "name": term.name,
                    "definition": term.definition or "",
                    "synonyms": [s.strip() for s in term.synonyms.split(";") if s.strip()] if term.synonyms else [],
                    "source": "ms",
                    "match_type": "fuzzy",
                    "category": term.category,
                }
            )

        return Response({"results": suggestions})

    @action(detail=False, methods=["get"])
    def instruments(self, request):
        """Get instrument-specific terms."""
        query = request.GET.get("q", "").strip()
        limit = int(request.GET.get("limit", 20))

        queryset = self.get_queryset().filter(category="instrument")
        if query:
            queryset = queryset.filter(models.Q(name__icontains=query) | models.Q(synonyms__icontains=query))

        return Response(
            {
                "results": [
                    {
                        "identifier": term.identifier,
                        "name": term.name,
                        "definition": term.definition or "",
                        "source": "ms",
                        "match_type": "exact",
                    }
                    for term in queryset[:limit]
                ]
            }
        )


# ===================================================================
# UNIFIED ONTOLOGY SEARCH ENDPOINT FOR SDRF VALIDATION
# ===================================================================


class CellOntologyViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for Cell Ontology (cell types and cell lines)."""

    queryset = CellOntology.objects.filter(obsolete=False)
    serializer_class = CellOntologySerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["name", "identifier", "cell_line", "organism", "source"]
    search_fields = ["name", "synonyms", "identifier", "organism", "tissue_origin"]
    ordering_fields = ["name", "identifier", "organism"]
    ordering = ["name"]

    @action(detail=False, methods=["get"])
    def search_suggestions(self, request):
        """Get cell type/cell line suggestions for SDRF validation."""
        query = request.GET.get("q", "").strip()
        cell_type = request.GET.get("cell_type", "").strip()  # 'cell_line', 'primary', or 'all'
        organism = request.GET.get("organism", "").strip()
        limit = int(request.GET.get("limit", 20))

        if not query or len(query) < 2:
            return Response({"results": []})

        # Search by name and synonyms
        queryset = self.get_queryset().filter(models.Q(name__icontains=query) | models.Q(synonyms__icontains=query))

        # Filter by cell type if provided
        if cell_type == "cell_line":
            queryset = queryset.filter(cell_line=True)
        elif cell_type == "primary":
            queryset = queryset.filter(cell_line=False)

        # Filter by organism if provided
        if organism:
            queryset = queryset.filter(organism__icontains=organism)

        queryset = queryset[:limit]

        suggestions = []
        for cell in queryset:
            suggestions.append(
                {
                    "identifier": cell.identifier,
                    "name": cell.name,
                    "definition": cell.definition or "",
                    "synonyms": [s.strip() for s in cell.synonyms.split(";") if s.strip()] if cell.synonyms else [],
                    "source": cell.source,
                    "match_type": "fuzzy",
                    "cell_type": "Cell Line" if cell.cell_line else "Cell Type",
                    "organism": cell.organism or "",
                    "tissue_origin": cell.tissue_origin or "",
                }
            )

        return Response({"results": suggestions})

    @action(detail=False, methods=["get"])
    def cell_lines(self, request):
        """Get cell lines for proteomics experiments."""
        queryset = self.get_queryset().filter(cell_line=True)
        organism = request.GET.get("organism", "").strip()

        if organism:
            queryset = queryset.filter(organism__icontains=organism)

        queryset = queryset[:100]  # Limit to prevent large responses
        serializer = self.get_serializer(queryset, many=True)
        return Response({"results": serializer.data})

    @action(detail=False, methods=["get"])
    def primary_cells(self, request):
        """Get primary cell types for proteomics experiments."""
        queryset = self.get_queryset().filter(cell_line=False)
        organism = request.GET.get("organism", "").strip()

        if organism:
            queryset = queryset.filter(organism__icontains=organism)

        queryset = queryset[:100]  # Limit to prevent large responses
        serializer = self.get_serializer(queryset, many=True)
        return Response({"results": serializer.data})


class OntologySearchViewSet(viewsets.ViewSet):
    """Unified ontology search for SDRF validation."""

    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=["get"])
    def suggest(self, request):
        """Get ontology suggestions across all sources for SDRF validation."""
        query = request.GET.get("q", "").strip()
        ontology_type = request.GET.get("type", "").strip()
        match_type = request.GET.get("match", "contains").strip().lower()  # 'contains' or 'startswith'
        limit = int(request.GET.get("limit", 50))

        if not query or len(query) < 2:
            return Response({"results": []})

        # Validate match_type parameter
        if match_type not in ["contains", "startswith"]:
            return Response(
                {"error": "Invalid match type. Use 'contains' or 'startswith'", "provided": match_type}, status=400
            )

        all_suggestions = []

        # Search based on exact ontology source names
        if ontology_type == "ncbi_taxonomy":
            # NCBI Taxonomy
            if match_type == "startswith":
                ncbi_results = NCBITaxonomy.objects.filter(
                    models.Q(scientific_name__istartswith=query)
                    | models.Q(common_name__istartswith=query)
                    | models.Q(synonyms__istartswith=query)
                )[:limit]
            else:  # contains
                ncbi_results = NCBITaxonomy.objects.filter(
                    models.Q(scientific_name__icontains=query)
                    | models.Q(common_name__icontains=query)
                    | models.Q(synonyms__icontains=query)
                )[:limit]

            for taxonomy in ncbi_results:
                all_suggestions.append(
                    {
                        "identifier": str(taxonomy.tax_id),
                        "name": taxonomy.scientific_name,
                        "definition": f"Rank: {taxonomy.rank}"
                        + (f", Common: {taxonomy.common_name}" if taxonomy.common_name else ""),
                        "source": "ncbi_taxonomy",
                        "match_type": match_type,
                    }
                )

        elif ontology_type == "species":
            # Species database
            if match_type == "startswith":
                species_results = Species.objects.filter(
                    models.Q(official_name__istartswith=query)
                    | models.Q(common_name__istartswith=query)
                    | models.Q(synonym__istartswith=query)
                )[:limit]
            else:  # contains
                species_results = Species.objects.filter(
                    models.Q(official_name__icontains=query)
                    | models.Q(common_name__icontains=query)
                    | models.Q(synonym__icontains=query)
                )[:limit]

            for species in species_results:
                all_suggestions.append(
                    {
                        "identifier": species.code,
                        "name": species.official_name,
                        "definition": f"Common name: {species.common_name or 'N/A'}",
                        "source": "species",
                        "match_type": match_type,
                    }
                )

        elif ontology_type == "mondo":
            # MONDO Disease Ontology
            if match_type == "startswith":
                mondo_results = MondoDisease.objects.filter(
                    models.Q(name__istartswith=query) | models.Q(synonyms__istartswith=query), obsolete=False
                )[:limit]
            else:  # contains
                mondo_results = MondoDisease.objects.filter(
                    models.Q(name__icontains=query) | models.Q(synonyms__icontains=query), obsolete=False
                )[:limit]

            for disease in mondo_results:
                all_suggestions.append(
                    {
                        "identifier": disease.identifier,
                        "name": disease.name,
                        "definition": disease.definition or "",
                        "source": "mondo",
                        "match_type": match_type,
                    }
                )

        elif ontology_type == "human_disease":
            # Human Disease database
            if match_type == "startswith":
                human_disease_results = HumanDisease.objects.filter(
                    models.Q(accession__istartswith=query)
                    | models.Q(synonyms__istartswith=query)
                    | models.Q(definition__istartswith=query)
                )[:limit]
            else:  # contains
                human_disease_results = HumanDisease.objects.filter(
                    models.Q(accession__icontains=query)
                    | models.Q(synonyms__icontains=query)
                    | models.Q(definition__icontains=query)
                )[:limit]

            for disease in human_disease_results:
                all_suggestions.append(
                    {
                        "identifier": disease.identifier,
                        "name": disease.accession,
                        "definition": disease.definition or "",
                        "source": "human_disease",
                        "match_type": match_type,
                    }
                )

        elif ontology_type == "uberon":
            # UBERON Anatomy Ontology
            if match_type == "startswith":
                uberon_results = UberonAnatomy.objects.filter(
                    models.Q(name__istartswith=query) | models.Q(synonyms__istartswith=query), obsolete=False
                )[:limit]
            else:  # contains
                uberon_results = UberonAnatomy.objects.filter(
                    models.Q(name__icontains=query) | models.Q(synonyms__icontains=query), obsolete=False
                )[:limit]

            for anatomy in uberon_results:
                all_suggestions.append(
                    {
                        "identifier": anatomy.identifier,
                        "name": anatomy.name,
                        "definition": anatomy.definition or "",
                        "source": "uberon",
                        "match_type": match_type,
                    }
                )

        elif ontology_type == "tissue":
            # Tissue database
            if match_type == "startswith":
                tissue_results = Tissue.objects.filter(
                    models.Q(accession__istartswith=query) | models.Q(synonyms__istartswith=query)
                )[:limit]
            else:  # contains
                tissue_results = Tissue.objects.filter(
                    models.Q(accession__icontains=query) | models.Q(synonyms__icontains=query)
                )[:limit]

            for tissue in tissue_results:
                all_suggestions.append(
                    {
                        "identifier": tissue.identifier,
                        "name": tissue.accession,
                        "definition": tissue.synonyms or "",
                        "source": "tissue",
                        "match_type": match_type,
                    }
                )

        elif ontology_type == "subcellular_location":
            # Subcellular Location database
            if match_type == "startswith":
                subcellular_results = SubcellularLocation.objects.filter(
                    models.Q(definition__istartswith=query) | models.Q(synonyms__istartswith=query)
                )[:limit]
            else:  # contains
                subcellular_results = SubcellularLocation.objects.filter(
                    models.Q(definition__icontains=query) | models.Q(synonyms__icontains=query)
                )[:limit]

            for location in subcellular_results:
                all_suggestions.append(
                    {
                        "identifier": location.accession,
                        "name": location.definition,
                        "definition": location.synonyms or "",
                        "source": "subcellular_location",
                        "match_type": match_type,
                    }
                )

        elif ontology_type == "chebi":
            # ChEBI Chemical Compounds
            if match_type == "startswith":
                chebi_results = ChEBICompound.objects.filter(
                    models.Q(name__istartswith=query)
                    | models.Q(synonyms__istartswith=query)
                    | models.Q(formula__istartswith=query),
                    obsolete=False,
                )[:limit]
            else:  # contains
                chebi_results = ChEBICompound.objects.filter(
                    models.Q(name__icontains=query)
                    | models.Q(synonyms__icontains=query)
                    | models.Q(formula__icontains=query),
                    obsolete=False,
                )[:limit]

            for compound in chebi_results:
                all_suggestions.append(
                    {
                        "identifier": compound.identifier,
                        "name": compound.name,
                        "definition": compound.definition or "",
                        "source": "chebi",
                        "match_type": match_type,
                    }
                )

        elif ontology_type == "cell_ontology":
            # Cell Ontology (primary cells only)
            if match_type == "startswith":
                cell_results = CellOntology.objects.filter(
                    models.Q(name__istartswith=query) | models.Q(synonyms__istartswith=query),
                    cell_line=False,
                    obsolete=False,
                )[:limit]
            else:  # contains
                cell_results = CellOntology.objects.filter(
                    models.Q(name__icontains=query) | models.Q(synonyms__icontains=query),
                    cell_line=False,
                    obsolete=False,
                )[:limit]

            for cell in cell_results:
                all_suggestions.append(
                    {
                        "identifier": cell.identifier,
                        "name": cell.name,
                        "definition": cell.definition or "",
                        "source": "cell_ontology",
                        "match_type": match_type,
                    }
                )

        elif ontology_type == "cellosaurus":
            # Cellosaurus (cell lines only)
            if match_type == "startswith":
                cell_results = CellOntology.objects.filter(
                    models.Q(name__istartswith=query) | models.Q(synonyms__istartswith=query),
                    cell_line=True,
                    obsolete=False,
                )[:limit]
            else:  # contains
                cell_results = CellOntology.objects.filter(
                    models.Q(name__icontains=query) | models.Q(synonyms__icontains=query),
                    cell_line=True,
                    obsolete=False,
                )[:limit]

            for cell in cell_results:
                all_suggestions.append(
                    {
                        "identifier": cell.identifier,
                        "name": cell.name,
                        "definition": cell.definition or "",
                        "source": "cellosaurus",
                        "match_type": match_type,
                    }
                )

        elif ontology_type == "psi_ms":
            # PSI-MS Ontology
            if match_type == "startswith":
                psims_results = PSIMSOntology.objects.filter(
                    models.Q(name__istartswith=query) | models.Q(synonyms__istartswith=query), obsolete=False
                )[:limit]
            else:  # contains
                psims_results = PSIMSOntology.objects.filter(
                    models.Q(name__icontains=query) | models.Q(synonyms__icontains=query), obsolete=False
                )[:limit]

            for term in psims_results:
                all_suggestions.append(
                    {
                        "identifier": term.identifier,
                        "name": term.name,
                        "definition": term.definition or "",
                        "source": "psi_ms",
                        "match_type": match_type,
                    }
                )

        elif ontology_type == "ms_unique_vocabularies":
            # MS Unique Vocabularies
            if match_type == "startswith":
                ms_vocab_results = MSUniqueVocabularies.objects.filter(
                    models.Q(name__istartswith=query) | models.Q(definition__istartswith=query)
                )[:limit]
            else:  # contains
                ms_vocab_results = MSUniqueVocabularies.objects.filter(
                    models.Q(name__icontains=query) | models.Q(definition__icontains=query)
                )[:limit]

            for vocab in ms_vocab_results:
                all_suggestions.append(
                    {
                        "identifier": vocab.accession,
                        "name": vocab.name,
                        "definition": vocab.definition or "",
                        "source": "ms_unique_vocabularies",
                        "match_type": match_type,
                    }
                )

        elif ontology_type == "unimod":
            # Unimod protein modifications
            if match_type == "startswith":
                unimod_results = Unimod.objects.filter(
                    models.Q(name__istartswith=query) | models.Q(definition__istartswith=query)
                )[:limit]
            else:  # contains
                unimod_results = Unimod.objects.filter(
                    models.Q(name__icontains=query) | models.Q(definition__icontains=query)
                )[:limit]

            for mod in unimod_results:
                all_suggestions.append(
                    {
                        "identifier": mod.accession,
                        "name": mod.name,
                        "definition": mod.definition or "",
                        "source": "unimod",
                        "match_type": match_type,
                    }
                )

        else:
            # No specific ontology type provided - return available types
            available_types = [
                "ncbi_taxonomy",
                "species",
                "mondo",
                "human_disease",
                "uberon",
                "tissue",
                "subcellular_location",
                "chebi",
                "cell_ontology",
                "cellosaurus",
                "psi_ms",
                "ms_unique_vocabularies",
                "unimod",
            ]

            return Response(
                {
                    "results": [],
                    "error": "Please specify an ontology type",
                    "available_types": available_types,
                    "match_types": ["contains", "startswith"],
                    "usage": "Add ?type=<ontology_type>&match=<match_type> to your request",
                    "examples": [
                        "?q=human&type=ncbi_taxonomy&match=contains",
                        "?q=cancer&type=mondo&match=startswith",
                        "?q=liver&type=uberon&match=contains",
                        "?q=HeLa&type=cellosaurus&match=startswith",
                        "?q=glucose&type=chebi&match=contains",
                        "?q=methylation&type=unimod&match=startswith",
                    ],
                }
            )

        # Sort by relevance (exact matches first, then partial matches)
        def sort_key(suggestion):
            name_lower = suggestion["name"].lower()
            query_lower = query.lower()

            if name_lower == query_lower:
                return (0, suggestion["name"])  # Exact match
            elif name_lower.startswith(query_lower):
                return (1, suggestion["name"])  # Starts with query
            else:
                return (2, suggestion["name"])  # Contains query

        all_suggestions.sort(key=sort_key)

        # Convert to standardized format using the serializer
        # Convert the raw suggestions to the standardized format
        standardized_suggestions = []
        for suggestion in all_suggestions[:limit]:
            # Create a data dict that matches the serializer expectation
            data = {
                "identifier": suggestion["identifier"],
                "name": suggestion["name"],
                "definition": suggestion.get("definition", ""),
                "ontology_type": ontology_type,
            }
            standardized_suggestions.append(data)

        serializer = OntologySuggestionSerializer(
            standardized_suggestions, many=True, context={"ontology_type": ontology_type}
        )

        return Response(
            {
                "results": serializer.data,
                "total": len(all_suggestions),
                "query": query,
                "ontology_type": ontology_type or "all",
            }
        )


class SchemaViewSet(FilterMixin, viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for Schema model - provides read-only access to schemas.
    Supports listing and retrieving schemas available to the user.
    """

    serializer_class = SchemaSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ["name", "display_name", "description"]
    filterset_fields = ["is_builtin", "is_active", "is_public", "tags"]

    def get_queryset(self):
        """Return schemas available to the user."""
        return Schema.get_available_schemas(user=self.request.user)

    @action(detail=False, methods=["get"])
    def available(self, request):
        """Get list of available schemas for creating templates."""
        schemas = self.get_queryset()
        serializer = self.get_serializer(schemas, many=True, context={"request": request})
        return Response(serializer.data)

    @action(detail=False, methods=["post"], permission_classes=[IsAdminUser])
    def sync_builtin(self, request):
        """Sync builtin schemas from sdrf-pipelines package (admin only)."""
        try:
            result = Schema.sync_builtin_schemas()

            if "error" in result:
                return Response(
                    {"error": f'Error syncing schemas: {result["error"]}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

            return Response(
                {
                    "message": "Schema sync completed successfully",
                    "created": result.get("created", 0),
                    "updated": result.get("updated", 0),
                }
            )

        except Exception as e:
            return Response(
                {"error": f"Error during schema sync: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SDRFDefaultsViewSet(viewsets.ViewSet):
    """
    ViewSet for SDRF default values - provides structured default values for SDRF columns.
    """

    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=["get"])
    def columns(self, request):
        """Get all available SDRF columns with their default values."""
        from .sdrf_defaults import get_all_column_defaults

        defaults = get_all_column_defaults()
        return Response({"columns": list(defaults.keys()), "total": len(defaults)})

    @action(detail=False, methods=["get"])
    def column_values(self, request):
        """Get default values for a specific column."""
        from .sdrf_defaults import get_column_defaults

        column_name = request.query_params.get("column")
        if not column_name:
            return Response({"error": "column parameter is required"}, status=status.HTTP_400_BAD_REQUEST)

        defaults = get_column_defaults(column_name)
        if not defaults:
            return Response({"error": f"No defaults found for column: {column_name}"}, status=status.HTTP_404_NOT_FOUND)

        return Response({"column": column_name, "config": defaults})

    @action(detail=False, methods=["get"])
    def search(self, request):
        """Search for columns containing the query string."""
        from .sdrf_defaults import search_columns

        query = request.query_params.get("q", "")
        if not query:
            return Response({"error": "q parameter is required"}, status=status.HTTP_400_BAD_REQUEST)

        results = search_columns(query)
        return Response({"query": query, "matches": results, "total": len(results)})

    @action(detail=False, methods=["get"])
    def structured_options(self, request):
        """Get options for structured fields (key-value pairs)."""
        from .sdrf_defaults import get_structured_field_options

        column_name = request.query_params.get("column")
        field_type = request.query_params.get("type")

        if not column_name:
            return Response({"error": "column parameter is required"}, status=status.HTTP_400_BAD_REQUEST)

        options = get_structured_field_options(column_name, field_type)
        if not options:
            return Response(
                {"error": f"No structured options found for column: {column_name}"}, status=status.HTTP_404_NOT_FOUND
            )

        return Response({"column": column_name, "field_type": field_type, "options": options})

    @action(detail=False, methods=["get"])
    def labels(self, request):
        """Get all label options (TMT, SILAC, label-free)."""
        from .sdrf_defaults import LABEL_VALUES

        return Response(
            {
                "label_free": LABEL_VALUES["label free sample"],
                "tmt_labels": LABEL_VALUES["tmt_labels"],
                "silac_labels": LABEL_VALUES["silac_labels"],
            }
        )

    @action(detail=False, methods=["get"])
    def modifications(self, request):
        """Get protein modification options (fixed and variable)."""
        from .sdrf_defaults import PROTEIN_MODIFICATIONS

        mod_type = request.query_params.get("type")
        if mod_type and mod_type in PROTEIN_MODIFICATIONS:
            return Response({"type": mod_type, "modifications": PROTEIN_MODIFICATIONS[mod_type]})

        return Response({"fixed": PROTEIN_MODIFICATIONS["fixed"], "variable": PROTEIN_MODIFICATIONS["variable"]})

    @action(detail=False, methods=["get"])
    def instruments(self, request):
        """Get list of common instrument models."""
        from .sdrf_defaults import INSTRUMENT_MODELS

        return Response({"instruments": INSTRUMENT_MODELS})

    @action(detail=False, methods=["get"])
    def cleavage_agents(self, request):
        """Get cleavage agent options."""
        from .sdrf_defaults import CLEAVAGE_AGENTS

        agent_type = request.query_params.get("type")
        if agent_type and agent_type in CLEAVAGE_AGENTS:
            return Response({"type": agent_type, "agents": CLEAVAGE_AGENTS[agent_type]})

        return Response({"enzymatic": CLEAVAGE_AGENTS["enzymatic"], "non_enzymatic": CLEAVAGE_AGENTS["non_enzymatic"]})

    @action(detail=False, methods=["get"])
    def compound_fields(self, request):
        """Get all compound fields that require special handling."""
        from .sdrf_defaults import get_all_compound_fields

        fields = get_all_compound_fields()
        return Response({"compound_fields": list(fields.keys()), "schemas": fields})

    @action(detail=False, methods=["get"])
    def compound_schema(self, request):
        """Get schema for a specific compound field."""
        from .sdrf_defaults import get_compound_field_schema

        column_name = request.query_params.get("column")
        if not column_name:
            return Response({"error": "column parameter is required"}, status=status.HTTP_400_BAD_REQUEST)

        schema = get_compound_field_schema(column_name)
        if not schema:
            return Response(
                {"error": f"No compound schema found for column: {column_name}"}, status=status.HTTP_404_NOT_FOUND
            )

        return Response({"column": column_name, "schema": schema})

    @action(detail=False, methods=["post"])
    def validate_compound(self, request):
        """Validate a compound field value against its schema."""
        from .sdrf_defaults import validate_compound_field_value

        column_name = request.data.get("column")
        value = request.data.get("value")

        if not column_name or value is None:
            return Response({"error": "column and value parameters are required"}, status=status.HTTP_400_BAD_REQUEST)

        validation_result = validate_compound_field_value(column_name, value)

        return Response({"column": column_name, "value": value, "validation": validation_result})

    @action(detail=False, methods=["get"])
    def suggestions(self, request):
        """Get column name suggestions based on partial input."""
        from .sdrf_defaults import get_column_suggestions

        partial = request.query_params.get("partial", "")
        if not partial:
            return Response({"error": "partial parameter is required"}, status=status.HTTP_400_BAD_REQUEST)

        suggestions = get_column_suggestions(partial)
        return Response({"partial": partial, "suggestions": suggestions, "total": len(suggestions)})

    @action(detail=False, methods=["get"])
    def quick_values(self, request):
        """Get quick access to commonly used value categories."""
        from .sdrf_defaults import get_quick_values

        category = request.query_params.get("category")
        if not category:
            return Response(
                {
                    "available_categories": [
                        "labels",
                        "instruments",
                        "organisms",
                        "diseases",
                        "cell_types",
                        "data_acquisition",
                        "fragmentation",
                        "replicates",
                    ]
                }
            )

        values = get_quick_values(category)
        if values is None:
            return Response({"error": f"Unknown category: {category}"}, status=status.HTTP_404_NOT_FOUND)

        return Response({"category": category, "values": values})
