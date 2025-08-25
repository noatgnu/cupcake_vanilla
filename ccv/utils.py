"""
CUPCAKE Vanilla - Utility functions for metadata management.

This module contains utility functions extracted from the main CUPCAKE project
for handling SDRF files, Excel templates, and metadata operations.
"""

import io
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sdrf_pipelines.sdrf.schemas import SchemaRegistry
from sdrf_pipelines.sdrf.sdrf import SDRFDataFrame

from .models import FavouriteMetadataOption, MetadataColumn, SamplePool, Schema


def sort_metadata(
    metadata_columns: List[MetadataColumn], sample_number: int, metadata_table=None
) -> Tuple[List[List[str]], Dict[int, Dict[str, Any]]]:
    """
    Sort metadata columns and create a structured output for SDRF export.

    Args:
        metadata_columns: List of MetadataColumn objects
        sample_number: Number of samples
        metadata_table: MetadataTable object (optional, used for pool information)

    Returns:
        Tuple of (result_data, id_map)
    """
    # Sort metadata columns by position and name
    sorted_metadata = sorted(metadata_columns, key=lambda x: (x.column_position or 0, x.name))

    # Get pool information for setting pooled sample column
    pooled_sample_status = {}  # sample_index -> status ("pooled", "not pooled", or None for not in any pool)
    if metadata_table:
        # Check ALL pools (both reference and non-reference) to determine pooled status
        all_pools = list(metadata_table.sample_pools.all())
        for pool in all_pools:
            # Samples in pooled_only_samples are marked as "pooled"
            for sample_idx in pool.pooled_only_samples:
                pooled_sample_status[sample_idx] = "pooled"

            # Samples in pooled_and_independent_samples are marked as "not pooled"
            # but they can still appear in SN= rows for reference pools
            for sample_idx in pool.pooled_and_independent_samples:
                pooled_sample_status[sample_idx] = "not pooled"

    # Create header row
    headers = []
    id_map = {}

    for i, metadata in enumerate(sorted_metadata):
        column_name = metadata.name
        headers.append(column_name)

        id_map[metadata.id] = {
            "column": i,
            "name": metadata.name,
            "type": metadata.type,
            "hidden": metadata.hidden,
        }

    # Create data rows
    result = [headers]

    for sample_idx in range(sample_number):
        row = []
        for metadata in sorted_metadata:
            value = ""

            # Special handling for pooled sample column
            if metadata.name.lower() == "pooled sample":
                sample_number_1_based = sample_idx + 1
                if sample_number_1_based in pooled_sample_status:
                    value = pooled_sample_status[sample_number_1_based]
                else:
                    # Independent samples that are not in any pool show as "not pooled"
                    value = "not pooled"
            else:
                # Check for sample-specific modifiers
                if metadata.modifiers and isinstance(metadata.modifiers, list):
                    # Look for sample-specific values in modifiers
                    for modifier in metadata.modifiers:
                        if isinstance(modifier, dict):
                            sample_range = modifier.get("samples", "")
                            # Parse sample range string (can be "1,2,3" or "4-6")
                            if sample_range:
                                sample_indices = []
                                for part in sample_range.split(","):
                                    part = part.strip()
                                    if "-" in part:
                                        start, end = part.split("-")
                                        sample_indices.extend(range(int(start), int(end) + 1))
                                    else:
                                        sample_indices.append(int(part))

                                if (sample_idx + 1) in sample_indices:
                                    value = modifier.get("value", "")
                                    break

                # Use default value if no modifier found
                if not value:
                    value = metadata.value or ""

            # Handle empty values with not_applicable flag
            if not value and metadata.not_applicable:
                value = "not applicable"
            elif not value and metadata.name.lower() != "pooled sample":
                # Don't set "not available" for pooled sample column - empty is fine
                value = "not available"

            row.append(value)
        result.append(row)

    return result, id_map


def sort_pool_metadata(
    metadata_columns: List[MetadataColumn], pools: List[SamplePool]
) -> Tuple[List[List[str]], Dict[int, Dict[str, Any]]]:
    """
    Sort pool metadata columns and create structured output.

    Args:
        metadata_columns: List of MetadataColumn objects for pools
        pools: List of SamplePool objects

    Returns:
        Tuple of (result_data, id_map)
    """
    # Sort metadata columns by position and name
    sorted_metadata = sorted(metadata_columns, key=lambda x: (x.column_position or 0, x.name))

    # Create header row
    headers = []
    id_map = {}

    for i, metadata in enumerate(sorted_metadata):
        column_name = metadata.name
        headers.append(column_name)

        id_map[metadata.id] = {
            "column": i,
            "name": metadata.name,
            "type": metadata.type,
            "hidden": metadata.hidden,
        }

    # Create data rows for pools
    result = [headers]

    for pool in pools:
        row = []
        for metadata in sorted_metadata:
            # For pools, use the pool name or metadata value
            if metadata.name.lower() == "pool name":
                value = pool.pool_name
            elif metadata.name.lower() == "pool description":
                value = pool.pool_description or ""
            else:
                value = metadata.value or ""

            row.append(value)
        result.append(row)

    return result, id_map


def convert_sdrf_to_metadata(name: str, value: str) -> str:
    """
    Convert SDRF values to standardized metadata format.
    Now handles favorite option format: [123] Human[*] -> Human

    Args:
        name: Metadata column name
        value: Raw value from SDRF

    Returns:
        Converted value
    """
    value = value.strip()

    # Handle favorite option format: [123] display_value[*] -> display_value
    if value.startswith("[") and "] " in value:
        # Extract the display value between "] " and the final "[*]" or "[**]" or "[***]"
        parts = value.split("] ", 1)
        if len(parts) == 2:
            display_part = parts[1]
            # Remove source markers [*], [**], [***] at the end
            if display_part.endswith("[*]"):
                value = display_part[:-3]
            elif display_part.endswith("[**]"):
                value = display_part[:-4]
            elif display_part.endswith("[***]"):
                value = display_part[:-5]
            else:
                value = display_part

    # Handle specific conversions based on name
    name_lower = name.lower()

    if name_lower in ["species", "organism"]:
        # Standardize species naming
        return value.strip()

    return value.strip()


def validate_sdrf(data: List[List[str]]) -> List[str]:
    """
    Validate SDRF data using the sdrf-pipelines library.

    Args:
        data: 2D array of SDRF data

    Returns:
        List of validation errors
    """
    errors = []

    try:
        # Convert data to DataFrame format expected by sdrf-pipelines
        df_string = "\n".join(["\t".join(row) for row in data])
        df = SDRFDataFrame.parse(io.StringIO(df_string))

        # Run validation
        errors.extend(df.validate())
        errors.extend(df.validate_experimental_design())

    except Exception as e:
        errors.append(f"SDRF validation error: {str(e)}")

    return errors


def detect_pooled_samples(data: List[List[str]], headers: List[str]) -> Tuple[Optional[int], List[int], List[int]]:
    """
    Detect pooled sample information in SDRF data.

    Args:
        data: 2D array of SDRF data (excluding headers)
        headers: List of column headers

    Returns:
        Tuple of (pooled_column_index, sn_rows, pooled_rows)
    """
    pooled_sample_column_index = None
    sn_rows = []
    pooled_rows = []

    # Find pooled sample column
    for i, header in enumerate(headers):
        header_lower = header.lower()
        if "pooled sample" in header_lower or "pooled_sample" in header_lower:
            pooled_sample_column_index = i
            break

    if pooled_sample_column_index is None:
        return None, [], []

    # Identify SN= rows and pooled rows
    for row_index, row in enumerate(data):
        if pooled_sample_column_index < len(row):
            cell_value = row[pooled_sample_column_index].strip()

            if cell_value.startswith("SN="):
                sn_rows.append(row_index)
            elif cell_value.lower() == "pooled":
                pooled_rows.append(row_index)

    return pooled_sample_column_index, sn_rows, pooled_rows


def get_favourite_metadata_options(user: Any, lab_group: Any = None, metadata_name: str = None) -> Dict[str, List[str]]:
    """
    Get favourite metadata options for dropdown generation.

    Args:
        user: User object
        lab_group: Lab group object (optional)
        metadata_name: Specific metadata name to filter (optional)

    Returns:
        Dictionary mapping metadata names to lists of options
    """
    favourites = {}

    # User-specific favourites
    user_favourites = FavouriteMetadataOption.objects.filter(user=user, lab_group__isnull=True, is_global=False)

    # Lab group favourites
    lab_group_favourites = (
        FavouriteMetadataOption.objects.filter(lab_group=lab_group, is_global=False)
        if lab_group
        else FavouriteMetadataOption.objects.none()
    )

    # Global favourites
    global_favourites = FavouriteMetadataOption.objects.filter(is_global=True)

    # Apply metadata name filter if provided
    if metadata_name:
        user_favourites = user_favourites.filter(name__iexact=metadata_name)
        lab_group_favourites = lab_group_favourites.filter(name__iexact=metadata_name)
        global_favourites = global_favourites.filter(name__iexact=metadata_name)

    # Process user favourites
    for fav in user_favourites:
        name_key = fav.name.lower()
        if name_key not in favourites:
            favourites[name_key] = []
        favourites[name_key].append(f"{fav.display_value or fav.value}[*]")

    # Process lab group favourites
    for fav in lab_group_favourites:
        name_key = fav.name.lower()
        if name_key not in favourites:
            favourites[name_key] = []
        favourites[name_key].append(f"{fav.display_value or fav.value}[**]")

    # Process global favourites
    for fav in global_favourites:
        name_key = fav.name.lower()
        if name_key not in favourites:
            favourites[name_key] = []
        favourites[name_key].append(f"{fav.display_value or fav.value}[***]")

    return favourites


def export_excel_template(
    metadata_columns: List[MetadataColumn],
    sample_number: int,
    pools: List[SamplePool] = None,
    favourites: Dict[str, List[str]] = None,
    field_mask_mapping: Dict[str, str] = None,
) -> Workbook:
    """
    Export metadata as Excel template with dropdowns and validation.

    Args:
        metadata_columns: List of metadata columns
        sample_number: Number of samples
        pools: Optional list of sample pools
        favourites: Optional dictionary of favourite options
        field_mask_mapping: Optional field name mappings

    Returns:
        Excel Workbook object
    """
    if favourites is None:
        favourites = {}
    if field_mask_mapping is None:
        field_mask_mapping = {}

    # Separate main and hidden metadata
    main_metadata = [m for m in metadata_columns if not m.hidden]
    hidden_metadata = [m for m in metadata_columns if m.hidden]

    # Sort metadata
    result_main, id_map_main = sort_metadata(main_metadata, sample_number)
    result_hidden, id_map_hidden = sort_metadata(hidden_metadata, sample_number) if hidden_metadata else ([], {})

    # Handle pools
    has_pools = pools and len(pools) > 0
    pool_result_main, pool_id_map_main = ([], {})

    if has_pools:
        pool_metadata = [m for m in metadata_columns if not m.hidden]  # Use same metadata for pools
        pool_result_main, pool_id_map_main = sort_pool_metadata(pool_metadata, pools)

    # Create workbook
    wb = Workbook()
    main_ws = wb.active
    main_ws.title = "main"

    # Create worksheets
    hidden_ws = wb.create_sheet(title="hidden")
    id_metadata_column_map_ws = wb.create_sheet(title="id_metadata_column_map")

    # Pool worksheets
    pool_main_ws = None
    if has_pools:
        pool_main_ws = wb.create_sheet(title="pool_main")

    # Fill ID mapping worksheet
    id_metadata_column_map_ws.append(["id", "column", "name", "type", "hidden"])
    for k, v in id_map_main.items():
        id_metadata_column_map_ws.append([k, v["column"], v["name"], v["type"], v["hidden"]])
    for k, v in id_map_hidden.items():
        id_metadata_column_map_ws.append([k, v["column"], v["name"], v["type"], v["hidden"]])

    # Styling
    fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Fill main worksheet
    if result_main:
        # Add headers and data
        main_ws.append(result_main[0])
        main_work_area = f"A1:{get_column_letter(len(result_main[0]))}{sample_number + 1}"

        for row in result_main[1:]:
            main_ws.append(row)

        # Apply styling
        for row in main_ws[main_work_area]:
            for cell in row:
                cell.fill = fill
                cell.border = thin_border

        # Auto-adjust column widths
        for col in main_ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except (TypeError, AttributeError):
                    pass
            adjusted_width = max_length + 2
            main_ws.column_dimensions[column].width = adjusted_width

        # Add notes
        note_texts = [
            "Note: Empty cells will be filled with 'not applicable' or " "'not available' when submitted.",
            "[*] User-specific favourite options.",
            "[**] Lab group-recommended options.",
            "[***] Global recommendations.",
        ]

        start_row = sample_number + 2
        for i, note_text in enumerate(note_texts):
            main_ws.merge_cells(
                start_row=start_row + i,
                start_column=1,
                end_row=start_row + i,
                end_column=len(result_main[0]),
            )
            note_cell = main_ws.cell(row=start_row + i, column=1)
            note_cell.value = note_text
            note_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Add dropdown validation
        _add_dropdown_validation(main_ws, result_main[0], favourites, field_mask_mapping, sample_number)

    # Fill hidden worksheet
    if result_hidden:
        hidden_ws.append(result_hidden[0])
        for row in result_hidden[1:]:
            hidden_ws.append(row)

        # Apply styling and validation
        _add_dropdown_validation(hidden_ws, result_hidden[0], favourites, field_mask_mapping, sample_number)

    # Fill pool worksheet
    if has_pools and pool_result_main:
        pool_main_ws.append(pool_result_main[0])
        for row in pool_result_main[1:]:
            pool_main_ws.append(row)

        # Apply styling and validation
        _add_dropdown_validation(
            pool_main_ws,
            pool_result_main[0],
            favourites,
            field_mask_mapping,
            len(pools),
        )

    return wb


def _add_dropdown_validation(
    worksheet,
    headers: List[str],
    favourites: Dict[str, List[str]],
    field_mask_mapping: Dict[str, str],
    row_count: int,
) -> None:
    """
    Add dropdown validation to a worksheet.

    Args:
        worksheet: Excel worksheet object
        headers: List of column headers
        favourites: Dictionary of favourite options
        field_mask_mapping: Field name mappings
        row_count: Number of data rows
    """
    required_metadata_names = {"tissue", "organism part", "disease", "species"}

    for i, header in enumerate(headers):
        name_splitted = header.split("[")
        if len(name_splitted) > 1:
            name = name_splitted[1].replace("]", "")
        else:
            name = name_splitted[0]

        required_column = name.lower() in required_metadata_names

        # Apply field masking
        name_capitalized = name.capitalize().replace("Ms1", "MS1").replace("Ms2", "MS2")
        if name_capitalized in field_mask_mapping:
            display_name = field_mask_mapping[name_capitalized]
            if len(name_splitted) > 1:
                worksheet.cell(row=1, column=i + 1).value = header.replace(
                    name_splitted[1].rstrip("]"), display_name.lower()
                )
            else:
                worksheet.cell(row=1, column=i + 1).value = display_name.lower()

        # Build option list
        option_list = []
        if required_column:
            option_list.append("not applicable")
        else:
            option_list.append("not available")

        if name.lower() in favourites:
            option_list.extend(favourites[name.lower()])

        # Add validation
        if option_list:
            dv = DataValidation(type="list", formula1=f'"{",".join(option_list)}"', showDropDown=False)
            col_letter = get_column_letter(i + 1)
            worksheet.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}{row_count + 1}")


# ===================================================================
# ONTOLOGY MAPPING AND VALIDATION FUNCTIONS
# ===================================================================


def detect_ontology_type(column_name: str, column_type: str) -> Optional[str]:
    """
    Automatically detect the appropriate ontology type based on column name and type.

    Args:
        column_name: Name of the metadata column
        column_type: Type of the metadata column (e.g., 'characteristics',
            'factor value')

    Returns:
        String indicating ontology type or None if no match
    """
    # Normalize column name for matching
    name_lower = column_name.lower().strip()

    # Species detection
    species_patterns = [
        "organism",
        "species",
        "taxonomy",
        "taxon",
        "organism_taxon_id",
        "ncbi taxid",
        "ncbi_taxid",
        "ncbitaxid",
    ]
    if any(pattern in name_lower for pattern in species_patterns):
        return "species"

    # Tissue detection
    tissue_patterns = [
        "tissue",
        "organ",
        "organism part",
        "organism_part",
        "body part",
        "anatomical part",
        "cell type",
        "cell_type",
        "sample type",
    ]
    if any(pattern in name_lower for pattern in tissue_patterns):
        return "tissue"

    # Disease detection
    disease_patterns = [
        "disease",
        "disorder",
        "condition",
        "pathology",
        "phenotype",
        "clinical finding",
        "medical condition",
    ]
    if any(pattern in name_lower for pattern in disease_patterns):
        return "disease"

    # Subcellular location detection
    subcellular_patterns = [
        "subcellular location",
        "cellular component",
        "cell component",
        "organelle",
        "compartment",
        "localization",
    ]
    if any(pattern in name_lower for pattern in subcellular_patterns):
        return "subcellular_location"

    # MS terms detection
    ms_patterns = [
        "instrument",
        "mass spectrometer",
        "ms instrument",
        "ionization",
        "fragmentation",
        "analyzer",
        "detector",
        "scan type",
        "mass accuracy",
    ]
    if any(pattern in name_lower for pattern in ms_patterns):
        return "ms_terms"

    # Unimod detection - be more specific to avoid false matches
    unimod_patterns = [
        "modification parameters",
        "ptm",
        "post-translational modification",
        "chemical modification",
        "protein modification",
        "unimod",
    ]
    if any(pattern in name_lower for pattern in unimod_patterns):
        return "unimod"

    return None


def apply_ontology_mapping_to_column(column: MetadataColumn) -> bool:
    """
    Apply automatic ontology detection and mapping to a metadata column.

    Args:
        column: MetadataColumn instance to update

    Returns:
        Boolean indicating if ontology type was detected and applied
    """
    if column.ontology_type:
        return False  # Already has ontology type

    detected_type = detect_ontology_type(column.name, column.type)
    if detected_type:
        column.ontology_type = detected_type
        column.save(update_fields=["ontology_type"])
        return True

    return False


def validate_sdrf_data_against_ontologies(
    headers: List[str],
    data_rows: List[List[str]],
    metadata_columns: List[MetadataColumn],
) -> Dict[str, List[Dict[str, Any]]]:
    """
    Validate SDRF data against ontologies and return validation results.

    Args:
        headers: List of column headers
        data_rows: List of data rows
        metadata_columns: List of MetadataColumn objects with ontology mappings

    Returns:
        Dictionary with validation results including errors and warnings
    """
    validation_results = {
        "errors": [],
        "warnings": [],
        "suggestions": {},
        "valid_values": {},
    }

    # Create header to column mapping
    header_to_column = {}
    for i, header in enumerate(headers):
        # Parse header format: name[type]
        if "[" in header and "]" in header:
            name = header.split("[")[0].strip()
            metadata_type = header.split("[")[1].rstrip("]").strip()
        else:
            name = header.strip()
            metadata_type = "characteristics"

        # Find matching metadata column
        for column in metadata_columns:
            if column.name == name and column.type == metadata_type:
                header_to_column[i] = column
                break

    # Validate each data cell
    for row_idx, row in enumerate(data_rows):
        for col_idx, value in enumerate(row):
            if col_idx in header_to_column:
                column = header_to_column[col_idx]
                header = headers[col_idx]

                if column.ontology_type and value.strip():
                    is_valid = column.validate_value_against_ontology(value)

                    if not is_valid:
                        validation_results["errors"].append(
                            {
                                "row": row_idx + 1,
                                "column": header,
                                "value": value,
                                "message": (f"Value '{value}' not found in " f"{column.ontology_type} ontology"),
                                "ontology_type": column.ontology_type,
                            }
                        )

                        # Get suggestions for invalid values
                        suggestions = column.get_ontology_suggestions(value, limit=5)
                        if suggestions:
                            key = f"{row_idx}_{col_idx}"
                            validation_results["suggestions"][key] = suggestions
                    else:
                        # Store valid values for reference
                        key = f"{header}_{value}"
                        validation_results["valid_values"][key] = True

    return validation_results


def get_ontology_suggestions_for_import(
    column_name: str, column_type: str, search_term: str = "", limit: int = 20
) -> List[Dict[str, Any]]:
    """
    Get ontology suggestions for import operations without requiring a MetadataColumn instance.

    Args:
        column_name: Name of the metadata column
        column_type: Type of the metadata column
        search_term: Optional search term to filter suggestions
        limit: Maximum number of suggestions to return

    Returns:
        List of suggestion dictionaries
    """
    ontology_type = detect_ontology_type(column_name, column_type)
    if not ontology_type:
        return []

    # Import ontology models dynamically to avoid circular imports
    from django.db import models

    from .models import HumanDisease, MSUniqueVocabularies, Species, SubcellularLocation, Tissue, Unimod

    ontology_mapping = {
        "species": Species,
        "tissue": Tissue,
        "disease": HumanDisease,
        "subcellular_location": SubcellularLocation,
        "ms_terms": MSUniqueVocabularies,
        "unimod": Unimod,
    }

    model_class = ontology_mapping.get(ontology_type)
    if not model_class:
        return []

    queryset = model_class.objects.all()

    # Apply search filtering based on model type
    if search_term:
        if ontology_type == "species":
            queryset = queryset.filter(
                models.Q(official_name__icontains=search_term)
                | models.Q(common_name__icontains=search_term)
                | models.Q(code__icontains=search_term)
            )
        elif ontology_type in ["tissue", "disease", "subcellular_location"]:
            queryset = queryset.filter(
                models.Q(accession__icontains=search_term) | models.Q(synonyms__icontains=search_term)
            )
        elif ontology_type in ["ms_terms", "unimod"]:
            queryset = queryset.filter(
                models.Q(name__icontains=search_term) | models.Q(definition__icontains=search_term)
            )

    return list(queryset[:limit].values())


def synchronize_pools_with_import_data(metadata_table, import_pools_data, metadata_columns, user):
    """
    Synchronize pools with import data for ccv (unified metadata system).
    Adapted from original CUPCAKE to work with ccv's simplified structure.
    """
    from .models import SamplePool

    # Get existing pools for this metadata table
    existing_pools = SamplePool.objects.filter(metadata_table=metadata_table)
    existing_pool_names = {pool.pool_name: pool for pool in existing_pools}
    existing_pool_ids = {pool.id: pool for pool in existing_pools}

    # Track pools found in import data
    import_pool_names = {pool_data["pool_name"] for pool_data in import_pools_data}

    # Process each pool from import data
    for pool_data in import_pools_data:
        pool_name = pool_data["pool_name"]
        pool_id = pool_data.get("pool_id")
        existing_pool = None

        # Try to find existing pool by ID first (more reliable), then by name
        if pool_id and pool_id in existing_pool_ids:
            existing_pool = existing_pool_ids[pool_id]
        elif pool_name in existing_pool_names:
            existing_pool = existing_pool_names[pool_name]

        if existing_pool:
            # Update existing pool
            existing_pool.pooled_only_samples = pool_data["pooled_only_samples"]
            existing_pool.pooled_and_independent_samples = pool_data["pooled_and_independent_samples"]
            existing_pool.is_reference = pool_data["is_reference"]
            existing_pool.save()

            # Clear existing metadata and recreate
            existing_pool.metadata_columns.clear()

            # Create new metadata for the updated pool
            _create_pool_metadata_from_import(existing_pool, pool_data, metadata_columns)
        else:
            # Create new pool
            new_pool = SamplePool.objects.create(
                pool_name=pool_name,
                pooled_only_samples=pool_data["pooled_only_samples"],
                pooled_and_independent_samples=pool_data["pooled_and_independent_samples"],
                is_reference=pool_data["is_reference"],
                metadata_table=metadata_table,
                created_by=user,
            )

            # Create metadata for the new pool
            _create_pool_metadata_from_import(new_pool, pool_data, metadata_columns)

    # Delete pools that are not in import data
    pools_to_delete = [pool for pool_name, pool in existing_pool_names.items() if pool_name not in import_pool_names]

    for pool in pools_to_delete:
        pool.delete()


def _calculate_most_common_value_for_column(data_rows, pooled_sample_indices, col_index, metadata_column):
    """
    Calculate the most common value for a column among pooled samples.

    Args:
        data_rows: All data rows from the import
        pooled_sample_indices: List of 1-based sample indices that belong to this pool
        col_index: Column index to analyze
        metadata_column: Metadata column for context

    Returns:
        Most common value among the pooled samples for this column
    """
    from collections import Counter

    values = []

    # Convert 1-based sample indices to 0-based row indices
    for sample_index in pooled_sample_indices:
        row_index = sample_index - 1  # Convert to 0-based
        if row_index < len(data_rows):
            row = data_rows[row_index]
            if col_index < len(row):
                cell_value = row[col_index].strip()
                if cell_value and cell_value != "":  # Exclude empty values
                    values.append(cell_value)

    if not values:
        # No values found - check if column allows not applicable
        if metadata_column.not_applicable:
            return "not applicable"
        else:
            return "not available"

    # Find the most common value
    value_counts = Counter(values)
    most_common_value = value_counts.most_common(1)[0][0]

    return most_common_value


def create_pool_metadata_from_table_columns(pool):
    """
    Create metadata columns for a pool based on the parent table's columns.
    This ensures the pool has all necessary metadata columns with appropriate values.

    Args:
        pool: SamplePool instance
    """
    from .models import MetadataColumn

    # Get the parent table's columns
    table_columns = list(pool.metadata_table.columns.all())

    # Clear existing pool metadata columns
    pool.metadata_columns.clear()

    for table_column in table_columns:
        if table_column.name.lower() == "pooled sample":
            # Set pooled sample column to SN= value
            pool_metadata_column = MetadataColumn.objects.create(
                name=table_column.name,
                type=table_column.type,
                value=pool.sdrf_value,  # This will generate SN=source1,source2,...
                mandatory=table_column.mandatory,
                hidden=table_column.hidden,
            )
        elif table_column.name.lower() == "source name":
            # Set source name column to pool name
            pool_metadata_column = MetadataColumn.objects.create(
                name=table_column.name,
                type=table_column.type,
                value=pool.pool_name,
                mandatory=table_column.mandatory,
                hidden=table_column.hidden,
            )
        else:
            # For other columns, calculate the most common value among pool samples
            pool_value = _calculate_most_common_value_for_pool_column(pool, table_column)
            pool_metadata_column = MetadataColumn.objects.create(
                name=table_column.name,
                type=table_column.type,
                value=pool_value,
                mandatory=table_column.mandatory,
                hidden=table_column.hidden,
            )

        # Add the metadata column to the pool
        pool.metadata_columns.add(pool_metadata_column)


def _calculate_most_common_value_for_pool_column(pool, table_column):
    """Calculate the most common value for a column among pool samples."""
    from collections import Counter

    pooled_sample_indices = pool.pooled_only_samples + pool.pooled_and_independent_samples
    values = []

    # Check if the table column has modifiers for specific samples
    if table_column.modifiers and isinstance(table_column.modifiers, list):
        for sample_idx in pooled_sample_indices:
            sample_value = None

            # Look for sample-specific modifiers
            for modifier in table_column.modifiers:
                if isinstance(modifier, dict):
                    sample_range = modifier.get("samples", "")
                    if sample_range:
                        sample_indices = []
                        for part in sample_range.split(","):
                            part = part.strip()
                            if "-" in part:
                                start, end = part.split("-")
                                sample_indices.extend(range(int(start), int(end) + 1))
                            else:
                                sample_indices.append(int(part))

                        if sample_idx in sample_indices:
                            sample_value = modifier.get("value", "")
                            break

            # Use default value if no modifier found
            if sample_value is None:
                sample_value = table_column.value or ""

            if sample_value:
                values.append(sample_value)
    else:
        # No modifiers, all samples use default value
        if table_column.value:
            return table_column.value

    if not values:
        # No values found - check if column allows not applicable
        if table_column.not_applicable:
            return "not applicable"
        else:
            return "not available"

    # Find the most common value
    value_counts = Counter(values)
    most_common_value = value_counts.most_common(1)[0][0]

    return most_common_value


def _create_pool_metadata_from_import(pool, pool_data, metadata_columns):
    """Create metadata for a pool from import data (ccv unified version)."""
    from .models import MetadataColumn

    row = pool_data.get("metadata_row")
    pool_name = pool_data["pool_name"]
    sdrf_value = pool_data["sdrf_value"]
    has_sn_pattern = sdrf_value.startswith("SN=")

    # Get all data rows for calculating most common values when no SN= pattern
    all_data_rows = pool_data.get("all_data_rows", [])
    pooled_sample_indices = pool_data["pooled_only_samples"] + pool_data["pooled_and_independent_samples"]

    for col_index, metadata_column in enumerate(metadata_columns):
        if metadata_column.name.lower() == "characteristics[pooled sample]":
            # Create pooled sample column with SN= value
            pool_metadata_column = MetadataColumn.objects.create(
                name=metadata_column.name,
                type=metadata_column.type,
                value=sdrf_value,
                mandatory=metadata_column.mandatory,
                hidden=metadata_column.hidden,
            )
        elif metadata_column.name.lower() == "source name":
            # Create source name column with pool name
            pool_metadata_column = MetadataColumn.objects.create(
                name=metadata_column.name,
                type=metadata_column.type,
                value=pool_name,
                mandatory=metadata_column.mandatory,
                hidden=metadata_column.hidden,
            )
        else:
            # Determine the value based on SN= pattern or most common value logic
            if has_sn_pattern and row:
                # Case 1: SN= pattern - use the SN= row's value directly
                raw_value = row[col_index] if col_index < len(row) else ""
            else:
                # Case 2: No SN= pattern - calculate most common value among pooled samples
                raw_value = _calculate_most_common_value_for_column(
                    all_data_rows, pooled_sample_indices, col_index, metadata_column
                )

            processed_value = raw_value

            # Process marker values for pools (same logic as main import)
            if raw_value and isinstance(raw_value, str):
                name = metadata_column.name.lower()
                if raw_value.endswith("[*]"):
                    processed_value = raw_value.replace("[*]", "")
                    # Try to find user favourite value
                    value_query = FavouriteMetadataOption.objects.filter(
                        user_id=pool.created_by.id,
                        name=name,
                        display_value=processed_value,
                        service_lab_group__isnull=True,
                        lab_group__isnull=True,
                    )
                    if value_query.exists():
                        processed_value = value_query.first().value
                elif raw_value.endswith("[**]") or raw_value.endswith("[***]"):
                    # Handle global/facility recommendations (simplified for ccv)
                    marker = "[**]" if raw_value.endswith("[**]") else "[***]"
                    processed_value = raw_value.replace(marker, "")
                    value_query = FavouriteMetadataOption.objects.filter(
                        name=name, display_value=processed_value, is_global=True
                    )
                    if value_query.exists():
                        processed_value = value_query.first().value
                elif raw_value.endswith("[****]"):
                    # Project suggestion - just strip the marker
                    processed_value = raw_value.replace("[****]", "")

            # Handle special cases for empty/not applicable values (same logic as main table)
            if not processed_value or processed_value.strip() == "":
                if metadata_column.not_applicable:
                    processed_value = "not applicable"
                else:
                    processed_value = "not available"

            pool_metadata_column = MetadataColumn.objects.create(
                name=metadata_column.name,
                type=metadata_column.type,
                value=processed_value,
                mandatory=metadata_column.mandatory,
                hidden=metadata_column.hidden,
            )

        # Add the metadata column to the pool
        pool.metadata_columns.add(pool_metadata_column)


def get_all_default_schema_names() -> List[str]:
    """
    Get all default schema names from the SchemaRegistry.

    Returns:
        List of default schema names.
    """
    return [schema for schema in SchemaRegistry().get_schema_names()]


def get_specific_default_schema(schema_name: str):
    """
    Get a specific default schema by name.

    Args:
        schema_name: Name of the schema to retrieve.

    Returns:
        The schema object if found, None otherwise.
    """
    registry = SchemaRegistry()
    if schema_name in registry.get_schema_names():
        return registry.get_schema(schema_name)
    return None


def compile_sdrf_columns_from_schemas(
    schema_names: list[str] = None, schema_ids: list[int] = None, schema_dir: str = None
) -> dict[str, list[str]]:
    """Compile SDRF column definitions from schemas."""
    import os
    import pickle
    import shutil
    import tempfile

    if schema_names and schema_ids:
        raise ValueError("Only one of schema_names or schema_ids should be provided, not both.")
    if not schema_names and not schema_ids:
        raise ValueError("At least one of schema_names or schema_ids must be provided.")

    sections = {
        "source_name": [],
        "characteristics": [],
        "special": [],
        "comment": [],
        "factor_value": [],
    }

    seen_columns = set()
    temp_dir = None

    try:
        if schema_ids:
            # Create temporary directory for schema files
            temp_dir = tempfile.mkdtemp(prefix="schema_processing_")

            # Copy schema files from database to temporary directory
            schema_objects = Schema.objects.filter(id__in=schema_ids, is_active=True)
            yml_files = []
            schemas = []
            for schema_obj in schema_objects:
                if schema_obj.schema_file:
                    # Get schema name without extension for SchemaRegistry
                    if schema_obj.schema_file.name.endswith(".yml"):
                        schema_filename = f"{schema_obj.name}.yml"
                        temp_schema_path = os.path.join(temp_dir, schema_filename)

                        # Copy the schema file content
                        with open(schema_obj.schema_file.path, "r") as source:
                            with open(temp_schema_path, "w") as dest:
                                dest.write(source.read())
                        yml_files.append(schema_obj)
                    elif schema_obj.schema_file.name.endswith(".pkl"):
                        schema = pickle.load(open(schema_obj.schema_file.path, "rb"))
                        schemas.append(schema)

            # Use temporary directory for SchemaRegistry
            registry = SchemaRegistry(temp_dir)

            # Get schemas by their names (without extension)

            for schema_obj in yml_files:
                schema = registry.get_schema(schema_obj.name)
                if schema:
                    schemas.append(schema)
        else:
            # Use schema names with provided or default schema directory
            registry = SchemaRegistry(schema_dir)
            schemas = []
            for schema_name in schema_names:
                schema = registry.get_schema(schema_name)
                if schema:
                    schemas.append(schema)

        # Process schemas to extract columns
        for schema in schemas:
            for column_def in schema.columns:
                column_name = column_def.name

                if column_name in seen_columns:
                    continue

                if column_name == "source name":
                    sections["source_name"].append(column_name)
                elif column_name.startswith("characteristics["):
                    sections["characteristics"].append(column_name)
                elif column_name.startswith("comment["):
                    sections["comment"].append(column_name)
                elif column_name.startswith("factor value["):
                    sections["factor_value"].append(column_name)
                else:
                    sections["special"].append(column_name)

                seen_columns.add(column_name)

    finally:
        # Clean up temporary directory
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)

    return sections


def reorder_columns_by_schema(
    input_file: str | Path,
    schema_names: list[str],
    output_file: Optional[str | Path] = None,
    schema_dir: Optional[str] = None,
) -> pd.DataFrame:
    """Reorder columns in a file according to schema definitions."""
    with open(input_file, "r") as f:
        header = f.readline().strip().lower().split("\t")

    df = pd.read_csv(input_file, sep="\t", dtype=str).fillna("")

    sections = compile_sdrf_columns_from_schemas(schema_names, schema_dir)

    processed_indices = set()

    final_col_positions_by_section = {
        "source_name": [],
        "characteristics": [],
        "special": [],
        "comment": [],
        "factor_value": [],
    }

    for section, columns in sections.items():
        for col in columns:
            found = False
            for i, orig_col in enumerate(header):
                if orig_col == col and i not in processed_indices:
                    final_col_positions_by_section[section].append(i)
                    processed_indices.add(i)
                    found = True
                    break

            if not found and col in header:
                indices = [i for i, x in enumerate(header) if x == col and i not in processed_indices]
                if indices:
                    final_col_positions_by_section[section].append(indices[0])
                    processed_indices.add(indices[0])

    for i, col in enumerate(header):
        if i in processed_indices:
            continue

        if col == "source name":
            final_col_positions_by_section["source_name"].append(i)
        elif col.startswith("characteristics["):
            final_col_positions_by_section["characteristics"].append(i)
        elif col.startswith("comment["):
            final_col_positions_by_section["comment"].append(i)
        elif col.startswith("factor value["):
            final_col_positions_by_section["factor_value"].append(i)
        else:
            final_col_positions_by_section["special"].append(i)

    final_col_positions = []
    final_col_positions.extend(final_col_positions_by_section["source_name"])
    final_col_positions.extend(final_col_positions_by_section["characteristics"])
    final_col_positions.extend(final_col_positions_by_section["special"])
    final_col_positions.extend(final_col_positions_by_section["comment"])
    final_col_positions.extend(final_col_positions_by_section["factor_value"])

    reordered_df = df.iloc[:, final_col_positions]

    if output_file:
        reordered_df.to_csv(output_file, sep="\t", index=False)

    return reordered_df
