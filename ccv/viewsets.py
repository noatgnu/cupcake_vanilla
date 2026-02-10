"""
Django REST Framework ViewSets for CUPCAKE Vanilla metadata management.
"""

import io
import json
import re

from django.contrib.auth.models import User
from django.db import models, transaction
from django.db.models import Q
from django.http import HttpResponse

from django_filters.rest_framework import DjangoFilterBackend
from django_filters.views import FilterMixin
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.filters import SearchFilter
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from rest_framework.response import Response

from ccc.models import LabGroup, ResourceRole, ResourceVisibility

from .models import (
    CellOntology,
    ChEBICompound,
    FavouriteMetadataOption,
    HumanDisease,
    MetadataColumn,
    MetadataColumnTemplate,
    MetadataColumnTemplateShare,
    MetadataTable,
    MetadataTableTemplate,
    MondoDisease,
    MSUniqueVocabularies,
    NCBITaxonomy,
    PSIMSOntology,
    SamplePool,
    Schema,
    Species,
    SubcellularLocation,
    Tissue,
    UberonAnatomy,
    Unimod,
)
from .permissions import MetadataColumnAccessPermission, MetadataTableAccessPermission
from .serializers import (
    CellOntologySerializer,
    ChEBICompoundSerializer,
    FavouriteMetadataOptionSerializer,
    HumanDiseaseSerializer,
    MetadataCollectionSerializer,
    MetadataColumnSerializer,
    MetadataColumnTemplateSerializer,
    MetadataColumnTemplateShareSerializer,
    MetadataExportSerializer,
    MetadataImportSerializer,
    MetadataTableSerializer,
    MetadataTableTemplateSerializer,
    MondoDiseaseSerializer,
    MSUniqueVocabulariesSerializer,
    NCBITaxonomySerializer,
    OntologySuggestionSerializer,
    PSIMSOntologySerializer,
    SamplePoolSerializer,
    SchemaSerializer,
    SpeciesSerializer,
    SubcellularLocationSerializer,
    TissueSerializer,
    UberonAnatomySerializer,
    UnimodSerializer,
)
from .tasks.import_utils import synchronize_pools_with_import_data
from .utils import (
    AutofillSpecValidator,
    SampleVariationGenerator,
    apply_ontology_mapping_to_column,
    detect_ontology_type,
    sort_metadata,
    validate_sdrf,
)


class MetadataTableViewSet(FilterMixin, viewsets.ModelViewSet):
    """ViewSet for managing MetadataTable objects."""

    queryset = MetadataTable.objects.all()
    serializer_class = MetadataTableSerializer
    permission_classes = [IsAuthenticated, MetadataTableAccessPermission]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ["name", "description", "owner__username"]
    filterset_fields = ["owner", "lab_group", "is_published", "is_locked"]

    def get_queryset(self):
        """
        Filter queryset based on user permissions and query parameters.

        By default, authenticated users only see their own tables. Users can
        opt-in to see shared content with ?show_shared=true. Unauthenticated
        users only see public tables.

        For list views, only shows metadata tables with source_app='ccv'.
        For detail views (accessing by ID), allows access to CCM tables too.
        """
        queryset = super().get_queryset()

        # For list views, only show CCV metadata tables
        # For detail views, allow access to CCM tables through permission classes
        if self.action == "list":
            queryset = queryset.filter(source_app="ccv")

        if hasattr(self.request, "user") and self.request.user.is_authenticated:
            user = self.request.user
            show_shared = self.request.query_params.get("show_shared", "false").lower() == "true"

            if show_shared:
                # Get all accessible lab groups (includes parent groups via bubble-up)
                from ccc.models import LabGroup

                accessible_groups = LabGroup.get_accessible_group_ids(user)

                permission_filter = (
                    Q(owner=user)
                    | Q(visibility=ResourceVisibility.PUBLIC)
                    | Q(visibility=ResourceVisibility.GROUP, lab_group_id__in=accessible_groups)
                    | Q(
                        resource_permissions__user=user,
                        resource_permissions__role__in=[
                            ResourceRole.VIEWER,
                            ResourceRole.EDITOR,
                            ResourceRole.ADMIN,
                            ResourceRole.OWNER,
                        ],
                    )
                )
                queryset = queryset.filter(permission_filter).distinct()
            else:
                queryset = queryset.filter(owner=user)
        else:
            queryset = queryset.filter(visibility=ResourceVisibility.PUBLIC)

        owner_id = self.request.query_params.get("owner_id")
        if owner_id:
            queryset = queryset.filter(owner_id=owner_id)

        lab_group_id = self.request.query_params.get("lab_group_id")
        if lab_group_id:
            queryset = queryset.filter(lab_group_id=lab_group_id)

        is_published = self.request.query_params.get("is_published")
        if is_published is not None:
            queryset = queryset.filter(is_published=is_published.lower() == "true")

        is_locked = self.request.query_params.get("is_locked")
        if is_locked is not None:
            queryset = queryset.filter(is_locked=is_locked.lower() == "true")

        return queryset.order_by("-created_at", "name")

    @action(detail=True, methods=["post"])
    def add_column(self, request, pk=None):
        """Add a new column to this metadata table."""
        table = self.get_object()

        if not table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        column_data = request.data.get("column_data", {})
        position = request.data.get("position")

        try:
            column = table.add_column(column_data, position)
            serializer = MetadataColumnSerializer(column)
            return Response(
                {
                    "message": "Column added successfully",
                    "column": serializer.data,
                }
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to add column: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def add_column_with_auto_reorder(self, request, pk=None):
        """Add a new column to this metadata table with automatic reordering."""
        table = self.get_object()

        if not table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        column_data = request.data.get("column_data", {})
        position = request.data.get("position")
        auto_reorder = request.data.get("auto_reorder", True)

        try:
            result = table.add_column_with_auto_reorder(column_data, position, auto_reorder)
            serializer = MetadataColumnSerializer(result["column"])

            return Response(
                {
                    "message": result["message"],
                    "column": serializer.data,
                    "reordered": result["reordered"],
                    "schema_ids_used": result["schema_ids_used"],
                }
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to add column with auto-reorder: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def remove_column(self, request, pk=None):
        """Remove a column from this metadata table."""
        table = self.get_object()

        if not table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        column_id = request.data.get("column_id")

        if not column_id:
            return Response(
                {"error": "column_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        success = table.remove_column(column_id)

        if success:
            return Response({"message": "Column removed successfully"})
        else:
            return Response(
                {"error": "Column not found or could not be removed"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def reorder_column(self, request, pk=None):
        """Reorder a column within this metadata table."""
        table = self.get_object()

        if not table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        column_id = request.data.get("column_id")
        new_position = request.data.get("new_position")

        if column_id is None or new_position is None:
            return Response(
                {"error": "column_id and new_position are required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        success = table.reorder_column(column_id, new_position)

        if success:
            return Response({"message": "Column reordered successfully"})
        else:
            return Response(
                {"error": "Column not found or could not be reordered"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def normalize_column_positions(self, request, pk=None):
        """Normalize column positions to be sequential starting from 0."""
        table = self.get_object()

        if not table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        table.normalize_column_positions()

        return Response({"message": "Column positions normalized successfully"})

    @action(detail=True, methods=["post"])
    def reorder_columns_by_schema_async(self, request, pk=None):
        """Start async reordering of table columns by schema."""
        from ccc.models import AsyncTaskStatus
        from ccv.tasks.reorder_tasks import reorder_metadata_table_columns_task

        table = self.get_object()

        if not table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        schema_ids = request.data.get("schema_ids", [])

        # Create task record
        task = AsyncTaskStatus.objects.create(
            task_type="REORDER_TABLE_COLUMNS",
            user=request.user,
            metadata_table=table,
            progress_current=0,
            progress_total=100,
            status="QUEUED",
            parameters={"schema_ids": schema_ids},
        )

        # Queue the task using delay
        job = reorder_metadata_table_columns_task.delay(
            metadata_table_id=table.id,
            user_id=request.user.id,
            schema_ids=schema_ids if schema_ids else None,
            task_id=str(task.id),
        )

        # Store job ID for tracking
        task.rq_job_id = job.id
        task.save()

        return Response(
            {
                "task_id": str(task.id),
                "message": "Column reordering task started",
                "metadata_table_id": table.id,
                "schema_ids": schema_ids,
            }
        )

    @action(detail=True, methods=["post"])
    def reorder_columns_by_schema(self, request, pk=None):
        """Reorder table columns by schema (sync or async based on environment)."""
        from ccv.tasks.reorder_tasks import reorder_metadata_table_columns_sync

        table = self.get_object()

        if not table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        schema_ids = request.data.get("schema_ids", [])
        force_async = request.data.get("async_processing", False)

        # Check if async processing is requested
        if force_async:
            # Delegate to async endpoint
            return self.reorder_columns_by_schema_async(request, pk)

        # Default to sync execution
        if True:
            # Sync execution
            result = reorder_metadata_table_columns_sync(
                metadata_table_id=table.id,
                user_id=request.user.id,
                schema_ids=schema_ids if schema_ids else None,
            )

            if result["success"]:
                return Response(
                    {
                        "message": "Column reordering completed successfully",
                        "metadata_table_id": table.id,
                        "schema_ids": schema_ids,
                        "result": result["result"],
                    }
                )
            else:
                return Response(
                    {"error": result["error"]},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            # Default to async when RQ is available
            return self.reorder_columns_by_schema_async(request, pk)

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def system_info(self, request):
        """Get system configuration information for frontend environment detection."""
        from django.conf import settings

        info = {
            "async_tasks_enabled": True,
            "is_electron_environment": getattr(settings, "IS_ELECTRON_ENVIRONMENT", False),
            "sync_operations_only": getattr(settings, "SYNC_OPERATIONS_ONLY", False),
        }

        # Add Electron-specific information if available
        if hasattr(settings, "ELECTRON_SETTINGS"):
            electron_settings = getattr(settings, "ELECTRON_SETTINGS", {})
            info.update(
                {
                    "electron_settings": {
                        "database_backend": electron_settings.get("DATABASE_BACKEND"),
                        "sync_operations_only": electron_settings.get("SYNC_OPERATIONS_ONLY", False),
                    }
                }
            )

        return Response(info)

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def admin_all_tables(self, request):
        """Admin endpoint to view all tables in the system."""
        if not (request.user.is_staff or request.user.is_superuser):
            return Response(
                {"error": "Permission denied: admin access required"},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Get all tables without filtering
        queryset = MetadataTable.objects.all()

        # Apply query parameter filters if provided
        owner_id = request.query_params.get("owner_id")
        if owner_id:
            queryset = queryset.filter(owner_id=owner_id)

        lab_group_id = request.query_params.get("lab_group_id")
        if lab_group_id:
            queryset = queryset.filter(lab_group_id=lab_group_id)

        is_published = request.query_params.get("is_published")
        if is_published is not None:
            queryset = queryset.filter(is_published=is_published.lower() == "true")

        is_locked = request.query_params.get("is_locked")
        if is_locked is not None:
            queryset = queryset.filter(is_locked=is_locked.lower() == "true")

        queryset = queryset.order_by("-created_at", "name")

        # Paginate the results
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["post"])
    def combine_columnwise(self, request):
        """
        Combine multiple metadata tables column-wise (side by side).

        Creates a new table with all columns from source tables combined,
        with the sample count set to the maximum among source tables.
        """
        data = request.data
        source_table_ids = data.get("source_table_ids", [])
        target_name = data.get("target_name", "")
        description = data.get("description", "")
        apply_schema_reordering = data.get("apply_schema_reordering", True)

        if not source_table_ids:
            return Response(
                {"error": "At least one source table ID is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not target_name:
            return Response(
                {"error": "Target table name is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Get source tables and check permissions
            source_tables = []
            for table_id in source_table_ids:
                try:
                    table = MetadataTable.objects.get(id=table_id)
                    if not table.can_view(request.user):
                        return Response(
                            {"error": f"Permission denied: cannot access table ID {table_id}"},
                            status=status.HTTP_403_FORBIDDEN,
                        )
                    source_tables.append(table)
                except MetadataTable.DoesNotExist:
                    return Response(
                        {"error": f"Table with ID {table_id} not found"},
                        status=status.HTTP_404_NOT_FOUND,
                    )

            # Combine tables column-wise
            combined_table = MetadataTable.combine_tables_columnwise(
                source_tables=source_tables,
                target_name=target_name,
                description=description,
                user=request.user,
                apply_schema_reordering=apply_schema_reordering,
            )

            # Serialize and return the new table
            serializer = self.get_serializer(combined_table)
            return Response(
                {"message": "Tables combined successfully (column-wise)", "combined_table": serializer.data},
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response(
                {"error": f"Failed to combine tables: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["post"])
    def combine_rowwise(self, request):
        """
        Combine multiple metadata tables row-wise (stacked vertically).

        Creates a new table with rows from all source tables stacked,
        using either union (all unique columns) or intersection (only common columns).
        """
        data = request.data
        source_table_ids = data.get("source_table_ids", [])
        target_name = data.get("target_name", "")
        description = data.get("description", "")
        apply_schema_reordering = data.get("apply_schema_reordering", True)
        merge_strategy = data.get("merge_strategy", "union")  # "union" or "intersection"

        if not source_table_ids:
            return Response(
                {"error": "At least one source table ID is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not target_name:
            return Response(
                {"error": "Target table name is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if merge_strategy not in ["union", "intersection"]:
            return Response(
                {"error": "Merge strategy must be 'union' or 'intersection'"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Get source tables and check permissions
            source_tables = []
            for table_id in source_table_ids:
                try:
                    table = MetadataTable.objects.get(id=table_id)
                    if not table.can_view(request.user):
                        return Response(
                            {"error": f"Permission denied: cannot access table ID {table_id}"},
                            status=status.HTTP_403_FORBIDDEN,
                        )
                    source_tables.append(table)
                except MetadataTable.DoesNotExist:
                    return Response(
                        {"error": f"Table with ID {table_id} not found"},
                        status=status.HTTP_404_NOT_FOUND,
                    )

            # Combine tables row-wise
            combined_table = MetadataTable.combine_tables_rowwise(
                source_tables=source_tables,
                target_name=target_name,
                description=description,
                user=request.user,
                apply_schema_reordering=apply_schema_reordering,
                merge_strategy=merge_strategy,
            )

            # Serialize and return the new table
            serializer = self.get_serializer(combined_table)
            return Response(
                {
                    "message": f"Tables combined successfully (row-wise, {merge_strategy})",
                    "combined_table": serializer.data,
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response(
                {"error": f"Failed to combine tables: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def perform_create(self, serializer):
        """Set the owner field when creating metadata tables."""
        serializer.save(owner=self.request.user)

    def perform_update(self, serializer):
        """Ensure user has edit permissions before updating."""
        table = self.get_object()
        if not table.can_edit(self.request.user):
            raise PermissionDenied("You don't have permission to edit this metadata table")
        serializer.save()

    def perform_destroy(self, instance):
        """Ensure user has edit permissions before deleting."""
        if not instance.can_edit(self.request.user):
            raise PermissionDenied("You don't have permission to delete this metadata table")
        super().perform_destroy(instance)

    @action(detail=True, methods=["post"])
    def validate_sample_count_change(self, request, pk=None):
        """
        Validate a sample count change and return affected data information.
        This allows the frontend to warn users about potential data loss.
        """
        from .serializers import SampleCountValidationSerializer

        metadata_table = self.get_object()

        # Check edit permissions
        if not metadata_table.can_edit(request.user):
            raise PermissionDenied("You don't have permission to edit this metadata table")

        serializer = SampleCountValidationSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        new_sample_count = serializer.validated_data["new_sample_count"]
        validation_result = metadata_table.validate_sample_count_change(new_sample_count)

        return Response(
            {
                "current_sample_count": metadata_table.sample_count,
                "new_sample_count": new_sample_count,
                **validation_result,
            }
        )

    @action(detail=True, methods=["post"])
    def validate_metadata_table(self, request, pk=None):
        """Validate metadata table (sync or async based on environment)."""
        from ccv.tasks.validation_tasks import validate_metadata_table_sync

        metadata_table = self.get_object()

        if not metadata_table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        validation_options = request.data.get("validation_options", {})
        force_async = request.data.get("async_processing", False)

        # Check if async processing is requested
        if force_async:
            # Delegate to async endpoint
            from .async_views import AsyncValidationViewSet

            async_view = AsyncValidationViewSet()
            # Prepare data for async view
            async_request_data = {
                "metadata_table_id": metadata_table.id,
                "validation_options": validation_options,
            }
            request._full_data = async_request_data
            return async_view.metadata_table(request)

        # Default to sync execution
        if True:
            # Sync execution
            result = validate_metadata_table_sync(
                metadata_table_id=metadata_table.id,
                user_id=request.user.id,
                validation_options=validation_options,
            )

            if result["success"]:
                return Response(
                    {
                        "message": "Validation completed successfully",
                        "metadata_table_id": metadata_table.id,
                        "validation_options": validation_options,
                        "result": result["result"],
                    }
                )
            else:
                return Response(
                    {"error": result["error"]},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            # Default to async when RQ is available
            from .async_views import AsyncValidationViewSet

            async_view = AsyncValidationViewSet()
            # Prepare data for async view
            async_request_data = {
                "metadata_table_id": metadata_table.id,
                "validation_options": validation_options,
            }
            request._full_data = async_request_data
            return async_view.metadata_table(request)

    @action(detail=True, methods=["post"])
    def update_sample_count(self, request, pk=None):
        """
        Update the sample count with user confirmation.
        This will clean up any modifiers and pools that reference invalid samples.
        """
        from .serializers import SampleCountUpdateSerializer

        metadata_table = self.get_object()

        # Check edit permissions
        if not metadata_table.can_edit(request.user):
            raise PermissionDenied("You don't have permission to edit this metadata table")

        serializer = SampleCountUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        new_sample_count = serializer.validated_data["new_sample_count"]
        confirmed = serializer.validated_data["confirmed"]

        # Validate the change first
        validation_result = metadata_table.validate_sample_count_change(new_sample_count)

        if not validation_result["valid"]:
            return Response(
                {"error": "Invalid sample count change", "validation_result": validation_result},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # If there are warnings (data will be removed) and user hasn't confirmed
        if validation_result["warnings"] and not confirmed:
            return Response(
                {
                    "error": "Confirmation required",
                    "message": "This change will remove data. Set confirmed=true to proceed.",
                    "validation_result": validation_result,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Apply the change
        old_sample_count = metadata_table.sample_count
        metadata_table.apply_sample_count_change(new_sample_count)

        return Response(
            {
                "message": f"Sample count updated from {old_sample_count} to {new_sample_count}",
                "old_sample_count": old_sample_count,
                "new_sample_count": new_sample_count,
                "cleanup_performed": new_sample_count < old_sample_count,
                "validation_result": validation_result,
            }
        )

    @action(detail=True, methods=["post"])
    def replace_column_value(self, request, pk=None):
        """
        Replace all occurrences of a specific value across columns in this table.

        Can target a specific column or all columns.

        Request body:
        - old_value: Value to replace (required)
        - new_value: New value (required)
        - column_id: Specific column ID to update (optional, if not provided updates all columns)
        - column_name: Specific column name to update (optional, alternative to column_id)
        - update_pools: Whether to update pool columns (default: true)
        """
        metadata_table = self.get_object()
        old_value = request.data.get("old_value")
        new_value = request.data.get("new_value")
        column_id = request.data.get("column_id")
        column_name = request.data.get("column_name")
        update_pools = request.data.get("update_pools", True)

        if old_value is None:
            return Response({"error": "old_value is required"}, status=status.HTTP_400_BAD_REQUEST)

        if new_value is None:
            return Response({"error": "new_value is required"}, status=status.HTTP_400_BAD_REQUEST)

        # Determine which columns to update
        if column_id:
            columns = metadata_table.columns.filter(id=column_id)
            if not columns.exists():
                return Response({"error": "Column not found in this table"}, status=status.HTTP_404_NOT_FOUND)
        elif column_name:
            columns = metadata_table.columns.filter(name=column_name)
            if not columns.exists():
                return Response(
                    {"error": f"Column '{column_name}' not found in this table"}, status=status.HTTP_404_NOT_FOUND
                )
        else:
            # Update all columns
            columns = metadata_table.columns.all()

        columns_updated = 0
        total_modifiers_merged = 0
        total_modifiers_deleted = 0
        total_samples_reverted_to_default = 0
        pools_updated = 0

        for column in columns:
            default_updated = False

            # Update column default value
            if column.value == old_value:
                column.value = new_value
                default_updated = True

            # Update modifiers with smart merging
            modifier_stats = MetadataColumnViewSet()._replace_value_in_modifiers(column, old_value, new_value)
            total_modifiers_merged += modifier_stats["modifiers_merged"]
            total_modifiers_deleted += modifier_stats["modifiers_deleted"]
            total_samples_reverted_to_default += modifier_stats["samples_reverted_to_default"]

            # Save column if anything changed
            if (
                default_updated
                or modifier_stats["modifiers_merged"] > 0
                or modifier_stats["modifiers_deleted"] > 0
                or modifier_stats["samples_reverted_to_default"] > 0
            ):
                column.save()
                columns_updated += 1

            # Update pool metadata columns
            if update_pools:
                sample_pools = metadata_table.sample_pools.all()
                for pool in sample_pools:
                    pool_columns = pool.metadata_columns.filter(name=column.name, value=old_value)
                    count = pool_columns.update(value=new_value)
                    pools_updated += count

        return Response(
            {
                "message": "Value replacement completed",
                "old_value": old_value,
                "new_value": new_value,
                "columns_checked": columns.count(),
                "columns_updated": columns_updated,
                "modifiers_merged": total_modifiers_merged,
                "modifiers_deleted": total_modifiers_deleted,
                "samples_reverted_to_default": total_samples_reverted_to_default,
                "pool_columns_updated": pools_updated,
            }
        )

    @action(detail=True, methods=["post"])
    def bulk_delete_columns(self, request, pk=None):
        """
        Delete multiple columns from this metadata table.

        For InstrumentJob metadata tables, respects staff-only permissions.
        """
        metadata_table = self.get_object()
        column_ids = request.data.get("column_ids", [])

        if not column_ids:
            return Response({"error": "column_ids is required"}, status=status.HTTP_400_BAD_REQUEST)

        if not metadata_table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        columns = metadata_table.columns.filter(id__in=column_ids)
        if not columns.exists():
            return Response({"error": "No columns found with the provided IDs"}, status=status.HTTP_404_NOT_FOUND)

        deleted_count = 0
        permission_denied_columns = []
        deleted_columns = []

        for column in columns:
            if metadata_table.source_app == "ccm":
                if hasattr(metadata_table, "instrument_jobs") and metadata_table.instrument_jobs.exists():
                    job = metadata_table.instrument_jobs.first()
                    if column.staff_only:
                        assigned_staff = job.staff.all()
                        has_assigned_staff = (
                            assigned_staff.exists() if hasattr(assigned_staff, "exists") else len(assigned_staff) > 0
                        )
                        if has_assigned_staff:
                            if request.user not in assigned_staff:
                                permission_denied_columns.append({"id": column.id, "name": column.name})
                                continue
                            if job.lab_group and not job.lab_group.is_member(request.user):
                                permission_denied_columns.append({"id": column.id, "name": column.name})
                                continue

            deleted_columns.append({"id": column.id, "name": column.name})
            column.delete()
            deleted_count += 1

        return Response(
            {
                "message": f"Deleted {deleted_count} column(s)",
                "deleted_count": deleted_count,
                "deleted_columns": deleted_columns,
                "permission_denied_columns": permission_denied_columns,
            }
        )

    @action(detail=True, methods=["post"])
    def bulk_update_staff_only(self, request, pk=None):
        """
        Mark or unmark multiple columns as staff-only.

        For InstrumentJob metadata tables, only assigned staff can modify staff_only status.
        """
        metadata_table = self.get_object()
        column_ids = request.data.get("column_ids", [])
        staff_only = request.data.get("staff_only")

        if not column_ids:
            return Response({"error": "column_ids is required"}, status=status.HTTP_400_BAD_REQUEST)

        if staff_only is None:
            return Response({"error": "staff_only is required (true or false)"}, status=status.HTTP_400_BAD_REQUEST)

        if not metadata_table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        columns = metadata_table.columns.filter(id__in=column_ids)
        if not columns.exists():
            return Response({"error": "No columns found with the provided IDs"}, status=status.HTTP_404_NOT_FOUND)

        updated_count = 0
        permission_denied_columns = []
        updated_columns = []

        for column in columns:
            if metadata_table.source_app == "ccm":
                if hasattr(metadata_table, "instrument_jobs") and metadata_table.instrument_jobs.exists():
                    job = metadata_table.instrument_jobs.first()
                    assigned_staff = job.staff.all()
                    has_assigned_staff = (
                        assigned_staff.exists() if hasattr(assigned_staff, "exists") else len(assigned_staff) > 0
                    )

                    if has_assigned_staff:
                        if request.user not in assigned_staff:
                            permission_denied_columns.append({"id": column.id, "name": column.name})
                            continue
                        if job.lab_group and not job.lab_group.is_member(request.user):
                            permission_denied_columns.append({"id": column.id, "name": column.name})
                            continue

            column.staff_only = staff_only
            column.save(update_fields=["staff_only"])
            updated_columns.append({"id": column.id, "name": column.name, "staff_only": staff_only})
            updated_count += 1

        return Response(
            {
                "message": f"Updated {updated_count} column(s)",
                "updated_count": updated_count,
                "updated_columns": updated_columns,
                "permission_denied_columns": permission_denied_columns,
                "staff_only": staff_only,
            }
        )

    @action(detail=True, methods=["post"])
    def advanced_autofill(self, request, pk=None):
        """
        Advanced autofill with template samples and column variations.

        Generates new samples based on template samples with specified
        column variations (ranges, lists, patterns, cartesian products).

        Request body:
        {
            "templateSamples": [1, 2],  # Sample indices to use as templates
            "targetSampleCount": 300,    # Total samples to generate
            "variations": [
                {
                    "columnId": 123,
                    "type": "range",     # range, list, pattern
                    "start": 1,          # for range type
                    "end": 30            # for range type
                },
                {
                    "columnId": 124,
                    "type": "list",
                    "values": ["TMT126", "TMT127N", "TMT127C", ...]
                }
            ],
            "fillStrategy": "cartesian_product"  # cartesian_product, sequential, interleaved
        }
        """
        table = self.get_object()

        if not table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"}, status=status.HTTP_403_FORBIDDEN
            )

        if table.is_locked:
            return Response({"error": "Cannot autofill locked table"}, status=status.HTTP_400_BAD_REQUEST)

        spec = request.data
        template_samples = spec.get("templateSamples", [])
        target_count = spec.get("targetSampleCount")
        variations = spec.get("variations", [])
        fill_strategy = spec.get("fillStrategy", "cartesian_product")

        validator = AutofillSpecValidator(spec, table)
        if not validator.is_valid():
            return Response({"errors": validator.errors}, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                generator = SampleVariationGenerator(table, template_samples, variations, target_count)
                variations_data = generator.generate_variations()

                if fill_strategy == "cartesian_product":
                    sample_variations = generator.cartesian_product(variations_data)
                elif fill_strategy == "sequential":
                    sample_variations = generator.sequential_fill(variations_data, target_count)
                elif fill_strategy == "interleaved":
                    sample_variations = generator.interleaved_fill(variations_data, target_count)
                else:
                    return Response(
                        {"error": f"Unknown fill strategy: {fill_strategy}"}, status=status.HTTP_400_BAD_REQUEST
                    )

                columns_to_update = generator.apply_variations_to_samples(sample_variations, target_count)
                MetadataColumn.objects.bulk_update(columns_to_update, ["modifiers"], batch_size=100)

            summary = generator.get_summary(sample_variations, fill_strategy)
            return Response(summary)

        except Exception as e:
            return Response({"error": f"Autofill failed: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class MetadataColumnViewSet(FilterMixin, viewsets.ModelViewSet):
    """ViewSet for managing MetadataColumn objects."""

    queryset = MetadataColumn.objects.all()
    serializer_class = MetadataColumnSerializer
    permission_classes = [IsAuthenticated, MetadataColumnAccessPermission]
    search_fields = ["name", "type", "value"]
    filterset_fields = ["metadata_table", "type", "mandatory", "hidden"]

    def get_queryset(self):
        """Filter queryset based on query parameters and source app permissions."""
        queryset = super().get_queryset()

        # Filter by metadata table
        metadata_table_id = self.request.query_params.get("metadata_table_id")
        if metadata_table_id:
            queryset = queryset.filter(metadata_table_id=metadata_table_id)

        # For list views, only show columns from CCV metadata tables
        # For detail views (accessing by ID), allow access to CCM table columns too
        if self.action == "list":
            # Filter out columns from CCM and other app tables
            queryset = queryset.filter(metadata_table__source_app="ccv")

        # Filter by metadata type
        metadata_type = self.request.query_params.get("type")
        if metadata_type:
            queryset = queryset.filter(type__icontains=metadata_type)

        # Filter by name
        name = self.request.query_params.get("name")
        if name:
            queryset = queryset.filter(name__icontains=name)

        # Filter by hidden status
        hidden = self.request.query_params.get("hidden")
        if hidden is not None:
            queryset = queryset.filter(hidden=hidden.lower() == "true")

        return queryset.order_by("column_position", "name")

    @action(detail=False, methods=["post"])
    def bulk_create(self, request):
        """Create multiple metadata columns at once."""
        serializer = self.get_serializer(data=request.data, many=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["post"])
    def validate_sdrf_data(self, request):
        """Validate SDRF data format."""
        metadata_ids = request.data.get("metadata_ids", [])
        sample_number = request.data.get("sample_number", 1)
        schema_name = request.data.get("schema_name", "default")
        skip_ontology = request.data.get("skip_ontology", False)

        metadata_columns = MetadataColumn.objects.filter(id__in=metadata_ids)
        result_data, _ = sort_metadata(list(metadata_columns), sample_number)

        validation_result = validate_sdrf(
            result_data,
            schema_name=schema_name,
            skip_ontology=skip_ontology,
        )

        errors = validation_result.get("errors", [])
        warnings = validation_result.get("warnings", [])

        return Response(
            {
                "valid": len(errors) == 0,
                "errors": errors,
                "warnings": warnings,
                "sample_count": len(result_data) - 1 if result_data else 0,
            }
        )

    @action(detail=False, methods=["get"])
    def ontology_suggestions(self, request):
        """
        Get ontology suggestions for a metadata column with enhanced search capabilities.

        Query Parameters:
        - column_id: ID of the metadata column
        - search: Search term to filter results
        - limit: Maximum number of results (default: 20, max: 100)
        - search_type: Type of search - 'icontains', 'istartswith', or 'exact' (default: 'icontains')
        """
        # Get column ID from query params
        column_id = request.query_params.get("column_id")
        if not column_id:
            return Response(
                {"error": "column_id parameter is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            column = MetadataColumn.objects.get(pk=column_id)
        except MetadataColumn.DoesNotExist:
            return Response(
                {"error": "Column not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        search_term = request.query_params.get("search", "")
        limit = min(int(request.query_params.get("limit", 20)), 100)  # Cap at 100
        search_type = request.query_params.get("search_type", "icontains")

        # Validate search_type
        valid_search_types = ["icontains", "istartswith", "exact"]
        if search_type not in valid_search_types:
            return Response(
                {"error": f"Invalid search_type. Must be one of: {', '.join(valid_search_types)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            suggestions = column.get_ontology_suggestions(search_term, limit, search_type)
        except Exception as e:
            return Response(
                {"error": f"Error retrieving suggestions: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # Serialize suggestions with proper field mappings
        serializer = OntologySuggestionSerializer(
            suggestions, many=True, context={"ontology_type": column.ontology_type}
        )

        return Response(
            {
                "ontology_type": column.ontology_type,
                "suggestions": serializer.data,
                "search_term": search_term,
                "search_type": search_type,
                "limit": limit,
                "count": len(suggestions),
                "custom_filters": column.custom_ontology_filters,
                "has_more": len(suggestions) >= limit,  # Indicate if there might be more results
            }
        )

    @action(detail=True, methods=["post"])
    def validate_value(self, request, pk=None):
        """Validate a value against the column's ontology."""
        column = self.get_object()
        value = request.data.get("value", "")

        is_valid = column.validate_value_against_ontology(value)

        response_data = {
            "valid": is_valid,
            "value": value,
            "ontology_type": column.ontology_type,
        }

        if not is_valid and column.ontology_type:
            # Provide suggestions for invalid values using enhanced search
            suggestions = column.get_ontology_suggestions(value, limit=5, search_type="icontains")
            response_data["suggestions"] = suggestions

        return Response(response_data)

    @action(detail=True, methods=["post"])
    def update_column_value(self, request, pk=None):
        """
        Update column value (default or sample-specific) with automatic modifier calculation.

        Request body:
        {
            "value": "new_value",
            "sample_indices": [1, 2, 3] or null,  # null means update default value
            "value_type": "default" | "sample_specific" | "replace_all"
        }

        Response:
        {
            "message": "Column value updated successfully",
            "column": {...},  # Updated column data
            "changes": {
                "old_default": "old_value",
                "new_default": "new_value",
                "old_modifiers": [...],
                "new_modifiers": [...],
                "updated_samples": [1, 2, 3]
            },
            "value_type": "sample_specific"
        }
        """
        column = self.get_object()

        if not column.metadata_table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        value = request.data.get("value", "")
        sample_indices = request.data.get("sample_indices")  # List of 1-based sample numbers or null
        value_type = request.data.get("value_type", "default")  # "default", "sample_specific", "replace_all"

        # Validate value_type
        valid_types = ["default", "sample_specific", "replace_all"]
        if value_type not in valid_types:
            return Response(
                {"error": f"Invalid value_type. Must be one of: {', '.join(valid_types)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate sample_indices for sample_specific updates
        if value_type == "sample_specific":
            if not sample_indices or not isinstance(sample_indices, list):
                return Response(
                    {"error": "sample_indices must be a non-empty list for sample_specific updates"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Validate that sample indices are within valid range
            max_samples = column.metadata_table.sample_count
            for idx in sample_indices:
                if not isinstance(idx, int) or idx < 1 or idx > max_samples:
                    return Response(
                        {"error": f"Invalid sample index {idx}. Must be between 1 and {max_samples}"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

        try:
            # Use the smart update method from the model
            changes = column.update_column_value_smart(
                value=value, sample_indices=sample_indices, value_type=value_type
            )

            # Save the column
            column.save()

            # Return updated column data
            serializer = MetadataColumnSerializer(column)

            return Response(
                {
                    "message": "Column value updated successfully",
                    "column": serializer.data,
                    "changes": changes,
                    "value_type": value_type,
                }
            )

        except Exception as e:
            return Response(
                {"error": f"Error updating column value: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=["post"])
    def bulk_update_sample_values(self, request, pk=None):
        """
        Bulk update column values for multiple samples with different values.

        Request body:
        {
            "updates": [
                {"sample_index": 1, "value": "run 1"},
                {"sample_index": 2, "value": "run 2"},
                {"sample_index": 3, "value": "run 3"}
            ]
        }

        Response:
        {
            "message": "Bulk update completed successfully",
            "updated_count": 3,
            "failed_count": 0,
            "column": {...}
        }
        """
        column = self.get_object()

        if not column.metadata_table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        updates = request.data.get("updates", [])
        if not updates or not isinstance(updates, list):
            return Response(
                {"error": "updates must be a non-empty list of {sample_index, value} objects"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        max_samples = column.metadata_table.sample_count
        updated_count = 0
        failed_updates = []

        for update in updates:
            sample_index = update.get("sample_index")
            value = update.get("value", "")

            if not isinstance(sample_index, int) or sample_index < 1 or sample_index > max_samples:
                failed_updates.append(
                    {
                        "sample_index": sample_index,
                        "error": f"Invalid sample index. Must be between 1 and {max_samples}",
                    }
                )
                continue

            try:
                column.update_column_value_smart(
                    value=value, sample_indices=[sample_index], value_type="sample_specific"
                )
                updated_count += 1
            except Exception as e:
                failed_updates.append({"sample_index": sample_index, "error": str(e)})

        column.save()

        serializer = MetadataColumnSerializer(column)

        response_data = {
            "message": "Bulk update completed",
            "updated_count": updated_count,
            "failed_count": len(failed_updates),
            "column": serializer.data,
        }

        if failed_updates:
            response_data["failed_updates"] = failed_updates

        return Response(response_data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"])
    def apply_ontology_mapping(self, request, pk=None):
        """Apply automatic ontology mapping to a column."""
        column = self.get_object()

        if not column.metadata_table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        was_applied = apply_ontology_mapping_to_column(column)

        return Response(
            {
                "applied": was_applied,
                "ontology_type": column.ontology_type,
                "message": (f"Ontology mapping {'applied' if was_applied else 'already present or not detected'}"),
            }
        )

    @action(detail=False, methods=["get"])
    def detect_ontology_type(self, request):
        """Detect ontology type for given column name and type."""

        column_name = request.query_params.get("name", "")
        column_type = request.query_params.get("type", "")

        if not column_name:
            return Response(
                {"error": "column name is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        detected_type = detect_ontology_type(column_name, column_type)

        return Response(
            {
                "column_name": column_name,
                "column_type": column_type,
                "detected_ontology_type": detected_type,
            }
        )

    def _replace_value_in_modifiers(self, column, old_value, new_value):
        """
        Replace value in modifiers, merging sample ranges if necessary.

        Handles several scenarios:
        1. New value matches default: Delete old modifiers (samples revert to default)
        2. New value exists in another modifier: Merge sample ranges
        3. New value is unique: Create new modifier with combined samples

        Returns:
            dict: Statistics about the replacement with keys:
                - modifiers_merged: Number of old modifiers merged into existing new_value modifier
                - modifiers_deleted: Number of old modifiers deleted (when new_value = default)
                - samples_reverted_to_default: Number of samples now using default value
        """
        if not column.modifiers:
            return {"modifiers_merged": 0, "modifiers_deleted": 0, "samples_reverted_to_default": 0}

        # Find modifiers with old_value and new_value
        old_modifiers = []
        new_modifier_idx = None
        new_modifier_samples = []

        for idx, modifier in enumerate(column.modifiers):
            if modifier.get("value") == old_value:
                old_modifiers.append((idx, modifier))
            elif modifier.get("value") == new_value:
                new_modifier_idx = idx
                new_modifier_samples = column._parse_sample_indices_from_modifier_string(modifier.get("samples", ""))

        if not old_modifiers:
            return {"modifiers_merged": 0, "modifiers_deleted": 0, "samples_reverted_to_default": 0}

        # Collect all sample indices from old_value modifiers
        all_old_samples = []
        for _, modifier in old_modifiers:
            samples = column._parse_sample_indices_from_modifier_string(modifier.get("samples", ""))
            all_old_samples.extend(samples)

        # Remove old modifiers (in reverse order to preserve indices)
        for idx, _ in reversed(old_modifiers):
            column.modifiers.pop(idx)

        modifiers_deleted = len(old_modifiers)
        samples_count = len(set(all_old_samples))

        # IMPORTANT: Check if new_value matches the default value
        # If so, the old modifiers are deleted and those samples automatically use the default
        if new_value == column.value:
            return {
                "modifiers_merged": 0,
                "modifiers_deleted": modifiers_deleted,
                "samples_reverted_to_default": samples_count,
            }

        # Merge with existing new_value modifier or create new one
        if new_modifier_idx is not None:
            # Merge sample ranges into existing modifier
            combined_samples = sorted(set(new_modifier_samples + all_old_samples))
            column.modifiers[new_modifier_idx]["samples"] = column._format_sample_indices_to_string(combined_samples)
            return {"modifiers_merged": len(old_modifiers), "modifiers_deleted": 0, "samples_reverted_to_default": 0}
        else:
            # Create new modifier with merged samples
            column.modifiers.append(
                {"samples": column._format_sample_indices_to_string(sorted(set(all_old_samples))), "value": new_value}
            )
            return {"modifiers_merged": 0, "modifiers_deleted": 0, "samples_reverted_to_default": 0}

    @action(detail=True, methods=["post"])
    def replace_value(self, request, pk=None):
        """
        Replace all occurrences of a specific value with a new value in this column.

        Intelligently handles modifiers by:
        - Merging sample ranges when replacing with an existing modifier value
        - Removing redundant modifiers when new value matches default
        - Preserving sample-specific assignments

        Updates:
        - Column default value
        - Values in modifiers (with smart merging)
        - Pool metadata columns (optional)

        Request body:
        - old_value: Value to replace (required)
        - new_value: New value (required)
        - update_pools: Whether to update pool columns (default: true)
        """
        column = self.get_object()
        old_value = request.data.get("old_value")
        new_value = request.data.get("new_value")
        update_pools = request.data.get("update_pools", True)

        if old_value is None:
            return Response({"error": "old_value is required"}, status=status.HTTP_400_BAD_REQUEST)

        if new_value is None:
            return Response({"error": "new_value is required"}, status=status.HTTP_400_BAD_REQUEST)

        default_updated = False

        # Update column default value
        if column.value == old_value:
            column.value = new_value
            default_updated = True

        # Update modifiers with smart merging
        modifier_stats = self._replace_value_in_modifiers(column, old_value, new_value)

        # Save column if anything changed
        if (
            default_updated
            or modifier_stats["modifiers_merged"] > 0
            or modifier_stats["modifiers_deleted"] > 0
            or modifier_stats["samples_reverted_to_default"] > 0
        ):
            column.save()

        # Update pool metadata columns
        pools_updated = 0
        if update_pools:
            sample_pools = column.metadata_table.sample_pools.all()
            for pool in sample_pools:
                pool_columns = pool.metadata_columns.filter(name=column.name, value=old_value)
                count = pool_columns.update(value=new_value)
                pools_updated += count

        return Response(
            {
                "message": "Value replacement completed",
                "old_value": old_value,
                "new_value": new_value,
                "default_value_updated": default_updated,
                "modifiers_merged": modifier_stats["modifiers_merged"],
                "modifiers_deleted": modifier_stats["modifiers_deleted"],
                "samples_reverted_to_default": modifier_stats["samples_reverted_to_default"],
                "pool_columns_updated": pools_updated,
            }
        )

    @action(detail=True, methods=["get"])
    def history(self, request, pk=None):
        """
        Get change history for a metadata column.

        Returns structured history with:
        - What changed (field names and values)
        - Who made the change (user)
        - When it happened (timestamp)
        - Change type (created, updated, deleted)

        Query Parameters:
        - limit: Maximum number of history records to return (default: 50, max: 200)
        - offset: Number of records to skip for pagination (default: 0)
        """
        column = self.get_object()

        limit = min(int(request.query_params.get("limit", 50)), 200)
        offset = int(request.query_params.get("offset", 0))

        history_records = list(column.history.all().order_by("-history_date")[offset : offset + limit])
        total_count = column.history.count()

        history_data = []

        for i, record in enumerate(history_records):
            next_record = history_records[i + 1] if i + 1 < len(history_records) else None
            changes = self._get_field_changes(record, next_record)

            history_entry = {
                "history_id": record.history_id,
                "history_date": record.history_date,
                "history_type": record.get_history_type_display(),
                "history_user": record.history_user.username if record.history_user else None,
                "history_user_id": record.history_user.id if record.history_user else None,
                "changes": changes,
                "snapshot": {
                    "name": record.name,
                    "type": record.type,
                    "value": record.value,
                    "column_position": record.column_position,
                    "mandatory": record.mandatory,
                    "hidden": record.hidden,
                    "readonly": record.readonly,
                    "modifiers": record.modifiers,
                    "ontology_type": record.ontology_type,
                    "not_applicable": record.not_applicable,
                    "not_available": record.not_available,
                },
            }

            history_data.append(history_entry)

        return Response(
            {
                "count": total_count,
                "limit": limit,
                "offset": offset,
                "has_more": (offset + limit) < total_count,
                "history": history_data,
            }
        )

    def _get_field_changes(self, current_record, previous_record):
        """
        Compare two historical records and return changed fields.

        Returns a list of changes with field name, old value, and new value.
        """
        if not previous_record:
            return []

        changes = []
        tracked_fields = [
            "name",
            "type",
            "value",
            "column_position",
            "mandatory",
            "hidden",
            "readonly",
            "modifiers",
            "ontology_type",
            "not_applicable",
            "not_available",
            "enable_typeahead",
            "staff_only",
        ]

        for field in tracked_fields:
            old_value = getattr(previous_record, field, None)
            new_value = getattr(current_record, field, None)

            if old_value != new_value:
                changes.append(
                    {
                        "field": field,
                        "old_value": old_value,
                        "new_value": new_value,
                    }
                )

        return changes

    def perform_update(self, serializer):
        """Update metadata column and sync hidden property to pool columns."""
        # Get the old instance to check what changed
        old_instance = self.get_object()
        old_hidden = old_instance.hidden

        # Save the updated instance
        instance = serializer.save()

        # If hidden property changed, sync to pool columns
        if old_hidden != instance.hidden:
            sample_pools = instance.metadata_table.sample_pools.all()
            for pool in sample_pools:
                pool_columns = pool.metadata_columns.filter(column_position=instance.column_position)
                if pool_columns.exists():
                    pool_columns.update(hidden=instance.hidden)


class SamplePoolViewSet(FilterMixin, viewsets.ModelViewSet):
    """ViewSet for managing SamplePool objects."""

    queryset = SamplePool.objects.all()
    serializer_class = SamplePoolSerializer
    permission_classes = [IsAuthenticated]
    search_fields = ["name", "description"]
    filterset_fields = ["owner", "lab_group"]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Filter by metadata table
        metadata_table_id = self.request.query_params.get("metadata_table_id")
        if metadata_table_id:
            queryset = queryset.filter(metadata_table_id=metadata_table_id)

        # Filter by reference status
        is_reference = self.request.query_params.get("is_reference")
        if is_reference is not None:
            queryset = queryset.filter(is_reference=is_reference.lower() == "true")

        return queryset.order_by("pool_name")

    def perform_create(self, serializer):
        """Set the created_by field and create pool metadata."""
        pool = serializer.save(created_by=self.request.user)

        # Create pool metadata columns based on table columns
        from .utils import create_pool_metadata_from_table_columns

        create_pool_metadata_from_table_columns(pool)

    def perform_update(self, serializer):
        """Update pool and refresh pooled sample columns."""
        pool = serializer.save()

        # Update pooled sample column values after pool changes
        from .utils import update_pooled_sample_column_for_table

        update_pooled_sample_column_for_table(pool.metadata_table)

    @action(detail=True, methods=["get"])
    def metadata_columns(self, request, pk=None):
        """Get metadata columns associated with this pool."""
        pool = self.get_object()

        # Get metadata columns for the same metadata table as the pool
        metadata_columns = MetadataColumn.objects.filter(metadata_table=pool.metadata_table)

        serializer = MetadataColumnSerializer(metadata_columns, many=True)
        return Response(serializer.data)


class MetadataTableTemplateViewSet(FilterMixin, viewsets.ModelViewSet):
    """ViewSet for managing MetadataTableTemplate objects."""

    queryset = MetadataTableTemplate.objects.all()
    serializer_class = MetadataTableTemplateSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ["name", "description"]
    filterset_fields = ["lab_group", "visibility", "is_default", "owner"]

    def get_queryset(self):
        """Filter queryset based on query parameters and user access."""
        queryset = super().get_queryset()

        # Filter by templates the user has access to:
        # 1. Templates created by the user
        # 2. Public templates
        # 3. Templates from lab groups the user is a member of (includes bubble-up from sub-groups)
        from ccc.models import LabGroup

        accessible_groups = LabGroup.get_accessible_group_ids(self.request.user)
        accessible_queryset = queryset.filter(
            Q(owner=self.request.user)
            | Q(visibility="public")  # User's own templates
            | Q(lab_group_id__in=accessible_groups)  # Lab group templates (includes bubble-up)
        ).distinct()

        # Apply additional query parameter filters
        owner_id = self.request.query_params.get("owner_id")
        if owner_id:
            accessible_queryset = accessible_queryset.filter(owner_id=owner_id)

        # Filter by lab group
        lab_group_id = self.request.query_params.get("lab_group_id")
        if lab_group_id:
            accessible_queryset = accessible_queryset.filter(lab_group_id=lab_group_id)

        # Filter by public status
        is_public = self.request.query_params.get("is_public")
        if is_public is not None:
            if is_public.lower() == "true":
                accessible_queryset = accessible_queryset.filter(visibility="public")
            else:
                accessible_queryset = accessible_queryset.exclude(visibility="public")

        # Filter by default status
        is_default = self.request.query_params.get("is_default")
        if is_default is not None:
            accessible_queryset = accessible_queryset.filter(is_default=is_default.lower() == "true")

        return accessible_queryset.order_by("-is_default", "name")

    @action(detail=True, methods=["post"])
    def add_column(self, request, pk=None):
        """Add a new column to this template."""
        template = self.get_object()

        if not template.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        column_data = request.data.get("column_data", {})
        position = request.data.get("position")

        try:
            column = template.add_column_to_template(column_data, position)
            serializer = MetadataColumnSerializer(column)
            return Response(
                {
                    "message": "Column added successfully",
                    "column": serializer.data,
                }
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to add column: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def remove_column(self, request, pk=None):
        """Remove a column from this template."""
        template = self.get_object()

        if not template.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        column_id = request.data.get("column_id")

        if not column_id:
            return Response(
                {"error": "column_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        success = template.remove_column_from_template(column_id)

        if success:
            return Response({"message": "Column removed successfully"})
        else:
            return Response(
                {"error": "Column not found or could not be removed"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def reorder_column(self, request, pk=None):
        """Reorder a column within this template."""
        template = self.get_object()

        if not template.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        column_id = request.data.get("column_id")
        new_position = request.data.get("new_position")

        if column_id is None or new_position is None:
            return Response(
                {"error": "column_id and new_position are required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        success = template.reorder_template_column(column_id, new_position)

        if success:
            return Response({"message": "Column reordered successfully"})
        else:
            return Response(
                {"error": "Column not found or could not be reordered"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def add_column_with_auto_reorder(self, request, pk=None):
        """Add a new column to this template with automatic reordering."""
        template = self.get_object()

        if not template.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        column_data = request.data.get("column_data", {})
        position = request.data.get("position")
        auto_reorder = request.data.get("auto_reorder", True)

        try:
            result = template.add_column_with_auto_reorder(column_data, position, auto_reorder)
            serializer = MetadataColumnSerializer(result["column"])

            return Response(
                {
                    "message": result["message"],
                    "column": serializer.data,
                    "reordered": result["reordered"],
                    "schema_ids_used": result["schema_ids_used"],
                }
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to add column: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def duplicate_column(self, request, pk=None):
        """Duplicate a column within this template."""
        template = self.get_object()

        if not template.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        column_id = request.data.get("column_id")
        new_name = request.data.get("new_name")

        if not column_id:
            return Response(
                {"error": "column_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        duplicated_column = template.duplicate_column_in_template(column_id, new_name)

        if duplicated_column:
            serializer = MetadataColumnSerializer(duplicated_column)
            return Response(
                {
                    "message": "Column duplicated successfully",
                    "column": serializer.data,
                }
            )
        else:
            return Response(
                {"error": "Column not found or could not be duplicated"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def normalize_column_positions(self, request, pk=None):
        """Normalize column positions to be sequential starting from 0."""
        template = self.get_object()

        if not template.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        template.normalize_template_column_positions()

        return Response({"message": "Column positions normalized successfully"})

    @action(detail=True, methods=["post"])
    def reorder_columns_by_schema_async(self, request, pk=None):
        """Start async reordering of template columns by schema."""
        from ccc.models import AsyncTaskStatus
        from ccv.tasks.reorder_tasks import reorder_metadata_table_template_columns_task

        template = self.get_object()

        if not template.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        schema_ids = request.data.get("schema_ids", [])

        # Create task record
        task = AsyncTaskStatus.objects.create(
            task_type="REORDER_TEMPLATE_COLUMNS",
            user=request.user,
            progress_current=0,
            progress_total=100,
            status="QUEUED",
            parameters={"schema_ids": schema_ids},
        )

        # Queue the task using delay
        job = reorder_metadata_table_template_columns_task.delay(
            template_id=template.id,
            user_id=request.user.id,
            schema_ids=schema_ids if schema_ids else None,
            task_id=str(task.id),
        )

        # Store job ID for tracking
        task.rq_job_id = job.id
        task.save()

        return Response(
            {
                "task_id": str(task.id),
                "message": "Template column reordering task started",
                "template_id": template.id,
                "schema_ids": schema_ids,
            }
        )

    @action(detail=True, methods=["post"])
    def reorder_columns_by_schema(self, request, pk=None):
        """Reorder template columns by schema (sync or async based on environment)."""
        from ccv.tasks.reorder_tasks import reorder_metadata_table_template_columns_sync

        template = self.get_object()

        if not template.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        schema_ids = request.data.get("schema_ids", [])
        force_async = request.data.get("async_processing", False)

        # Check if async processing is requested
        if force_async:
            # Delegate to async endpoint
            return self.reorder_columns_by_schema_async(request, pk)

        # Default to sync execution
        if True:
            # Sync execution
            result = reorder_metadata_table_template_columns_sync(
                template_id=template.id,
                user_id=request.user.id,
                schema_ids=schema_ids if schema_ids else None,
            )

            if result["success"]:
                return Response(
                    {
                        "message": "Template column reordering completed successfully",
                        "template_id": template.id,
                        "schema_ids": schema_ids,
                        "result": result["result"],
                    }
                )
            else:
                return Response(
                    {"error": result["error"]},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            # Default to async when RQ is available
            return self.reorder_columns_by_schema_async(request, pk)

    @action(detail=True, methods=["post"])
    def apply_to_metadata_table(self, request, pk=None):
        """Apply this template to a metadata table."""
        template = self.get_object()
        metadata_table_id = request.data.get("metadata_table_id")

        if not metadata_table_id:
            return Response(
                {"error": "metadata_table_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            metadata_table = MetadataTable.objects.get(id=metadata_table_id)
        except MetadataTable.DoesNotExist:
            return Response(
                {"error": "Invalid metadata_table_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not metadata_table.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Create metadata columns for the metadata table based on template
        created_columns = []
        for template_column in template.user_columns.all():
            new_column = MetadataColumn.objects.create(
                name=template_column.name,
                type=template_column.type,
                column_position=template_column.column_position,
                value=template_column.value,
                mandatory=template_column.mandatory,
                hidden=template_column.hidden,
                auto_generated=template_column.auto_generated,
                readonly=template_column.readonly,
                not_applicable=template_column.not_applicable,
                not_available=template_column.not_available,
                modifiers=template_column.modifiers,
                metadata_table=metadata_table,
            )
            created_columns.append(new_column)

        serializer = MetadataColumnSerializer(created_columns, many=True)
        return Response(
            {
                "message": f"Template applied successfully. Created {len(created_columns)} metadata columns.",
                "created_columns": serializer.data,
            }
        )

    @action(detail=False, methods=["get"])
    def available_schemas(self, request):
        """Get list of available schemas for creating templates."""

        # Get schemas available to the user
        schemas = Schema.get_available_schemas(user=request.user)

        schema_info = []
        for schema in schemas:
            schema_info.append(
                {
                    "id": schema.id,
                    "name": schema.name,
                    "display_name": schema.display_name,
                    "description": schema.description,
                    "is_builtin": schema.is_builtin,
                    "tags": schema.tags,
                    "usage_count": schema.usage_count,
                    "file_size": schema.file_size,
                    "created_at": schema.created_at.isoformat() if schema.created_at else None,
                }
            )

        return Response(schema_info)

    @action(detail=False, methods=["post"])
    def create_from_schema(self, request):
        """Create a new template from schema definitions."""

        # Validate required fields
        name = request.data.get("name")
        schema_ids = request.data.get("schema_ids", [])

        # Support legacy 'schemas' parameter for backward compatibility
        if not schema_ids and "schemas" in request.data:
            schema_names = request.data.get("schemas", [])
            # Convert names to IDs
            available_schemas = Schema.get_available_schemas(user=request.user)
            schema_name_to_id = {s.name: s.id for s in available_schemas}
            schema_ids = [schema_name_to_id.get(name) for name in schema_names if name in schema_name_to_id]
            schema_ids = [sid for sid in schema_ids if sid is not None]

        if not name:
            return Response({"error": "Template name is required"}, status=status.HTTP_400_BAD_REQUEST)

        if not schema_ids:
            default_schema = Schema.objects.filter(
                name__in=["ms-proteomics", "base", "default", "minimum"], is_active=True
            ).first()
            if default_schema:
                schema_ids = [default_schema.id]
            else:
                return Response(
                    {"error": "No default schema available. Please specify schema_ids."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Validate schemas exist and user has access
        available_schemas = Schema.get_available_schemas(user=request.user)
        available_ids = {s.id for s in available_schemas}
        invalid_schema_ids = [sid for sid in schema_ids if sid not in available_ids]
        if invalid_schema_ids:
            return Response(
                {"error": f'Invalid or inaccessible schema IDs: {", ".join(map(str, invalid_schema_ids))}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Get optional parameters
            description = request.data.get("description")
            lab_group_id = request.data.get("lab_group_id")
            is_public = request.data.get("is_public", False)
            is_default = request.data.get("is_default", False)

            # Convert is_public to visibility
            visibility = "public" if is_public else "private"

            # Get lab group if specified
            lab_group = None
            if lab_group_id:
                try:
                    lab_group = LabGroup.objects.get(id=lab_group_id)
                    # Check if user is member of this lab group (includes bubble-up from sub-groups)
                    if not lab_group.is_member(request.user):
                        return Response(
                            {"error": "You are not a member of the specified lab group"},
                            status=status.HTTP_403_FORBIDDEN,
                        )
                except LabGroup.DoesNotExist:
                    return Response({"error": "Lab group not found"}, status=status.HTTP_404_NOT_FOUND)

            # Create template from schemas
            template = MetadataTableTemplate.create_from_schemas(
                name=name,
                schema_ids=schema_ids,
                owner=request.user,
                lab_group=lab_group,
                description=description,
                visibility=visibility,
                is_default=is_default,
            )

            # Serialize and return the created template
            serializer = self.get_serializer(template)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {"error": f"Failed to create template: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=["post"])
    def create_table_from_template(self, request):
        """Create a new metadata table from an existing template."""
        # Validate required fields
        name = request.data.get("name")
        template_id = request.data.get("template_id")

        if not name:
            return Response({"error": "Table name is required"}, status=status.HTTP_400_BAD_REQUEST)

        if not template_id:
            return Response({"error": "Template ID is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Get the template
            template = MetadataTableTemplate.objects.get(id=template_id)

            # Check if user has access to this template
            if not template.can_view(request.user):
                return Response({"error": "You do not have access to this template"}, status=status.HTTP_403_FORBIDDEN)

            # Get optional parameters
            description = request.data.get("description")
            sample_count = request.data.get("sample_count", 1)
            lab_group_id = request.data.get("lab_group_id")

            # Get lab group if specified
            lab_group = None
            if lab_group_id:
                try:
                    lab_group = LabGroup.objects.get(id=lab_group_id)
                    # Check if user is member of this lab group (includes bubble-up from sub-groups)
                    if not lab_group.is_member(request.user):
                        return Response(
                            {"error": "You are not a member of the specified lab group"},
                            status=status.HTTP_403_FORBIDDEN,
                        )
                except LabGroup.DoesNotExist:
                    return Response({"error": "Lab group not found"}, status=status.HTTP_404_NOT_FOUND)

            # Create metadata table from template
            metadata_table = template.create_table_from_template(
                table_name=name,
                creator=request.user,
                sample_count=sample_count,
                description=description,
                lab_group=lab_group,
            )

            # Serialize and return the created table
            serializer = MetadataTableSerializer(metadata_table)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        except MetadataTableTemplate.DoesNotExist:
            return Response({"error": "Template not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response(
                {"error": f"Failed to create table from template: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["post"])
    def create_table_from_schemas(self, request):
        """Create a new metadata table directly from schema definitions."""

        # Validate required fields
        name = request.data.get("name")
        schema_ids = request.data.get("schema_ids", [])

        # Support legacy 'schemas' parameter for backward compatibility
        if not schema_ids and "schemas" in request.data:
            schema_names = request.data.get("schemas", [])
            # Convert names to IDs
            available_schemas = Schema.get_available_schemas(user=request.user)
            schema_name_to_id = {s.name: s.id for s in available_schemas}
            schema_ids = [schema_name_to_id.get(name) for name in schema_names if name in schema_name_to_id]
            schema_ids = [sid for sid in schema_ids if sid is not None]

        if not name:
            return Response({"error": "Table name is required"}, status=status.HTTP_400_BAD_REQUEST)

        if not schema_ids:
            default_schema = Schema.objects.filter(
                name__in=["ms-proteomics", "base", "default", "minimum"], is_active=True
            ).first()
            if default_schema:
                schema_ids = [default_schema.id]
            else:
                return Response(
                    {"error": "No default schema available. Please specify schema_ids."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Validate schemas exist and user has access
        available_schemas = Schema.get_available_schemas(user=request.user)
        available_ids = {s.id for s in available_schemas}
        invalid_schema_ids = [sid for sid in schema_ids if sid not in available_ids]
        if invalid_schema_ids:
            return Response(
                {"error": f'Invalid or inaccessible schema IDs: {", ".join(map(str, invalid_schema_ids))}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Get optional parameters
            description = request.data.get("description")
            sample_count = request.data.get("sample_count", 1)
            lab_group_id = request.data.get("lab_group_id")

            # Get lab group if specified
            lab_group = None
            if lab_group_id:
                try:
                    lab_group = LabGroup.objects.get(id=lab_group_id)
                    # Check if user is member of this lab group (includes bubble-up from sub-groups)
                    if not lab_group.is_member(request.user):
                        return Response(
                            {"error": "You are not a member of the specified lab group"},
                            status=status.HTTP_403_FORBIDDEN,
                        )
                except LabGroup.DoesNotExist:
                    return Response({"error": "Lab group not found"}, status=status.HTTP_404_NOT_FOUND)

            # Create a temporary template to leverage the schema functionality
            temp_template = MetadataTableTemplate(
                name=f"temp_template_for_{name}",
                owner=request.user,
                lab_group=lab_group,
                visibility="private",  # Set default visibility
            )

            # Create metadata table from schemas using the template method
            metadata_table = temp_template.create_table_from_schemas(
                table_name=name,
                schema_ids=schema_ids,
                owner=request.user,
                sample_count=sample_count,
                description=description,
                lab_group=lab_group,
            )

            # Serialize and return the created table
            serializer = MetadataTableSerializer(metadata_table)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {"error": f"Failed to create table from schemas: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=["post"])
    def update_field_mask(self, request, pk=None):
        """
        Update the field mask mapping for this template.

        Field masks allow customizing display names for columns in the frontend.
        The field_mask_mapping is a dictionary where keys are column names
        and values are the display names to show in the UI.

        Example request body:
        {
            "field_mask_mapping": {
                "characteristics[organism]": "Organism",
                "characteristics[organism part]": "Tissue",
                "source name": "Sample Name"
            }
        }
        """
        template = self.get_object()

        if not template.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        field_mask_mapping = request.data.get("field_mask_mapping", {})

        # Validate that the input is a dictionary
        if not isinstance(field_mask_mapping, dict):
            return Response(
                {"error": "field_mask_mapping must be a dictionary"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate keys are column names that exist in this template and values are strings
        template_column_names = set(col.name for col in template.user_columns.all())

        for key, value in field_mask_mapping.items():
            # Key should be a string (column name)
            if not isinstance(key, str):
                return Response(
                    {"error": "All keys in field_mask_mapping must be strings (column names)"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Optionally verify the column name exists in this template
            # (Comment this out if you want to allow field masks for columns that might be added later)
            if key not in template_column_names:
                return Response(
                    {"error": f"Column name '{key}' does not exist in this template"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Value should be a string (display name)
            if not isinstance(value, str):
                return Response(
                    {"error": "All values in field_mask_mapping must be strings (display names)"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        try:
            # Update the template's field mask mapping
            template.field_mask_mapping = field_mask_mapping
            template.save(update_fields=["field_mask_mapping"])

            return Response(
                {
                    "message": "Field mask mapping updated successfully",
                    "field_mask_mapping": template.field_mask_mapping,
                }
            )

        except Exception as e:
            return Response(
                {"error": f"Failed to update field mask mapping: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=["get"])
    def get_field_mask(self, request, pk=None):
        """
        Get the current field mask mapping for this template.

        Returns the field_mask_mapping dictionary that maps column names
        to their display names for use in the frontend.
        """
        template = self.get_object()

        if not template.can_view(request.user):
            return Response(
                {"error": "Permission denied: cannot access this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        return Response(
            {
                "field_mask_mapping": template.field_mask_mapping,
                "template_name": template.name,
                "template_id": template.id,
            }
        )

    @action(detail=True, methods=["get"])
    def get_columns_with_field_masks(self, request, pk=None):
        """
        Get all columns in this template along with their field masks.

        Returns a list of columns with their original names, display names (if field mask exists),
        and other column information to help the frontend build field mask UI.
        """
        template = self.get_object()

        if not template.can_view(request.user):
            return Response(
                {"error": "Permission denied: cannot access this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        columns = template.user_columns.all().order_by("column_position", "name")
        field_masks = template.field_mask_mapping

        column_data = []
        for column in columns:
            column_data.append(
                {
                    "id": column.id,
                    "name": column.name,
                    "type": column.type,
                    "display_name": field_masks.get(column.name, column.name),
                    "has_field_mask": column.name in field_masks,
                    "column_position": column.column_position,
                    "mandatory": column.mandatory,
                    "hidden": column.hidden,
                }
            )

        return Response(
            {
                "template_name": template.name,
                "template_id": template.id,
                "columns": column_data,
                "field_mask_mapping": field_masks,
            }
        )

    @action(detail=True, methods=["post"])
    def bulk_delete_columns(self, request, pk=None):
        """
        Delete multiple columns from this metadata table template.

        Request body:
        - column_ids: List of column IDs to delete
        """
        template = self.get_object()
        column_ids = request.data.get("column_ids", [])

        if not column_ids:
            return Response({"error": "column_ids is required"}, status=status.HTTP_400_BAD_REQUEST)

        if not template.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        columns = template.user_columns.filter(id__in=column_ids)
        if not columns.exists():
            return Response({"error": "No columns found with the provided IDs"}, status=status.HTTP_404_NOT_FOUND)

        deleted_count = 0
        deleted_columns = []

        for column in columns:
            deleted_columns.append({"id": column.id, "name": column.name})
            column.delete()
            deleted_count += 1

        return Response(
            {
                "message": f"Deleted {deleted_count} column(s)",
                "deleted_count": deleted_count,
                "deleted_columns": deleted_columns,
            }
        )

    @action(detail=True, methods=["post"])
    def bulk_update_staff_only(self, request, pk=None):
        """
        Mark or unmark multiple columns as staff-only in this template.

        Request body:
        - column_ids: List of column IDs to update
        - staff_only: Boolean value (true or false)
        """
        template = self.get_object()
        column_ids = request.data.get("column_ids", [])
        staff_only = request.data.get("staff_only")

        if not column_ids:
            return Response({"error": "column_ids is required"}, status=status.HTTP_400_BAD_REQUEST)

        if staff_only is None:
            return Response({"error": "staff_only is required (true or false)"}, status=status.HTTP_400_BAD_REQUEST)

        if not template.can_edit(request.user):
            return Response(
                {"error": "Permission denied: cannot edit this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        columns = template.user_columns.filter(id__in=column_ids)
        if not columns.exists():
            return Response({"error": "No columns found with the provided IDs"}, status=status.HTTP_404_NOT_FOUND)

        updated_count = 0
        updated_columns = []

        for column in columns:
            column.staff_only = staff_only
            column.save(update_fields=["staff_only"])
            updated_columns.append({"id": column.id, "name": column.name, "staff_only": staff_only})
            updated_count += 1

        return Response(
            {
                "message": f"Updated {updated_count} column(s)",
                "updated_count": updated_count,
                "updated_columns": updated_columns,
                "staff_only": staff_only,
            }
        )


class FavouriteMetadataOptionViewSet(FilterMixin, viewsets.ModelViewSet):
    """ViewSet for managing FavouriteMetadataOption objects."""

    queryset = FavouriteMetadataOption.objects.all()
    serializer_class = FavouriteMetadataOptionSerializer
    permission_classes = [IsAuthenticated]
    search_fields = ["name", "value", "display_value"]
    filterset_fields = ["is_global", "user", "lab_group"]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Filter by user
        user_id = self.request.query_params.get("user_id")
        if user_id:
            queryset = queryset.filter(user_id=user_id)

        # Filter by lab group
        lab_group_id = self.request.query_params.get("lab_group_id")
        if lab_group_id:
            queryset = queryset.filter(lab_group_id=lab_group_id)

        # Filter by global status
        is_global = self.request.query_params.get("is_global")
        if is_global is not None:
            queryset = queryset.filter(is_global=is_global.lower() == "true")

        # Filter by metadata name
        name = self.request.query_params.get("name")
        if name:
            queryset = queryset.filter(name__icontains=name)

        # Filter by metadata type
        metadata_type = self.request.query_params.get("type")
        if metadata_type:
            queryset = queryset.filter(type__icontains=metadata_type)

        return queryset.order_by("name", "type")


class MetadataManagementViewSet(viewsets.GenericViewSet):
    """ViewSet for metadata management operations like import/export."""

    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=["post"])
    def export_excel_template(self, request):
        """Export metadata as Excel template (matching original CUPCAKE)."""
        serializer = MetadataExportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data

        # Get metadata table
        try:
            metadata_table = MetadataTable.objects.get(id=data["metadata_table_id"])
        except (MetadataTable.DoesNotExist, KeyError):
            return Response(
                {"error": "metadata_table_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not metadata_table.can_view(request.user):
            return Response(
                {"error": "Permission denied: cannot view this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Check if async processing is requested
        if data.get("async_processing", False):
            # Queue async task
            from .async_views import AsyncExportViewSet

            async_view = AsyncExportViewSet()
            return async_view.excel_template(request)

        # Get metadata columns (user can specify specific columns or get all)
        if data.get("metadata_column_ids"):
            metadata_columns = metadata_table.columns.filter(id__in=data["metadata_column_ids"])
        else:
            metadata_columns = metadata_table.columns.all()

        # Separate main and hidden metadata (original CUPCAKE logic)
        main_metadata = [m for m in metadata_columns if not m.hidden]
        hidden_metadata = [m for m in metadata_columns if m.hidden]

        # Sort metadata and get structured output (original CUPCAKE approach)
        result_main, id_map_main = sort_metadata(main_metadata, metadata_table.sample_count, metadata_table)
        result_hidden = []
        id_map_hidden = {}
        if hidden_metadata:
            result_hidden, id_map_hidden = sort_metadata(hidden_metadata, metadata_table.sample_count, metadata_table)

        # Get pools and prepare pool data (original CUPCAKE logic)
        pools = list(metadata_table.sample_pools.all())
        has_pools = len(pools) > 0
        pool_id_map_main, pool_id_map_hidden = {}, {}

        if has_pools and data.get("include_pools", True):
            # Note: Pools in ccv don't have separate metadata like original CUPCAKE
            # We'll use the same metadata structure but organize by pool
            pass

        # Get favourites for each metadata column based on column names
        favourites = {}

        # Get column names from actual metadata columns being exported
        column_names = set(column.name.lower() for column in metadata_columns)

        # User-specific favourites - use case-insensitive name matching
        user_favourites = FavouriteMetadataOption.objects.filter(
            user=request.user,
            lab_group__isnull=True,
            name__iregex=r"^(" + "|".join(re.escape(name) for name in column_names) + ")$",
        )
        for fav in user_favourites:
            if fav.name.lower() not in favourites:
                favourites[fav.name.lower()] = []
            favourites[fav.name.lower()].append(f"[{fav.id}] {fav.display_value}[*]")

        # Lab group favourites
        lab_group_ids = data.get("lab_group_ids")
        if lab_group_ids is not None:  # Check for None vs empty list
            if lab_group_ids == []:
                # Empty list means "all lab groups"
                lab_favourites = FavouriteMetadataOption.objects.filter(
                    lab_group__isnull=False,
                    name__iregex=r"^(" + "|".join(re.escape(name) for name in column_names) + ")$",
                )
            else:
                # Specific lab group IDs
                lab_favourites = FavouriteMetadataOption.objects.filter(
                    lab_group_id__in=lab_group_ids,
                    name__iregex=r"^(" + "|".join(re.escape(name) for name in column_names) + ")$",
                )

            for fav in lab_favourites:
                if fav.name.lower() not in favourites:
                    favourites[fav.name.lower()] = []
                favourites[fav.name.lower()].append(f"[{fav.id}] {fav.display_value}[**]")
                # Add "not applicable" for required metadata
                if fav.name.lower() == "tissue" or fav.name.lower() == "organism part":
                    favourites[fav.name.lower()].append("not applicable")

        # Global recommendations
        global_favourites = FavouriteMetadataOption.objects.filter(
            is_global=True, name__iregex=r"^(" + "|".join(re.escape(name) for name in column_names) + ")$"
        )
        for fav in global_favourites:
            if fav.name.lower() not in favourites:
                favourites[fav.name.lower()] = []
            favourites[fav.name.lower()].append(f"[{fav.id}] {fav.display_value}[***]")

        # Create Excel workbook with multiple sheets (original CUPCAKE structure)
        wb = Workbook()
        main_ws = wb.active
        main_ws.title = "main"
        hidden_ws = wb.create_sheet(title="hidden")
        id_metadata_column_map_ws = wb.create_sheet(title="id_metadata_column_map")

        # Create pool sheets if pools exist
        pool_id_metadata_column_map_ws = None
        pool_object_map_ws = None

        if has_pools and data.get("include_pools", True):
            # Create pool worksheets with actual pool data
            pool_main_ws = wb.create_sheet(title="pool_main")
            pool_hidden_ws = wb.create_sheet(title="pool_hidden")
            pool_id_metadata_column_map_ws = wb.create_sheet(title="pool_id_metadata_column_map")
            pool_object_map_ws = wb.create_sheet(title="pool_object_map")

        # Fill ID metadata column mapping (original CUPCAKE format)
        id_metadata_column_map_ws.append(["id", "column", "name", "type", "hidden"])
        for k, v in id_map_main.items():
            id_metadata_column_map_ws.append([k, v["column"], v["name"], v["type"], v["hidden"]])
        for k, v in id_map_hidden.items():
            id_metadata_column_map_ws.append([k, v["column"], v["name"], v["type"], v["hidden"]])

        # Fill pool data and mapping if pools exist
        if has_pools and data.get("include_pools", True) and pool_id_metadata_column_map_ws:
            # Get pool metadata columns
            pool_main_metadata = [m for m in metadata_columns if not m.hidden]
            pool_hidden_metadata = [m for m in metadata_columns if m.hidden]

            # Generate pool data using sort_pool_metadata utility
            from .utils import sort_pool_metadata

            result_pool_main, pool_id_map_main = sort_pool_metadata(pool_main_metadata, pools)
            result_pool_hidden, pool_id_map_hidden = (
                sort_pool_metadata(pool_hidden_metadata, pools) if pool_hidden_metadata else ([], {})
            )

            # Fill pool_main sheet
            for row_data in result_pool_main:
                pool_main_ws.append(row_data)

            # Fill pool_hidden sheet if there's hidden data
            if result_pool_hidden:
                for row_data in result_pool_hidden:
                    pool_hidden_ws.append(row_data)

            # Fill pool ID metadata column mapping
            pool_id_metadata_column_map_ws.append(["id", "column", "name", "type", "hidden"])
            for k, v in pool_id_map_main.items():
                pool_id_metadata_column_map_ws.append([k, v["column"], v["name"], v["type"], v["hidden"]])
            for k, v in pool_id_map_hidden.items():
                pool_id_metadata_column_map_ws.append([k, v["column"], v["name"], v["type"], v["hidden"]])

            # Fill pool object mapping sheet
            pool_object_map_ws.append(
                [
                    "pool_name",
                    "pooled_only_samples",
                    "pooled_and_independent_samples",
                    "is_reference",
                    "sdrf_value",
                ]
            )
            for pool in pools:
                pool_object_map_ws.append(
                    [
                        pool.pool_name,
                        (json.dumps(pool.pooled_only_samples) if pool.pooled_only_samples else "[]"),
                        (
                            json.dumps(pool.pooled_and_independent_samples)
                            if pool.pooled_and_independent_samples
                            else "[]"
                        ),
                        pool.is_reference,
                        pool.sdrf_value or "",
                    ]
                )

        # Excel styling (original CUPCAKE formatting)
        fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Populate main worksheet (original CUPCAKE logic)
        if result_main and len(result_main) > 0:
            main_ws.append(result_main[0])
            main_work_area = f"A1:{get_column_letter(len(result_main[0]))}{metadata_table.sample_count + 1}"

            for row in result_main[1:]:
                main_ws.append(row)
        else:
            # Add placeholder data if no metadata columns exist
            main_ws.append(["No metadata columns"])
            main_work_area = "A1:A1"

        # Apply styling to main worksheet
        for row in main_ws[main_work_area]:
            for cell in row:
                cell.fill = fill
                cell.border = thin_border

        # Auto-adjust column widths for main worksheet
        for col in main_ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except (TypeError, AttributeError):
                    pass
            adjusted_width = max_length + 2
            main_ws.column_dimensions[column].width = adjusted_width

        # Add informational notes (original CUPCAKE notes)
        if result_main and len(result_main) > 0:
            note_texts = [
                "Note: Cells that are empty will automatically be filled with 'not applicable' or "
                "'not available' depending on the column when submitted.",
                "[ID] Format: [favorite_option_id] display_value[source_marker]",
                "[*] User-specific favourite options.",
                "[**] Lab group-recommended options.",
                "[***] Global recommendations.",
            ]

            start_row = metadata_table.sample_count + 2
            for i, note_text in enumerate(note_texts):
                main_ws.merge_cells(
                    start_row=start_row + i,
                    start_column=1,
                    end_row=start_row + i,
                    end_column=len(result_main[0]),
                )
                note_cell = main_ws.cell(row=start_row + i, column=1)
                note_cell.value = note_text
                note_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Populate hidden worksheet (original CUPCAKE logic)
        if result_hidden and len(result_hidden) > 0:
            hidden_work_area = f"A1:{get_column_letter(len(result_hidden[0]))}{metadata_table.sample_count + 1}"
            hidden_ws.append(result_hidden[0])
            for row in result_hidden[1:]:
                hidden_ws.append(row)
        else:
            # Add placeholder data if no hidden columns exist
            hidden_ws.append(["No hidden columns"])
            hidden_work_area = "A1:A1"

        # Apply styling to hidden worksheet
        for row in hidden_ws[hidden_work_area]:
            for cell in row:
                cell.fill = fill
                cell.border = thin_border

        # Auto-adjust column widths for hidden worksheet
        for col in hidden_ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except (TypeError, AttributeError):
                    pass
            adjusted_width = max_length + 2
            hidden_ws.column_dimensions[column].width = adjusted_width

        # Add data validation dropdowns for main worksheet (original CUPCAKE logic)
        if result_main and len(result_main) > 0:
            for i, header in enumerate(result_main[0]):
                name = header

                # Find the corresponding metadata column
                metadata_column = next((col for col in main_metadata if col.name == name), None)

                # Build option list
                option_list = []
                # Use column flags to determine appropriate empty value option
                if metadata_column:
                    if metadata_column.not_applicable:
                        option_list.append("not applicable")
                    elif metadata_column.not_available:
                        option_list.append("not available")
                else:
                    # Fallback for columns without metadata definition
                    option_list.append("not available")

                # Add favourites if available
                if name.lower() in favourites:
                    option_list.extend(favourites[name.lower()])

                # Create data validation
                if option_list:
                    dv = DataValidation(
                        type="list",
                        formula1=f'"{",".join(option_list)}"',
                        showDropDown=False,
                    )
                    col_letter = get_column_letter(i + 1)
                    main_ws.add_data_validation(dv)
                    dv.add(f"{col_letter}2:{col_letter}{metadata_table.sample_count + 1}")

        # Add data validation dropdowns for hidden worksheet (original CUPCAKE logic)
        if result_hidden and len(result_hidden) > 0:
            for i, header in enumerate(result_hidden[0]):
                name = header

                # Find the corresponding metadata column
                metadata_column = next((col for col in hidden_metadata if col.name == name), None)

                # Build option list
                option_list = []
                # Use column flags to determine appropriate empty value option
                if metadata_column:
                    if metadata_column.not_applicable:
                        option_list.append("not applicable")
                    elif metadata_column.not_available:
                        option_list.append("not available")
                else:
                    # Fallback for columns without metadata definition
                    option_list.append("not available")

                # Add favourites if available
                if name.lower() in favourites:
                    option_list.extend(favourites[name.lower()])

                # Create data validation
                if option_list:
                    dv = DataValidation(
                        type="list",
                        formula1=f'"{",".join(option_list)}"',
                        showDropDown=False,
                    )
                    col_letter = get_column_letter(i + 1)
                    hidden_ws.add_data_validation(dv)
                    dv.add(f"{col_letter}2:{col_letter}{metadata_table.sample_count + 1}")

        # Return Excel file as HTTP response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            content=output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="metadata_template.xlsx"'
        return response

    def _find_matching_column_template(self, column_name, metadata_table=None, templates=None):
        """
        Intelligent column template matching for SDRF import.

        Args:
            column_name: Name of the column from SDRF
            metadata_table: Target metadata table to search in first
            templates: List of available templates (from table's associated templates)

        Returns:
            MetadataColumnTemplate or None
        """
        # Normalize column name for matching
        normalized_name = column_name.lower().strip()

        # Step 1: Look in current table's existing columns
        if metadata_table:
            existing_column = metadata_table.columns.filter(name__iexact=column_name).first()
            if existing_column and existing_column.template:
                return existing_column.template

        # Step 2: Look in table's associated templates (from schemas)
        if templates:
            # Try exact match first
            exact_match = None
            for template in templates:
                if template.column_name.lower() == normalized_name:
                    exact_match = template
                    break

            if exact_match:
                return exact_match

            # Try partial match
            partial_matches = []
            for template in templates:
                template_name_lower = template.column_name.lower()
                # Check if column name contains template name or vice versa
                if normalized_name in template_name_lower or template_name_lower in normalized_name:
                    partial_matches.append(template)

            # Return the best partial match (shortest name = more specific)
            if partial_matches:
                return min(partial_matches, key=lambda t: len(t.column_name))

        # Step 3: Look in global/system templates
        # Try exact match in system templates
        system_template = MetadataColumnTemplate.objects.filter(
            column_name__iexact=column_name, is_system_template=True, is_active=True
        ).first()

        if system_template:
            return system_template

        # Try partial match in system templates
        system_partial = MetadataColumnTemplate.objects.filter(
            column_name__icontains=normalized_name, is_system_template=True, is_active=True
        ).first()

        return system_partial

    def _get_column_type_from_name(self, column_name):
        """Determine column type based on SDRF naming conventions."""
        name_lower = column_name.lower()

        if name_lower.startswith("characteristics[") or "characteristics" in name_lower:
            return "characteristics"
        elif name_lower.startswith("factor value[") or "factor value" in name_lower:
            return "factor_value"
        elif name_lower.startswith("comment[") or "comment" in name_lower:
            return "comment"
        elif name_lower in ["source name", "source_name"]:
            return "source_name"
        elif name_lower in ["assay name", "assay_name"]:
            return "assay_name"
        elif name_lower in ["technology name", "technology_name"]:
            return "technology_name"
        else:
            return "characteristics"  # Default fallback

    def _organize_columns_by_schema(self, metadata_table, created_columns):
        """
        Reorder columns based on schemas found in the table's templates.

        Organizing by sections and putting unmatched columns last.
        """
        # Get all schemas associated with the table's templates
        table_templates = []

        # Get templates from table's lab group if it exists
        if metadata_table.lab_group:
            table_templates.extend(
                MetadataColumnTemplate.objects.filter(lab_group=metadata_table.lab_group, is_active=True)
            )

        # Get templates from the table owner
        if metadata_table.owner:
            table_templates.extend(MetadataColumnTemplate.objects.filter(owner=metadata_table.owner, is_active=True))

        # Get system templates
        table_templates.extend(MetadataColumnTemplate.objects.filter(is_system_template=True, is_active=True))

        # Get unique schemas from these templates
        schema_ids = set()
        for template in table_templates:
            if hasattr(template, "source_schema") and template.source_schema:
                try:
                    schema = Schema.objects.get(name=template.source_schema, is_active=True)
                    schema_ids.add(schema.id)
                except Schema.DoesNotExist:
                    continue

        # If we have schemas, use the table's reorder method
        if schema_ids:
            metadata_table.reorder_columns_by_schema(schema_ids=list(schema_ids))
        else:
            # Fallback: basic section-based ordering
            self._basic_column_reordering(metadata_table, created_columns)

        # Apply the same reordering logic to sample pool columns
        created_pools = list(metadata_table.sample_pools.all())
        if created_pools:
            for pool in created_pools:
                if pool.metadata_columns.exists():
                    if schema_ids:
                        try:
                            pool.reorder_pool_columns_by_schema(schema_ids=list(schema_ids))
                        except Exception as e:
                            # Log error but don't fail the import, fall back to basic reordering
                            print(f"Warning: Failed to reorder pool '{pool.pool_name}' columns by schema: {e}")
                            pool.basic_pool_column_reordering()
                    else:
                        # No schemas found, use basic section-based ordering for pools
                        pool.basic_pool_column_reordering()

    def _basic_column_reordering(self, metadata_table, created_columns):
        """Basic column reordering by section when no schemas are available."""
        sections = {
            "source_name": [],
            "characteristics": [],
            "assay_name": [],
            "technology_name": [],
            "comment": [],
            "factor_value": [],
            "other": [],
        }

        # Group columns by type
        for column in metadata_table.columns.all():
            col_type = column.type.lower()
            if col_type in sections:
                sections[col_type].append(column)
            else:
                sections["other"].append(column)

        # Assign positions
        position = 0
        section_order = [
            "source_name",
            "characteristics",
            "assay_name",
            "technology_name",
            "comment",
            "factor_value",
            "other",
        ]

        for section in section_order:
            for column in sections[section]:
                column.column_position = position
                column.save(update_fields=["column_position"])
                position += 1

    def _get_table_template(self, metadata_table):
        """Get the table template that existing columns originate from."""

        # Find table template by looking at existing columns' template references
        existing_columns = list(metadata_table.columns.all())
        if existing_columns:
            # Look for columns that have template references
            template_column_templates = set()
            for col in existing_columns:
                if col.template:
                    template_column_templates.add(col.template)

            if template_column_templates:
                # Find which table template contains the most of these column templates
                best_template = None
                best_match_count = 0

                for table_template in MetadataTableTemplate.objects.all():
                    # Get all template columns in this table template
                    table_template_columns = set()
                    for template_col in table_template.user_columns.all():
                        if hasattr(template_col, "template") and template_col.template:
                            table_template_columns.add(template_col.template)

                    # Count matches
                    match_count = len(template_column_templates.intersection(table_template_columns))
                    if match_count > best_match_count:
                        best_match_count = match_count
                        best_template = table_template

                if best_template and best_match_count > 0:
                    return best_template

        # If no template found from existing columns, try to infer from table metadata
        # This could be enhanced to look at table creation history or naming patterns
        return None

    def _find_or_create_matching_column(
        self, clean_name, metadata_type, metadata_table, table_template, column_position, occurrence_number
    ):
        """Find existing column or create new one using template properties."""

        # Get all existing columns with the same name (case insensitive)
        existing_columns = list(metadata_table.columns.filter(name__iexact=clean_name).order_by("column_position"))

        if occurrence_number <= len(existing_columns):
            # Replace content of existing column at this occurrence
            existing_column = existing_columns[occurrence_number - 1]
            existing_column.column_position = column_position
            existing_column.save(update_fields=["column_position"])
            return existing_column
        else:
            # Create new column (occurrence_number > existing columns count)
            # Look for template column to copy properties from
            template_column = None
            if table_template:
                # Try to find exact match in template - user_columns contains MetadataColumn objects
                template_column = table_template.user_columns.filter(name__iexact=clean_name).first()

            # Create new column with template properties if available
            if template_column:
                # template_column is a MetadataColumn from the table template
                metadata_column = MetadataColumn.objects.create(
                    name=clean_name,
                    type=metadata_type or template_column.type,
                    column_position=column_position,
                    metadata_table=metadata_table,
                    template=template_column.template,  # Reference the MetadataColumnTemplate if it exists
                    ontology_type=template_column.ontology_type,
                    mandatory=template_column.mandatory,
                    hidden=template_column.hidden,
                    readonly=template_column.readonly,
                    auto_generated=template_column.auto_generated,
                    not_applicable=template_column.not_applicable,
                    not_available=template_column.not_available,
                )
            else:
                # Create basic column without template
                metadata_column = MetadataColumn.objects.create(
                    name=clean_name,
                    type=metadata_type,
                    column_position=column_position,
                    metadata_table=metadata_table,
                )

                # Note: Ontology mapping should only be applied on user request, not automatically

            return metadata_column

    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser])
    def import_sdrf_file(self, request):
        """Import metadata from SDRF file with intelligent column matching and schema organization."""
        serializer = MetadataImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data

        # Check if async processing is requested
        if data.get("async_processing", False):
            # Queue async task
            from .async_views import AsyncImportViewSet

            async_view = AsyncImportViewSet()
            return async_view.sdrf_file(request)

        try:
            # Read and parse SDRF file
            file_content = data["file"].read().decode("utf-8")
            lines = file_content.strip().split("\n")

            if not lines:
                return Response({"error": "Empty file"}, status=status.HTTP_400_BAD_REQUEST)

            # Parse headers and data
            headers = lines[0].split("\t")
            data_rows = [line.split("\t") for line in lines[1:]]

            # Check for pooled sample column and identify pool data (original CUPCAKE logic)
            pooled_column_index = None
            pooled_rows = []
            sn_rows = []

            for i, header in enumerate(headers):
                header_lower = header.lower()
                if "pooled sample" in header_lower or "pooled_sample" in header_lower:
                    pooled_column_index = i
                    break

            # If we found a pooled sample column, process the data
            if pooled_column_index is not None:
                for row_index, row in enumerate(data_rows):
                    if pooled_column_index < len(row):
                        pooled_value = row[pooled_column_index].strip()
                        if pooled_value.startswith("SN="):
                            sn_rows.append(row_index)
                        elif pooled_value.lower() == "pooled":
                            pooled_rows.append(row_index)

            # Get target metadata table
            try:
                metadata_table = MetadataTable.objects.get(id=data["metadata_table_id"])
            except MetadataTable.DoesNotExist:
                return Response(
                    {"error": "Invalid metadata_table_id"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if not metadata_table.can_edit(request.user):
                return Response(
                    {"error": "Permission denied: cannot edit this metadata table"},
                    status=status.HTTP_403_FORBIDDEN,
                )

            # Clear existing columns if replace_existing is True
            if data.get("replace_existing"):
                metadata_table.columns.all().delete()
                metadata_table.sample_pools.all().delete()

            # Remove SN= rows from data before general processing (original CUPCAKE logic)
            sn_data = []
            if pooled_column_index is not None and sn_rows:
                # Store SN= rows for later pool creation
                for row_index in sorted(sn_rows, reverse=True):  # Reverse order to maintain indices
                    if row_index < len(data_rows):
                        sn_data.append(data_rows[row_index])
                        del data_rows[row_index]

            # Update sample count and extend/truncate data if needed (original CUPCAKE logic)
            expected_sample_count = metadata_table.sample_count or len(data_rows)
            if len(data_rows) != expected_sample_count:
                # Extend the number of samples to match expected count and fill with empty strings
                if len(data_rows) < expected_sample_count:
                    data_rows.extend(
                        [["" for i in range(len(headers))] for j in range(expected_sample_count - len(data_rows))]
                    )
                else:
                    data_rows = data_rows[:expected_sample_count]

                    # If we removed SN= rows and truncated data, add SN= rows back for pool creation
                    if pooled_column_index is not None and sn_rows and sn_data:
                        # Add SN= rows back to the end of the truncated data
                        data_rows.extend(sn_data)
                        # Update sn_rows indices to reflect their new positions at the end
                        sn_rows = list(range(expected_sample_count, len(data_rows)))

            # Update metadata table sample count
            metadata_table.sample_count = len(data_rows)
            metadata_table.save(update_fields=["sample_count"])

            # Enhanced SDRF import with intelligent column matching and schema organization
            created_columns = []
            column_name_usage = {}  # Track how many times each name is used

            # Get the metadata table template that existing columns come from
            table_template = self._get_table_template(metadata_table)

            # Process headers with intelligent template matching
            for i, header in enumerate(headers):
                # Parse header format: type[name] (original CUPCAKE format)
                header_lower = header.lower()
                if "[" in header_lower and "]" in header_lower:
                    metadata_type = header_lower.split("[")[0].strip()
                    name = header_lower.strip()
                else:
                    # Handle "source name" specifically, everything else remains "special"
                    name = header_lower.strip()
                    if name == "source name":
                        metadata_type = "source_name"
                    else:
                        metadata_type = "special"
                # Count usage of this column name
                if name not in column_name_usage:
                    column_name_usage[name] = 0
                column_name_usage[name] += 1

                # Try to find matching column in existing table or template
                metadata_column = self._find_or_create_matching_column(
                    name, metadata_type, metadata_table, table_template, i, column_name_usage[name]
                )

                created_columns.append(metadata_column)

            # retrieve metadata columns after reordering
            for i, metadata_column in enumerate(created_columns):
                metadata_value_map = {}

                # Process each data row for this column
                for j, row in enumerate(data_rows):
                    if i < len(row) and row[i]:
                        cell_value = row[i].strip()

                        if cell_value == "":
                            continue

                        # Handle "not applicable" and "not available" as sample-specific values
                        if cell_value == "not applicable":
                            value = "not applicable"
                        elif cell_value == "not available":
                            value = "not available"
                        else:
                            value = metadata_column.convert_sdrf_to_metadata(cell_value)

                        if value not in metadata_value_map:
                            metadata_value_map[value] = []
                        metadata_value_map[value].append(j)

                # Set the most common value as default
                max_count = 0
                max_value = None
                for value in metadata_value_map:
                    if len(metadata_value_map[value]) > max_count:
                        max_count = len(metadata_value_map[value])
                        max_value = value

                if max_value:
                    metadata_column.value = max_value

                # Calculate modifiers for other values (exact original CUPCAKE logic)
                modifiers = []
                for value in metadata_value_map:
                    if value != max_value:
                        modifier = {"samples": [], "value": value}
                        # Sort from lowest to highest. Add samples index. For continuous samples, add range
                        samples = metadata_value_map[value]
                        samples.sort()
                        start = samples[0]
                        end = samples[0]
                        for i2 in range(1, len(samples)):
                            if samples[i2] == end + 1:
                                end = samples[i2]
                            else:
                                if start == end:
                                    modifier["samples"].append(str(start + 1))
                                else:
                                    modifier["samples"].append(f"{start + 1}-{end + 1}")
                                start = samples[i2]
                                end = samples[i2]
                        if start == end:
                            modifier["samples"].append(str(start + 1))
                        else:
                            modifier["samples"].append(f"{start + 1}-{end + 1}")
                        if len(modifier["samples"]) == 1:
                            modifier["samples"] = modifier["samples"][0]
                        else:
                            modifier["samples"] = ",".join(modifier["samples"])
                        modifiers.append(modifier)

                if modifiers:
                    # Store modifiers as native JSON object
                    metadata_column.modifiers = modifiers

                metadata_column.save()

            created_pools = []
            if data.get("create_pools") and pooled_column_index is not None:
                # Find source name column
                source_name_column_index = None
                for idx, header in enumerate(headers):
                    header_lower = header.lower()
                    if "source name" in header_lower or "source_name" in header_lower:
                        source_name_column_index = idx
                        break

                # Pool synchronization: track pools from import data (exact original CUPCAKE logic)
                import_pools_data = []

                if sn_rows and sn_data:
                    # Case 1: There are rows with SN= values - create pools from them
                    for pool_index, row in enumerate(sn_data):
                        if pooled_column_index < len(row):
                            sdrf_value = row[pooled_column_index].strip()

                            # Extract source names from SN= value
                            if sdrf_value.startswith("SN="):
                                source_names = sdrf_value[3:].split(",")
                                source_names = [name.strip() for name in source_names]

                                # Get pool name from source name column or use default
                                pool_name = (
                                    row[source_name_column_index]
                                    if source_name_column_index is not None and source_name_column_index < len(row)
                                    else f"Pool {pool_index + 1}"
                                )

                                # Find sample indices that match these source names
                                pooled_only_samples = []
                                pooled_and_independent_samples = []

                                for sample_index, sample_row in enumerate(data_rows):
                                    if source_name_column_index is not None and source_name_column_index < len(
                                        sample_row
                                    ):
                                        sample_source_name = sample_row[source_name_column_index].strip()
                                        if sample_source_name in source_names:
                                            # Check if this sample is also marked as "not pooled" or independent
                                            sample_pooled_value = ""
                                            if pooled_column_index < len(sample_row):
                                                sample_pooled_value = sample_row[pooled_column_index].strip().lower()

                                            if (
                                                sample_pooled_value == "not pooled"
                                                or sample_pooled_value == ""
                                                or sample_pooled_value == "independent"
                                            ):
                                                # Sample exists both in pool and as independent
                                                pooled_and_independent_samples.append(sample_index + 1)
                                            else:
                                                # Sample is only in pool
                                                pooled_only_samples.append(sample_index + 1)

                                # Store pool data for synchronization
                                import_pools_data.append(
                                    {
                                        "pool_name": pool_name,
                                        "pooled_only_samples": pooled_only_samples,
                                        "pooled_and_independent_samples": pooled_and_independent_samples,
                                        "is_reference": True,  # SN= pools are reference pools
                                        "metadata_row": row,
                                        "sdrf_value": sdrf_value,
                                        "all_data_rows": data_rows,
                                    }
                                )

                elif pooled_rows:
                    # Case 2: No SN= rows but there are "pooled" rows - create a pool from them
                    # Get source names of all pooled samples
                    pooled_source_names = []
                    pooled_only_samples = []

                    for row_index in pooled_rows:
                        if (
                            source_name_column_index is not None
                            and row_index < len(data_rows)
                            and source_name_column_index < len(data_rows[row_index])
                        ):
                            source_name = data_rows[row_index][source_name_column_index].strip()
                            if source_name:
                                pooled_source_names.append(source_name)
                                pooled_only_samples.append(row_index + 1)

                    if pooled_source_names:
                        # Create SN= value from source names
                        sdrf_value = "SN=" + ",".join(pooled_source_names)
                        pool_name = "Pool 1"
                        template_row = data_rows[pooled_rows[0]]

                        # Store pool data for synchronization
                        import_pools_data.append(
                            {
                                "pool_name": pool_name,
                                "pooled_only_samples": pooled_only_samples,
                                "pooled_and_independent_samples": [],
                                "is_reference": False,  # Pooled rows are not reference pools by default
                                "metadata_row": template_row,
                                "sdrf_value": sdrf_value,
                                "all_data_rows": data_rows,
                            }
                        )

                # Synchronize pools with sophisticated logic (matching original CUPCAKE)
                if import_pools_data:
                    synchronize_pools_with_import_data(metadata_table, import_pools_data, created_columns, request.user)
                    # Get the updated pools list for response
                    created_pools = list(metadata_table.sample_pools.all())

            if table_template:
                if table_template.user_columns:
                    schema_ids = set()
                    for template_column in table_template.user_columns.all():
                        if template_column.template:
                            if template_column.template.schema:
                                schema_ids.add(template_column.template.schema.id)
                    if schema_ids:
                        try:
                            metadata_table.reorder_columns_by_schema(schema_ids=list(schema_ids))
                        except Exception:
                            metadata_table.normalize_column_positions()
                    else:
                        # No schemas found, just normalize positions
                        metadata_table.normalize_column_positions()

                    # Apply the same reordering logic to sample pool columns
                    if created_pools:
                        for pool in created_pools:
                            if pool.metadata_columns.exists():
                                if schema_ids:
                                    try:
                                        pool.reorder_pool_columns_by_schema(schema_ids=list(schema_ids))
                                    except Exception:
                                        pool.basic_pool_column_reordering()
                                else:
                                    # No schemas found, use basic section-based ordering for pools
                                    pool.basic_pool_column_reordering()
            else:
                # No template available, just normalize positions
                metadata_table.normalize_column_positions()

                # Apply basic reordering to sample pool columns when no template available
                if created_pools:
                    for pool in created_pools:
                        if pool.metadata_columns.exists():
                            pool.basic_pool_column_reordering()

            # Include validation results in response
            response_data = {
                "message": "SDRF file imported successfully",
                "created_columns": len(created_columns),
                "created_pools": len(created_pools),
                "pools_detected": pooled_column_index is not None,
                "sn_rows_count": len(sn_rows),
                "pooled_rows_count": len(pooled_rows),
                "sample_rows": len(data_rows),
            }

            return Response(response_data)

        except Exception as e:
            return Response(
                {"error": f"Import failed: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser])
    def import_excel_file(self, request):
        """Import metadata from Excel file (matching original CUPCAKE)."""
        serializer = MetadataImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data

        # Check if async processing is requested
        if data.get("async_processing", False):
            # Queue async task
            from .async_views import AsyncImportViewSet

            async_view = AsyncImportViewSet()
            return async_view.excel_file(request)

        try:
            # Read Excel workbook
            file_content = data["file"].read()
            wb = load_workbook(io.BytesIO(file_content))

            # Get main worksheet
            if "main" not in wb.sheetnames:
                return Response(
                    {"error": "Excel file must contain a 'main' worksheet"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            main_ws = wb["main"]
            main_headers = [cell.value for cell in main_ws[1]]
            main_data = [list(row) for row in main_ws.iter_rows(min_row=2, values_only=True)]

            # Get hidden worksheet if exists
            hidden_headers = []
            hidden_data = []
            if "hidden" in wb.sheetnames:
                hidden_ws = wb["hidden"]
                if hidden_ws.max_row > 1:
                    hidden_headers = [cell.value for cell in hidden_ws[1]]
                    hidden_data = [list(row) for row in hidden_ws.iter_rows(min_row=2, values_only=True)]

            # Get ID metadata column mapping
            if "id_metadata_column_map" not in wb.sheetnames:
                return Response(
                    {"error": "Excel file must contain an 'id_metadata_column_map' worksheet"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            id_metadata_column_map_ws = wb["id_metadata_column_map"]
            id_metadata_column_map_list = [
                list(row) for row in id_metadata_column_map_ws.iter_rows(min_row=2, values_only=True)
            ]
            id_metadata_column_map = {}
            for row in id_metadata_column_map_list:
                if row[0] is not None:
                    id_metadata_column_map[int(row[0])] = {
                        "column": row[1],
                        "name": row[2],
                        "type": row[3],
                        "hidden": row[4],
                    }

            # Check for pool sheets (original CUPCAKE logic)
            # Pool data extraction - simplified in ccv
            # pool_main_headers, pool_main_data = [], []
            # pool_hidden_headers, pool_hidden_data = [], []
            pool_id_metadata_column_map = {}
            pool_object_map_data = []

            if "pool_main" in wb.sheetnames:
                pool_main_ws = wb["pool_main"]
                if pool_main_ws.max_row > 1:
                    # pool_main_headers = [cell.value for cell in pool_main_ws[1]]
                    # pool_main_data = [
                    #     list(row)
                    #     for row in pool_main_ws.iter_rows(min_row=2, values_only=True)
                    # ]
                    pass

            if "pool_hidden" in wb.sheetnames:
                pool_hidden_ws = wb["pool_hidden"]
                if pool_hidden_ws.max_row > 1:
                    # pool_hidden_headers = [cell.value for cell in pool_hidden_ws[1]]
                    # pool_hidden_data = [
                    #     list(row)
                    #     for row in pool_hidden_ws.iter_rows(min_row=2, values_only=True)
                    # ]
                    pass

            if "pool_id_metadata_column_map" in wb.sheetnames:
                pool_id_metadata_column_map_ws = wb["pool_id_metadata_column_map"]
                pool_id_metadata_column_map_list = [
                    list(row) for row in pool_id_metadata_column_map_ws.iter_rows(min_row=2, values_only=True)
                ]
                for row in pool_id_metadata_column_map_list:
                    if row[0] is not None:
                        pool_id_metadata_column_map[int(row[0])] = {
                            "column": row[1],
                            "name": row[2],
                            "type": row[3],
                            "hidden": row[4],
                        }

            if "pool_object_map" in wb.sheetnames:
                pool_object_map_ws = wb["pool_object_map"]
                if pool_object_map_ws.max_row > 1:
                    pool_object_map_data = [
                        list(row) for row in pool_object_map_ws.iter_rows(min_row=2, values_only=True)
                    ]

            # Get target metadata table
            try:
                metadata_table = MetadataTable.objects.get(id=data["metadata_table_id"])
            except MetadataTable.DoesNotExist:
                return Response(
                    {"error": "Invalid metadata_table_id"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

                if not metadata_table.can_edit(request.user):
                    return Response(
                        {"error": "Permission denied: cannot edit this metadata table"},
                        status=status.HTTP_403_FORBIDDEN,
                    )

            # Clear existing columns if replace_existing is True
            if data.get("replace_existing"):
                metadata_table.columns.all().delete()
                metadata_table.sample_pools.all().delete()

            # Combine main and hidden data (original CUPCAKE logic)
            if hidden_data:
                headers = main_headers + hidden_headers
                combined_data = [main_row + hidden_row for main_row, hidden_row in zip(main_data, hidden_data)]
            else:
                headers = main_headers
                combined_data = main_data

            # Update sample count and extend/truncate data if needed (original CUPCAKE logic)
            expected_sample_count = metadata_table.sample_count or len(combined_data)
            if len(combined_data) != expected_sample_count:
                if len(combined_data) < expected_sample_count:
                    combined_data.extend(
                        [["" for i in range(len(headers))] for j in range(expected_sample_count - len(combined_data))]
                    )
                else:
                    combined_data = combined_data[:expected_sample_count]

            # Update metadata table sample count
            metadata_table.sample_count = len(combined_data)
            metadata_table.save(update_fields=["sample_count"])

            # Create metadata columns from main headers (original CUPCAKE logic)
            created_columns = []
            for n, header in enumerate(main_headers):
                # Find metadata column ID from mapping
                id_from_map = 0
                for map_id, map_data in id_metadata_column_map.items():
                    if map_data["column"] == n and not map_data["hidden"]:
                        id_from_map = map_id
                        break

                # Try to get existing column or create new one
                if id_from_map > 0:
                    try:
                        metadata_column = MetadataColumn.objects.get(id=id_from_map)
                    except MetadataColumn.DoesNotExist:
                        metadata_column = self._create_column_from_header(header, False, metadata_table, n)
                else:
                    metadata_column = self._create_column_from_header(header, False, metadata_table, n)

                # Note: Ontology mapping should only be applied on user request, not automatically
                created_columns.append(metadata_column)

            # Create metadata columns from hidden headers (original CUPCAKE logic)
            for n, header in enumerate(hidden_headers):
                # Find metadata column ID from mapping
                id_from_map = 0
                for map_id, map_data in id_metadata_column_map.items():
                    if map_data["column"] == n and map_data["hidden"]:
                        id_from_map = map_id
                        break

                # Try to get existing column or create new one
                if id_from_map > 0:
                    try:
                        metadata_column = MetadataColumn.objects.get(id=id_from_map)
                    except MetadataColumn.DoesNotExist:
                        metadata_column = self._create_column_from_header(
                            header, True, metadata_table, len(main_headers) + n
                        )
                else:
                    metadata_column = self._create_column_from_header(
                        header, True, metadata_table, len(main_headers) + n
                    )

                # Note: Ontology mapping should only be applied on user request, not automatically
                created_columns.append(metadata_column)

            # Process data to populate column values and modifiers (exact original CUPCAKE logic)
            for i, metadata_column in enumerate(created_columns):
                metadata_value_map = {}

                # Process each data row for this column
                for j, row in enumerate(combined_data):
                    if i < len(row) and row[i]:
                        cell_value = str(row[i]).strip()

                        if cell_value == "":
                            continue

                        if cell_value == "not applicable":
                            value = "not applicable"
                        elif cell_value == "not available":
                            value = "not available"
                        else:
                            # Use MetadataColumn instance method for proper SDRF conversion with ontology lookup
                            value = metadata_column.convert_sdrf_to_metadata(cell_value)

                        if value not in metadata_value_map:
                            metadata_value_map[value] = []
                        metadata_value_map[value].append(j)

                max_count = 0
                max_value = None
                for value in metadata_value_map:
                    if len(metadata_value_map[value]) > max_count:
                        max_count = len(metadata_value_map[value])
                        max_value = value

                if max_value:
                    metadata_column.value = max_value

                modifiers = []
                for value in metadata_value_map:
                    if value != max_value:
                        modifier = {"samples": [], "value": value}
                        samples = metadata_value_map[value]
                        samples.sort()
                        start = samples[0]
                        end = samples[0]
                        for i2 in range(1, len(samples)):
                            if samples[i2] == end + 1:
                                end = samples[i2]
                            else:
                                if start == end:
                                    modifier["samples"].append(str(start + 1))
                                else:
                                    modifier["samples"].append(f"{start + 1}-{end + 1}")
                                start = samples[i2]
                                end = samples[i2]
                        if start == end:
                            modifier["samples"].append(str(start + 1))
                        else:
                            modifier["samples"].append(f"{start + 1}-{end + 1}")
                        if len(modifier["samples"]) == 1:
                            modifier["samples"] = modifier["samples"][0]
                        else:
                            modifier["samples"] = ",".join(modifier["samples"])
                        modifiers.append(modifier)

                if modifiers:
                    # Store modifiers as native JSON object
                    metadata_column.modifiers = modifiers

                metadata_column.save()

            created_pools = []
            if pool_object_map_data and data.get("create_pools"):
                for pool_row in pool_object_map_data:
                    pool_name = pool_row[0] if pool_row[0] else f"Pool {len(created_pools) + 1}"
                    pooled_only_samples = json.loads(pool_row[1]) if pool_row[1] else []
                    pooled_and_independent_samples = json.loads(pool_row[2]) if pool_row[2] else []
                    is_reference = pool_row[3] if len(pool_row) > 3 else False
                    sample_pool = SamplePool.objects.create(
                        metadata_table=metadata_table,
                        pool_name=pool_name,
                        pooled_only_samples=pooled_only_samples,
                        pooled_and_independent_samples=pooled_and_independent_samples,
                        is_reference=is_reference,
                        created_by=request.user,
                    )

                    from .utils import create_pool_metadata_from_table_columns

                    create_pool_metadata_from_table_columns(sample_pool)

                    created_pools.append(sample_pool)

            response_data = {
                "message": "Excel file imported successfully",
                "created_columns": len(created_columns),
                "created_pools": len(created_pools),
                "sample_rows": len(combined_data),
                "has_hidden_data": len(hidden_data) > 0,
                "has_pool_data": len(pool_object_map_data) > 0,
            }

            return Response(response_data)

        except Exception as e:
            return Response(
                {"error": f"Excel import failed: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    def _create_column_from_header(self, header, is_hidden, metadata_table, position):
        """Helper method to create MetadataColumn from header (original CUPCAKE logic)."""
        header = header.lower() if header else ""

        # Parse SDRF-style headers with brackets (e.g., "factor value[genetic modification]")
        if "[" in header and header.endswith("]"):
            # This is an SDRF-style header like "factor value[genetic modification]"
            name = header
            base_part = header.split("[")[0].strip()

            # Determine type based on SDRF conventions
            if base_part.startswith("characteristics"):
                metadata_type = "characteristics"
            elif base_part.startswith("factor value"):
                metadata_type = "factor value"
            elif base_part.startswith("comment"):
                metadata_type = "comment"
            else:
                metadata_type = "special"
        else:
            # Plain header name - handle "source name" specifically, everything else is "special"
            name = header
            header_lower = header.lower().strip()
            if header_lower == "source name":
                metadata_type = "source_name"
            else:
                metadata_type = "special"

        # Create metadata column with proper naming (match original CUPCAKE)
        metadata_column = MetadataColumn.objects.create(
            name=name.capitalize().replace("Ms1", "MS1").replace("Ms2", "MS2"),
            type=metadata_type.capitalize(),
            column_position=position,
            metadata_table=metadata_table,
            hidden=is_hidden,
            readonly=False,
        )

        return metadata_column

    @action(detail=False, methods=["get"])
    def collect_metadata(self, request):
        """Collect metadata from metadata table."""
        serializer = MetadataCollectionSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data

        try:
            metadata_table = MetadataTable.objects.get(id=data["metadata_table_id"])
        except MetadataTable.DoesNotExist:
            return Response(
                {"error": "Invalid metadata_table_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Get metadata columns for the specified table
        metadata_columns = MetadataColumn.objects.filter(metadata_table=metadata_table)

        # Filter by metadata names if specified
        if data.get("metadata_names"):
            metadata_columns = metadata_columns.filter(name__in=data["metadata_names"])

        if data.get("unique_values_only"):
            # Return unique values grouped by metadata name
            unique_metadata = {}
            for column in metadata_columns:
                name = column.name.lower()
                if name not in unique_metadata:
                    unique_metadata[name] = {
                        "name": column.name,
                        "types": {},
                        "unique_values": set(),
                        "value_count": 0,
                    }

                if column.type not in unique_metadata[name]["types"]:
                    unique_metadata[name]["types"][column.type] = 0
                unique_metadata[name]["types"][column.type] += 1

                if column.value:
                    unique_metadata[name]["unique_values"].add(column.value)
                    unique_metadata[name]["value_count"] += 1

            # Convert sets to lists for JSON serialization
            for name_data in unique_metadata.values():
                name_data["unique_values"] = sorted(list(name_data["unique_values"]))

            return Response(
                {
                    "metadata_table_id": data["metadata_table_id"],
                    "metadata_table_name": metadata_table.name,
                    "unique_values_only": True,
                    "metadata_columns": unique_metadata,
                    "total_unique_values": sum(len(md["unique_values"]) for md in unique_metadata.values()),
                }
            )
        else:
            # Return full metadata column details
            serializer = MetadataColumnSerializer(metadata_columns, many=True)
            return Response(
                {
                    "metadata_table_id": data["metadata_table_id"],
                    "metadata_table_name": metadata_table.name,
                    "unique_values_only": False,
                    "metadata_columns": serializer.data,
                    "total_columns": metadata_columns.count(),
                }
            )

    @action(detail=False, methods=["post"])
    def export_sdrf_file(self, request):
        """Export metadata as SDRF file (matching original CUPCAKE)."""
        serializer = MetadataExportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data

        # Get metadata table
        try:
            metadata_table = MetadataTable.objects.get(id=data["metadata_table_id"])
        except (MetadataTable.DoesNotExist, KeyError):
            return Response(
                {"error": "metadata_table_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check if async processing is requested
        if data.get("async_processing", False):
            # Queue async task
            from .async_views import AsyncExportViewSet

            async_view = AsyncExportViewSet()
            return async_view.sdrf_file(request)

        if not metadata_table.can_view(request.user):
            return Response(
                {"error": "Permission denied: cannot view this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Get metadata columns (user can specify specific columns or get all)
        if data.get("metadata_column_ids"):
            metadata_columns = metadata_table.columns.filter(id__in=data["metadata_column_ids"])
        else:
            metadata_columns = metadata_table.columns.all()

        # Filter out hidden columns for SDRF export (SDRF standard doesn't support hidden columns)
        visible_metadata = metadata_columns.filter(hidden=False)

        # Sort metadata and get structured output for SDRF format
        result_data, _ = sort_metadata(list(visible_metadata), metadata_table.sample_count, metadata_table)

        # Add pool data if requested and pools exist (original CUPCAKE logic)
        if data.get("include_pools", True):
            pools = list(metadata_table.sample_pools.all())
            if pools:
                # Add SN= rows for reference pools (original CUPCAKE approach)
                for pool in pools:
                    if pool.is_reference and pool.sdrf_value:
                        # Create a pool row following original CUPCAKE format
                        pool_row = ["" for _ in result_data[0]]  # Initialize with empty values

                        # Find the pooled sample column index
                        pooled_column_index = None
                        for i, header in enumerate(result_data[0]):
                            header_lower = header.lower()
                            if "pooled sample" in header_lower or "pooled_sample" in header_lower:
                                pooled_column_index = i
                                break

                        # Find the source name column index
                        source_name_column_index = None
                        for i, header in enumerate(result_data[0]):
                            header_lower = header.lower()
                            if "source name" in header_lower or "source_name" in header_lower:
                                source_name_column_index = i
                                break

                        # Set pool data in the row
                        if pooled_column_index is not None:
                            pool_row[pooled_column_index] = pool.sdrf_value

                        if source_name_column_index is not None:
                            pool_row[source_name_column_index] = pool.pool_name

                        # Fill other columns with pool-specific default values
                        for i, metadata_column in enumerate(visible_metadata):
                            if i < len(pool_row) and not pool_row[i]:  # Only fill empty cells
                                # Use column default value or check column flags for appropriate empty value
                                if metadata_column.value:
                                    pool_row[i] = metadata_column.value
                                elif metadata_column.not_applicable:
                                    pool_row[i] = "not applicable"
                                elif metadata_column.not_available:
                                    pool_row[i] = "not available"

                        # Add pool row to results
                        result_data.append(pool_row)

        # Convert to tab-separated format (SDRF standard)
        sdrf_content = []
        for row in result_data:
            # Convert all values to strings and handle None values
            str_row = [str(cell) if cell is not None else "" for cell in row]
            sdrf_content.append("\t".join(str_row))

        sdrf_text = "\n".join(sdrf_content)

        # Validate SDRF format if validation was requested
        validation_results = None
        if data.get("validate_sdrf", False):
            try:
                validation_results = validate_sdrf(result_data)
            except Exception as e:
                validation_results = {
                    "errors": [f"SDRF validation failed: {str(e)}"],
                    "warnings": [],
                    "suggestions": [],
                }

        # Return as downloadable file
        response = HttpResponse(sdrf_text, content_type="text/tab-separated-values")

        # Create filename based on metadata table name
        safe_filename = "".join(c for c in metadata_table.name if c.isalnum() or c in (" ", "-", "_")).rstrip()
        safe_filename = safe_filename.replace(" ", "_")
        response["Content-Disposition"] = f'attachment; filename="{safe_filename}_metadata.sdrf"'

        # Add validation results as headers if available
        if validation_results:
            response["X-SDRF-Validation-Errors"] = str(len(validation_results.get("errors", [])))
            response["X-SDRF-Validation-Warnings"] = str(len(validation_results.get("warnings", [])))
            if validation_results.get("errors"):
                # Include first few errors in header (truncated for header length limits)
                error_summary = "; ".join(validation_results["errors"][:3])
                if len(error_summary) > 200:
                    error_summary = error_summary[:197] + "..."
                response["X-SDRF-Validation-Error-Summary"] = error_summary

        return response


# ===================================================================
# ONTOLOGY AND CONTROLLED VOCABULARY VIEWSETS
# ===================================================================


class SpeciesViewSet(FilterMixin, viewsets.ReadOnlyModelViewSet):
    """ViewSet for Species controlled vocabulary (read-only)."""

    queryset = Species.objects.all()
    serializer_class = SpeciesSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ["official_name", "common_name", "synonym"]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Filter by species code
        code = self.request.query_params.get("code")
        if code:
            queryset = queryset.filter(code__iexact=code)

        # Filter by taxon
        taxon = self.request.query_params.get("taxon")
        if taxon:
            queryset = queryset.filter(taxon=taxon)

        return queryset.order_by("official_name")


class TissueViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for Tissue controlled vocabulary (read-only)."""

    queryset = Tissue.objects.all()
    serializer_class = TissueSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Search by identifier or accession
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                models.Q(identifier__icontains=search)
                | models.Q(accession__icontains=search)
                | models.Q(synonyms__icontains=search)
            )

        return queryset.order_by("identifier")


class HumanDiseaseViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for HumanDisease controlled vocabulary (read-only)."""

    queryset = HumanDisease.objects.all()
    serializer_class = HumanDiseaseSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Search by identifier, accession, or definition
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                models.Q(identifier__icontains=search)
                | models.Q(accession__icontains=search)
                | models.Q(definition__icontains=search)
                | models.Q(synonyms__icontains=search)
                | models.Q(acronym__icontains=search)
            )

        return queryset.order_by("identifier")


class SubcellularLocationViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for SubcellularLocation controlled vocabulary (read-only)."""

    queryset = SubcellularLocation.objects.all()
    serializer_class = SubcellularLocationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Search by various fields
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                models.Q(accession__icontains=search)
                | models.Q(location_identifier__icontains=search)
                | models.Q(definition__icontains=search)
                | models.Q(synonyms__icontains=search)
            )

        return queryset.order_by("accession")


class MSUniqueVocabulariesViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for MSUniqueVocabularies controlled vocabulary (read-only)."""

    queryset = MSUniqueVocabularies.objects.all()
    serializer_class = MSUniqueVocabulariesSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Filter by term type
        term_type = self.request.query_params.get("term_type")
        if term_type:
            queryset = queryset.filter(term_type__iexact=term_type)

        # Search by accession, name, or definition
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                models.Q(accession__icontains=search)
                | models.Q(name__icontains=search)
                | models.Q(definition__icontains=search)
            )

        return queryset.order_by("accession")


class UnimodViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for Unimod controlled vocabulary (read-only)."""

    queryset = Unimod.objects.all()
    serializer_class = UnimodSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter queryset based on query parameters."""
        queryset = super().get_queryset()

        # Search by accession, name, or definition
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                models.Q(accession__icontains=search)
                | models.Q(name__icontains=search)
                | models.Q(definition__icontains=search)
            )

        return queryset.order_by("accession")


# ===================================================================
# METADATA COLUMN TEMPLATE VIEWSETS
# ===================================================================


class MetadataColumnTemplateViewSet(FilterMixin, viewsets.ModelViewSet):
    """ViewSet for managing MetadataColumnTemplate objects."""

    queryset = MetadataColumnTemplate.objects.all()
    serializer_class = MetadataColumnTemplateSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ["name", "description", "column_name", "tags"]
    filterset_fields = ["visibility", "owner", "lab_group", "is_active"]

    def get_queryset(self):
        """Filter queryset based on visibility and user permissions (includes bubble-up from sub-groups)."""
        user = self.request.user
        queryset = super().get_queryset()

        from ccc.models import LabGroup

        accessible_groups = LabGroup.get_accessible_group_ids(user)

        queryset = queryset.filter(
            models.Q(visibility="global")
            | models.Q(visibility="public")
            | models.Q(visibility="private", owner=user)
            | models.Q(visibility="lab_group", lab_group_id__in=accessible_groups)
            | models.Q(shared_with_users=user)
        ).distinct()

        # Apply additional filters
        visibility = self.request.query_params.get("visibility")
        if visibility:
            queryset = queryset.filter(visibility=visibility)

        owner_id = self.request.query_params.get("owner_id")
        if owner_id:
            queryset = queryset.filter(owner_id=owner_id)

        lab_group_id = self.request.query_params.get("lab_group_id")
        if lab_group_id:
            queryset = queryset.filter(lab_group_id=lab_group_id)

        ontology_type = self.request.query_params.get("ontology_type")
        if ontology_type:
            queryset = queryset.filter(ontology_type=ontology_type)

        category = self.request.query_params.get("category")
        if category:
            queryset = queryset.filter(category__icontains=category)

        is_active = self.request.query_params.get("is_active")
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == "true")

        schema_id = self.request.query_params.get("schema_id")
        if schema_id:
            queryset = queryset.filter(schema_id=schema_id)

        source_schema = self.request.query_params.get("source_schema")
        if source_schema:
            queryset = queryset.filter(source_schema__icontains=source_schema)

        return queryset.order_by("-usage_count", "name")

    def perform_create(self, serializer):
        """Ensure owner is set when creating templates."""
        serializer.save(owner=self.request.user)

    def perform_update(self, serializer):
        """Check permissions before updating."""
        template = self.get_object()
        if not template.can_edit(self.request.user):
            raise PermissionDenied("You don't have permission to edit this template")
        serializer.save()

    def perform_destroy(self, instance):
        """Check permissions before deleting."""
        if not instance.can_delete(self.request.user):
            raise PermissionDenied("You don't have permission to delete this template")
        super().perform_destroy(instance)

    @action(detail=True, methods=["post"])
    def create_metadata_column(self, request, pk=None):
        """Create a MetadataColumn from this template."""
        template = self.get_object()

        # Check if user can view this template
        if not template.can_view(request.user):
            return Response(
                {"error": "You don't have permission to use this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        metadata_table_id = request.data.get("metadata_table_id")
        position = request.data.get("position")

        if not metadata_table_id:
            return Response(
                {"error": "metadata_table_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            metadata_table = MetadataTable.objects.get(id=metadata_table_id)
        except MetadataTable.DoesNotExist:
            return Response(
                {"error": "Invalid metadata_table_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check if user can edit the metadata table
        if not metadata_table.can_edit(request.user):
            return Response(
                {"error": "You don't have permission to edit this metadata table"},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Create the metadata column
        try:
            column = template.create_metadata_column(metadata_table, position)

            return Response(
                {
                    "message": "Metadata column created successfully from template",
                    "column_id": column.id,
                    "template_usage_count": template.usage_count,
                },
                status=status.HTTP_201_CREATED,
            )
        except Exception as e:
            return Response(
                {"error": f"Failed to create column: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=True, methods=["post"])
    def share_template(self, request, pk=None):
        """Share this template with another user."""
        template = self.get_object()

        # Check if user can edit this template (only creators/staff can share)
        if not template.can_edit(request.user):
            return Response(
                {"error": "You don't have permission to share this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        user_id = request.data.get("user_id")
        permission_level = request.data.get("permission_level", "use")

        if not user_id:
            return Response(
                {"error": "user_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            target_user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response(
                {"error": "Invalid user_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Create or update share
        share, created = MetadataColumnTemplateShare.objects.update_or_create(
            template=template,
            user=target_user,
            defaults={
                "shared_by": request.user,
                "permission_level": permission_level,
            },
        )

        action = "created" if created else "updated"
        return Response(
            {
                "message": f"Template share {action} successfully",
                "share_id": share.id,
                "user": target_user.username,
                "permission_level": permission_level,
            }
        )

    @action(detail=True, methods=["delete"])
    def unshare_template(self, request, pk=None):
        """Remove template sharing for a user."""
        template = self.get_object()

        # Check if user can edit this template
        if not template.can_edit(request.user):
            return Response(
                {"error": "You don't have permission to manage sharing for this template"},
                status=status.HTTP_403_FORBIDDEN,
            )

        user_id = request.data.get("user_id")
        if not user_id:
            return Response(
                {"error": "user_id is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            share = MetadataColumnTemplateShare.objects.get(template=template, user_id=user_id)
            share.delete()
            return Response({"message": "Template sharing removed successfully"})
        except MetadataColumnTemplateShare.DoesNotExist:
            return Response(
                {"error": "Template share not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

    @action(detail=False, methods=["get"])
    def my_templates(self, request):
        """Get templates created by the current user."""
        templates = self.get_queryset().filter(owner=request.user)

        # Apply search if provided
        search = request.query_params.get("search")
        if search:
            templates = templates.filter(
                models.Q(name__icontains=search)
                | models.Q(description__icontains=search)
                | models.Q(column_name__icontains=search)
            )

        serializer = self.get_serializer(templates, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def popular_templates(self, request):
        """Get most popular public templates."""
        try:
            limit = int(request.query_params.get("limit", 10))
            limit = max(0, limit)  # Ensure non-negative
        except (ValueError, TypeError):
            limit = 10  # Default fallback
        templates = (
            self.get_queryset()
            .filter(visibility__in=["public", "global"], is_active=True)
            .order_by("-usage_count")[:limit]
        )

        serializer = self.get_serializer(templates, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def ontology_suggestions(self, request):
        """Get ontology suggestions for this column template."""
        # Import serializer locally to avoid circular import issues

        # Get template ID from query params
        template_id = request.query_params.get("template_id")
        if not template_id:
            return Response(
                {"error": "template_id parameter is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Get template directly without the viewset's filtered queryset
        try:
            template = MetadataColumnTemplate.objects.get(pk=template_id)
            # Check if user can view this template
            if not template.can_view(request.user):
                return Response(
                    {"error": "You don't have permission to access this template"},
                    status=status.HTTP_403_FORBIDDEN,
                )
        except MetadataColumnTemplate.DoesNotExist:
            return Response(
                {"error": "Template not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

        search_term = request.query_params.get("search", "")
        limit = min(int(request.query_params.get("limit", 20)), 100)  # Cap at 100
        search_type = request.query_params.get("search_type", "icontains")

        # Validate search_type
        valid_search_types = ["icontains", "istartswith", "exact"]
        if search_type not in valid_search_types:
            return Response(
                {"error": f"Invalid search_type. Must be one of: {', '.join(valid_search_types)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not template.ontology_type:
            return Response(
                {
                    "suggestions": [],
                    "has_more": False,
                    "total_count": 0,
                    "message": "No ontology type specified for this template",
                }
            )

        try:
            # Use the template's own method for getting suggestions
            suggestions = template.get_ontology_suggestions(search_term, limit, search_type)
            total_count = len(suggestions)

            # Serialize suggestions with enhanced data
            serializer = OntologySuggestionSerializer(
                suggestions, many=True, context={"ontology_type": template.ontology_type}
            )

            return Response(
                {
                    "suggestions": serializer.data,
                    "has_more": len(suggestions) == limit,  # Template method doesn't provide total count
                    "total_count": total_count,
                    "ontology_type": template.ontology_type,
                    "search_term": search_term,
                    "search_type": search_type,
                }
            )

        except Exception as e:
            return Response(
                {"error": f"Failed to fetch ontology suggestions: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class MetadataColumnTemplateShareViewSet(FilterMixin, viewsets.ModelViewSet):
    """ViewSet for managing MetadataColumnTemplateShare objects."""

    queryset = MetadataColumnTemplateShare.objects.all()
    serializer_class = MetadataColumnTemplateShareSerializer
    permission_classes = [IsAuthenticated]
    search_fields = ["template__name", "template__description"]
    filterset_fields = ["template", "shared_by", "shared_with", "permission_level"]

    def get_queryset(self):
        """Filter queryset based on user permissions."""
        user = self.request.user
        queryset = super().get_queryset()

        # Users can only see shares they created, received, or are staff
        if not user.is_staff:
            queryset = queryset.filter(models.Q(shared_by=user) | models.Q(user=user))

        # Filter by template
        template_id = self.request.query_params.get("template_id")
        if template_id:
            queryset = queryset.filter(template_id=template_id)

        # Filter by permission level
        permission_level = self.request.query_params.get("permission_level")
        if permission_level:
            queryset = queryset.filter(permission_level=permission_level)

        return queryset.order_by("-shared_at")

    def perform_create(self, serializer):
        """Ensure shared_by is set when creating shares."""
        template = serializer.validated_data["template"]

        # Check if user can edit the template
        if not template.can_edit(self.request.user):
            raise PermissionDenied("You don't have permission to share this template")

        serializer.save(shared_by=self.request.user)

    def perform_update(self, serializer):
        """Check permissions before updating shares."""
        share = self.get_object()
        if share.shared_by != self.request.user and not self.request.user.is_staff:
            raise PermissionDenied("You can only update shares you created")
        serializer.save()

    def perform_destroy(self, instance):
        """Check permissions before deleting shares."""
        if (
            instance.shared_by != self.request.user
            and instance.user != self.request.user
            and not self.request.user.is_staff
        ):
            raise PermissionDenied("You can only delete shares you created or received")
        super().perform_destroy(instance)


# ===================================================================
# COMPREHENSIVE ONTOLOGY VIEWSETS FOR SDRF VALIDATION
# ===================================================================


class MondoDiseaseViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for MONDO Disease Ontology."""

    queryset = MondoDisease.objects.filter(obsolete=False)
    serializer_class = MondoDiseaseSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["name", "identifier"]
    search_fields = ["name", "definition", "synonyms", "identifier"]
    ordering_fields = ["name", "identifier", "created_at"]
    ordering = ["name"]

    @action(detail=False, methods=["get"])
    def search_suggestions(self, request):
        """Get disease suggestions for SDRF validation."""
        query = request.GET.get("q", "").strip()
        limit = int(request.GET.get("limit", 20))

        if not query or len(query) < 2:
            return Response({"results": []})

        # Search by name and synonyms
        queryset = self.get_queryset().filter(models.Q(name__icontains=query) | models.Q(synonyms__icontains=query))[
            :limit
        ]

        suggestions = []
        for disease in queryset:
            suggestions.append(
                {
                    "identifier": disease.identifier,
                    "name": disease.name,
                    "definition": disease.definition or "",
                    "synonyms": [s.strip() for s in disease.synonyms.split(";") if s.strip()]
                    if disease.synonyms
                    else [],
                    "source": "mondo",
                    "match_type": "fuzzy",
                }
            )

        return Response({"results": suggestions})


class UberonAnatomyViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for UBERON Anatomy Ontology."""

    queryset = UberonAnatomy.objects.filter(obsolete=False)
    serializer_class = UberonAnatomySerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["name", "identifier"]
    search_fields = ["name", "definition", "synonyms", "identifier"]
    ordering_fields = ["name", "identifier", "created_at"]
    ordering = ["name"]

    @action(detail=False, methods=["get"])
    def search_suggestions(self, request):
        """Get anatomy/tissue suggestions for SDRF validation."""
        query = request.GET.get("q", "").strip()
        limit = int(request.GET.get("limit", 20))

        if not query or len(query) < 2:
            return Response({"results": []})

        # Search by name and synonyms
        queryset = self.get_queryset().filter(models.Q(name__icontains=query) | models.Q(synonyms__icontains=query))[
            :limit
        ]

        suggestions = []
        for anatomy in queryset:
            suggestions.append(
                {
                    "identifier": anatomy.identifier,
                    "name": anatomy.name,
                    "definition": anatomy.definition or "",
                    "synonyms": [s.strip() for s in anatomy.synonyms.split(";") if s.strip()]
                    if anatomy.synonyms
                    else [],
                    "source": "uberon",
                    "match_type": "fuzzy",
                }
            )

        return Response({"results": suggestions})


class NCBITaxonomyViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for NCBI Taxonomy."""

    queryset = NCBITaxonomy.objects.all()
    serializer_class = NCBITaxonomySerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["scientific_name", "common_name", "rank", "tax_id"]
    search_fields = ["scientific_name", "common_name", "synonyms"]
    ordering_fields = ["scientific_name", "tax_id", "rank"]
    ordering = ["scientific_name"]

    @action(detail=False, methods=["get"])
    def search_suggestions(self, request):
        """Get organism suggestions for SDRF validation."""
        query = request.GET.get("q", "").strip()
        limit = int(request.GET.get("limit", 20))

        if not query or len(query) < 2:
            return Response({"results": []})

        # Search by scientific name, common name, and synonyms
        queryset = self.get_queryset().filter(
            models.Q(scientific_name__icontains=query)
            | models.Q(common_name__icontains=query)
            | models.Q(synonyms__icontains=query)
        )[:limit]

        suggestions = []
        for taxonomy in queryset:
            suggestions.append(
                {
                    "identifier": str(taxonomy.tax_id),
                    "name": taxonomy.scientific_name,
                    "definition": f"Rank: {taxonomy.rank}"
                    + (f", Common: {taxonomy.common_name}" if taxonomy.common_name else ""),
                    "synonyms": [s.strip() for s in taxonomy.synonyms.split(";") if s.strip()]
                    if taxonomy.synonyms
                    else [],
                    "source": "ncbitaxon",
                    "match_type": "fuzzy",
                }
            )

        return Response({"results": suggestions})


class ChEBICompoundViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for ChEBI Chemical Compounds."""

    queryset = ChEBICompound.objects.filter(obsolete=False)
    serializer_class = ChEBICompoundSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["name", "identifier", "formula"]
    search_fields = ["name", "definition", "synonyms", "identifier", "formula"]
    ordering_fields = ["name", "identifier", "mass"]
    ordering = ["name"]

    @action(detail=False, methods=["get"])
    def search_suggestions(self, request):
        """Get chemical compound suggestions for SDRF validation."""
        query = request.GET.get("q", "").strip()
        limit = int(request.GET.get("limit", 20))

        if not query or len(query) < 2:
            return Response({"results": []})

        # Search by name, synonyms, and formula
        queryset = self.get_queryset().filter(
            models.Q(name__icontains=query) | models.Q(synonyms__icontains=query) | models.Q(formula__icontains=query)
        )[:limit]

        suggestions = []
        for compound in queryset:
            suggestions.append(
                {
                    "identifier": compound.identifier,
                    "name": compound.name,
                    "definition": compound.definition or "",
                    "synonyms": [s.strip() for s in compound.synonyms.split(";") if s.strip()]
                    if compound.synonyms
                    else [],
                    "source": "chebi",
                    "match_type": "fuzzy",
                }
            )

        return Response({"results": suggestions})


class PSIMSOntologyViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for PSI-MS Ontology."""

    queryset = PSIMSOntology.objects.filter(obsolete=False)
    serializer_class = PSIMSOntologySerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["name", "identifier", "category"]
    search_fields = ["name", "definition", "synonyms", "identifier"]
    ordering_fields = ["name", "identifier", "category"]
    ordering = ["name"]

    @action(detail=False, methods=["get"])
    def search_suggestions(self, request):
        """Get mass spectrometry term suggestions for SDRF validation."""
        query = request.GET.get("q", "").strip()
        category = request.GET.get("category", "").strip()
        limit = int(request.GET.get("limit", 20))

        if not query or len(query) < 2:
            return Response({"results": []})

        # Search by name and synonyms
        queryset = self.get_queryset().filter(models.Q(name__icontains=query) | models.Q(synonyms__icontains=query))

        # Filter by category if provided
        if category:
            queryset = queryset.filter(category=category)

        queryset = queryset[:limit]

        suggestions = []
        for term in queryset:
            suggestions.append(
                {
                    "identifier": term.identifier,
                    "name": term.name,
                    "definition": term.definition or "",
                    "synonyms": [s.strip() for s in term.synonyms.split(";") if s.strip()] if term.synonyms else [],
                    "source": "ms",
                    "match_type": "fuzzy",
                    "category": term.category,
                }
            )

        return Response({"results": suggestions})

    @action(detail=False, methods=["get"])
    def instruments(self, request):
        """Get instrument-specific terms."""
        query = request.GET.get("q", "").strip()
        limit = int(request.GET.get("limit", 20))

        queryset = self.get_queryset().filter(category="instrument")
        if query:
            queryset = queryset.filter(models.Q(name__icontains=query) | models.Q(synonyms__icontains=query))

        return Response(
            {
                "results": [
                    {
                        "identifier": term.identifier,
                        "name": term.name,
                        "definition": term.definition or "",
                        "source": "ms",
                        "match_type": "exact",
                    }
                    for term in queryset[:limit]
                ]
            }
        )


# ===================================================================
# UNIFIED ONTOLOGY SEARCH ENDPOINT FOR SDRF VALIDATION
# ===================================================================


class CellOntologyViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for Cell Ontology (cell types and cell lines)."""

    queryset = CellOntology.objects.filter(obsolete=False)
    serializer_class = CellOntologySerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["name", "identifier", "cell_line", "organism", "source"]
    search_fields = ["name", "synonyms", "identifier", "organism", "tissue_origin"]
    ordering_fields = ["name", "identifier", "organism"]
    ordering = ["name"]

    @action(detail=False, methods=["get"])
    def search_suggestions(self, request):
        """Get cell type/cell line suggestions for SDRF validation."""
        query = request.GET.get("q", "").strip()
        cell_type = request.GET.get("cell_type", "").strip()  # 'cell_line', 'primary', or 'all'
        organism = request.GET.get("organism", "").strip()
        limit = int(request.GET.get("limit", 20))

        if not query or len(query) < 2:
            return Response({"results": []})

        # Search by name and synonyms
        queryset = self.get_queryset().filter(models.Q(name__icontains=query) | models.Q(synonyms__icontains=query))

        # Filter by cell type if provided
        if cell_type == "cell_line":
            queryset = queryset.filter(cell_line=True)
        elif cell_type == "primary":
            queryset = queryset.filter(cell_line=False)

        # Filter by organism if provided
        if organism:
            queryset = queryset.filter(organism__icontains=organism)

        queryset = queryset[:limit]

        suggestions = []
        for cell in queryset:
            suggestions.append(
                {
                    "identifier": cell.identifier,
                    "name": cell.name,
                    "definition": cell.definition or "",
                    "synonyms": [s.strip() for s in cell.synonyms.split(";") if s.strip()] if cell.synonyms else [],
                    "source": cell.source,
                    "match_type": "fuzzy",
                    "cell_type": "Cell Line" if cell.cell_line else "Cell Type",
                    "organism": cell.organism or "",
                    "tissue_origin": cell.tissue_origin or "",
                }
            )

        return Response({"results": suggestions})

    @action(detail=False, methods=["get"])
    def cell_lines(self, request):
        """Get cell lines for proteomics experiments."""
        queryset = self.get_queryset().filter(cell_line=True)
        organism = request.GET.get("organism", "").strip()

        if organism:
            queryset = queryset.filter(organism__icontains=organism)

        try:
            limit = int(request.query_params.get("limit", 10))
            limit = max(0, limit)  # Ensure non-negative
        except (ValueError, TypeError):
            limit = 10  # Default fallback
        queryset = queryset[:limit]
        serializer = self.get_serializer(queryset, many=True)
        return Response({"results": serializer.data})


class OntologySearchViewSet(viewsets.ViewSet):
    """Unified ontology search for SDRF validation."""

    permission_classes = [IsAuthenticated]

    def _get_ontology_suggestions_unified(
        self, ontology_type: str, search_term: str, limit: int, search_type: str, custom_filters: dict = None
    ):
        """
        Get ontology suggestions using the unified approach like MetadataColumn.get_ontology_suggestions.
        """

        # Create a temporary object to use the unified search method
        class TempOntologySearcher:
            def __init__(self, ontology_type, custom_ontology_filters=None):
                self.ontology_type = ontology_type
                self.custom_ontology_filters = custom_ontology_filters or {}

            def get_ontology_model(self):
                """Get the Django model class for this ontology type."""
                ontology_mapping = {
                    "species": Species,
                    "tissue": Tissue,
                    "human_disease": HumanDisease,
                    "subcellular_location": SubcellularLocation,
                    "ms_unique_vocabularies": MSUniqueVocabularies,
                    "unimod": Unimod,
                    "chebi": ChEBICompound,
                    "ncbi_taxonomy": NCBITaxonomy,
                    "mondo": MondoDisease,
                    "uberon": UberonAnatomy,
                    "cell_ontology": CellOntology,
                    "psi_ms": PSIMSOntology,
                }
                return ontology_mapping.get(self.ontology_type)

            def get_ontology_suggestions(self, search_term: str = "", limit: int = 20, search_type: str = "icontains"):
                """Use the same logic as MetadataColumn.get_ontology_suggestions."""
                model_class = self.get_ontology_model()
                if not model_class:
                    return []
                queryset = model_class.objects.all()

                # Apply custom ontology filters first
                # Custom filters can be structured as:
                # 1. {"ontology_type": {"field": "value"}} - wrapped with ontology type (MetadataColumn format)
                # 2. {"field": "value"} - direct filters (standalone format)
                if self.custom_ontology_filters:
                    # Check if filters are wrapped with ontology type
                    actual_filters = self.custom_ontology_filters.get(self.ontology_type, self.custom_ontology_filters)

                    for field, filter_value in actual_filters.items():
                        # Skip if this is still the ontology type wrapper
                        if field == self.ontology_type:
                            continue

                        if isinstance(filter_value, dict):
                            # Handle complex filter values like {'icontains': 'value'} or {'exact': 'value'}
                            for lookup, value in filter_value.items():
                                filter_kwargs = {f"{field}__{lookup}": value}
                                queryset = queryset.filter(**filter_kwargs)
                        else:
                            # Handle simple filter values
                            queryset = queryset.filter(**{field: filter_value})

                # Apply search filtering based on search_type and model type
                if search_term:
                    search_queries = []

                    # Build search queries based on ontology type and search type
                    if self.ontology_type == "species":
                        search_fields = ["official_name", "common_name", "code"]
                        for field in search_fields:
                            search_queries.append(models.Q(**{f"{field}__{search_type}": search_term}))

                    elif self.ontology_type == "tissue":
                        search_fields = ["identifier", "accession", "synonyms"]
                        for field in search_fields:
                            search_queries.append(models.Q(**{f"{field}__{search_type}": search_term}))

                    elif self.ontology_type == "human_disease":
                        search_fields = ["identifier", "acronym", "accession", "definition", "synonyms", "keywords"]
                        for field in search_fields:
                            search_queries.append(models.Q(**{f"{field}__{search_type}": search_term}))

                    elif self.ontology_type == "subcellular_location":
                        search_fields = ["accession", "location_identifier", "definition", "synonyms", "content"]
                        for field in search_fields:
                            search_queries.append(models.Q(**{f"{field}__{search_type}": search_term}))

                    elif self.ontology_type == "ms_unique_vocabularies":
                        search_fields = ["accession", "name", "definition"]
                        for field in search_fields:
                            search_queries.append(models.Q(**{f"{field}__{search_type}": search_term}))

                    elif self.ontology_type == "unimod":
                        search_fields = ["accession", "name", "definition"]
                        for field in search_fields:
                            search_queries.append(models.Q(**{f"{field}__{search_type}": search_term}))

                    elif self.ontology_type == "chebi":
                        search_fields = ["identifier", "name", "definition", "synonyms", "formula"]
                        for field in search_fields:
                            search_queries.append(models.Q(**{f"{field}__{search_type}": search_term}))

                    elif self.ontology_type == "ncbi_taxonomy":
                        search_fields = ["scientific_name", "common_name", "synonyms"]
                        for field in search_fields:
                            search_queries.append(models.Q(**{f"{field}__{search_type}": search_term}))

                    elif self.ontology_type == "mondo":
                        search_fields = ["identifier", "name", "definition", "synonyms"]
                        # Add obsolete filter for mondo
                        queryset = queryset.filter(obsolete=False)
                        for field in search_fields:
                            search_queries.append(models.Q(**{f"{field}__{search_type}": search_term}))

                    elif self.ontology_type == "uberon":
                        search_fields = ["identifier", "name", "definition", "synonyms"]
                        for field in search_fields:
                            search_queries.append(models.Q(**{f"{field}__{search_type}": search_term}))

                    elif self.ontology_type == "cell_ontology":
                        search_fields = ["identifier", "name", "definition", "synonyms"]
                        for field in search_fields:
                            search_queries.append(models.Q(**{f"{field}__{search_type}": search_term}))

                    elif self.ontology_type == "psi_ms":
                        search_fields = ["identifier", "name", "definition"]
                        for field in search_fields:
                            search_queries.append(models.Q(**{f"{field}__{search_type}": search_term}))

                    # Apply search queries
                    if search_queries:
                        combined_query = search_queries[0]
                        for query in search_queries[1:]:
                            combined_query |= query
                        queryset = queryset.filter(combined_query)

                # Apply limit and return results
                return queryset[:limit]

        # Use the temporary searcher
        searcher = TempOntologySearcher(ontology_type, custom_filters)
        return searcher.get_ontology_suggestions(search_term, limit, search_type)

    def _format_ontology_suggestion(self, result, ontology_type: str, match_type: str):
        """Format a single ontology result into the expected suggestion format."""
        try:
            if ontology_type == "ncbi_taxonomy":
                return {
                    "identifier": str(result.tax_id),
                    "name": result.scientific_name,
                    "definition": f"Rank: {result.rank}"
                    + (f", Common: {result.common_name}" if result.common_name else ""),
                    "source": "ncbi_taxonomy",
                    "match_type": match_type,
                }
            elif ontology_type == "species":
                return {
                    "identifier": result.code,
                    "name": result.official_name,
                    "definition": f"Common name: {result.common_name or 'N/A'}",
                    "source": "species",
                    "match_type": match_type,
                }
            elif ontology_type == "mondo":
                return {
                    "identifier": result.identifier,
                    "name": result.name,
                    "definition": result.definition or "",
                    "source": "mondo",
                    "match_type": match_type,
                }
            elif ontology_type == "human_disease":
                return {
                    "identifier": result.accession,
                    "name": result.identifier,
                    "definition": result.definition or "",
                    "source": "human_disease",
                    "match_type": match_type,
                }
            elif ontology_type == "tissue":
                return {
                    "identifier": result.accession,
                    "name": result.identifier,
                    "definition": result.synonyms or "",
                    "source": "tissue",
                    "match_type": match_type,
                }
            elif ontology_type == "subcellular_location":
                return {
                    "identifier": result.accession,
                    "name": result.location_identifier,
                    "definition": result.definition or "",
                    "source": "subcellular_location",
                    "match_type": match_type,
                }
            elif ontology_type == "ms_unique_vocabularies":
                return {
                    "identifier": result.accession,
                    "name": result.name,
                    "definition": result.definition or "",
                    "source": "ms_unique_vocabularies",
                    "match_type": match_type,
                }
            elif ontology_type == "unimod":
                return {
                    "identifier": result.accession,
                    "name": result.name,
                    "definition": result.definition or "",
                    "source": "unimod",
                    "match_type": match_type,
                }
            elif ontology_type == "chebi":
                return {
                    "identifier": result.identifier,
                    "name": result.name,
                    "definition": result.definition or result.formula or "",
                    "source": "chebi",
                    "match_type": match_type,
                }
            elif ontology_type == "uberon":
                return {
                    "identifier": result.identifier,
                    "name": result.name,
                    "definition": result.definition or "",
                    "source": "uberon",
                    "match_type": match_type,
                }
            elif ontology_type == "cell_ontology":
                return {
                    "identifier": result.identifier,
                    "name": result.name,
                    "definition": result.definition or "",
                    "source": "cell_ontology",
                    "match_type": match_type,
                }
            elif ontology_type == "psi_ms":
                return {
                    "identifier": result.identifier,
                    "name": result.name,
                    "definition": result.definition or "",
                    "source": "psi_ms",
                    "match_type": match_type,
                }
            else:
                # Generic fallback
                return {
                    "identifier": getattr(result, "identifier", "")
                    or getattr(result, "accession", "")
                    or getattr(result, "code", ""),
                    "name": getattr(result, "name", "")
                    or getattr(result, "official_name", "")
                    or getattr(result, "scientific_name", ""),
                    "definition": getattr(result, "definition", "") or "",
                    "source": ontology_type,
                    "match_type": match_type,
                }
        except Exception:
            # Log error and return None to skip this result
            return None

    @action(detail=False, methods=["get"])
    def suggest(self, request):
        """
        Get ontology suggestions across all sources for SDRF validation.

        Query parameters:
        - q: Search query (required, min 2 characters)
        - type: Ontology type to search (optional)
        - match: Match type - 'contains' or 'startswith' (default: 'contains')
        - limit: Maximum number of results (default: 50)
        - custom_filters: JSON string of custom filters to apply (optional)
          Example: {"organism": "Homo sapiens"} or {"organism": {"icontains": "human"}}
        """
        query = request.GET.get("q", "").strip()
        ontology_type = request.GET.get("type", "").strip()
        match_type = request.GET.get("match", "contains").strip().lower()  # 'contains' or 'startswith'
        limit = int(request.GET.get("limit", 50))
        custom_filters_param = request.GET.get("custom_filters", "").strip()

        if not query or len(query) < 2:
            return Response({"results": []})

        # Validate match_type parameter
        if match_type not in ["contains", "startswith"]:
            return Response(
                {"error": "Invalid match type. Use 'contains' or 'startswith'", "provided": match_type}, status=400
            )

        # Parse custom filters if provided
        custom_filters = {}
        if custom_filters_param:
            try:
                import json

                custom_filters = json.loads(custom_filters_param)
                if not isinstance(custom_filters, dict):
                    return Response({"error": "custom_filters must be a JSON object"}, status=400)
            except json.JSONDecodeError as e:
                return Response({"error": f"Invalid JSON in custom_filters: {str(e)}"}, status=400)

        # Convert match_type to search_type format used by unified method
        search_type = "istartswith" if match_type == "startswith" else "icontains"

        suggestions_with_types = []

        # If specific ontology type requested, use unified search
        if ontology_type:
            # Use the unified search method
            results = self._get_ontology_suggestions_unified(ontology_type, query, limit, search_type, custom_filters)

            # Keep model instances with their ontology type
            for result in results:
                suggestions_with_types.append((result, ontology_type))
        else:
            # If no specific ontology type, search across common ones
            common_ontologies = ["species", "ncbi_taxonomy", "mondo", "human_disease", "tissue", "uberon"]

            for ont_type in common_ontologies:
                # Limit per ontology type to avoid overwhelming results
                per_type_limit = max(1, limit // len(common_ontologies))
                results = self._get_ontology_suggestions_unified(
                    ont_type, query, per_type_limit, search_type, custom_filters
                )

                for result in results:
                    suggestions_with_types.append((result, ont_type))

                # Stop if we have enough suggestions
                if len(suggestions_with_types) >= limit:
                    break

        # Limit final results
        suggestions_with_types = suggestions_with_types[:limit]

        # Use OntologySuggestionSerializer to format results consistently with other routes
        formatted_results = []
        for model_instance, ont_type in suggestions_with_types:
            serializer = OntologySuggestionSerializer(model_instance, context={"ontology_type": ont_type})
            formatted_results.append(serializer.data)

        return Response(
            {
                "ontology_type": ontology_type or "all",
                "suggestions": formatted_results,
                "search_term": query,
                "search_type": match_type,
                "limit": limit,
                "count": len(formatted_results),
                "custom_filters": custom_filters if custom_filters else None,
                "has_more": len(formatted_results) >= limit,
            }
        )


class SchemaViewSet(FilterMixin, viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for Schema model - provides read-only access to schemas.
    Supports listing and retrieving schemas available to the user.
    """

    serializer_class = SchemaSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ["name", "display_name", "description"]
    filterset_fields = ["is_builtin", "is_active", "is_public"]

    def get_queryset(self):
        """Return schemas available to the user."""
        return Schema.get_available_schemas(user=self.request.user)

    @action(detail=False, methods=["get"])
    def available(self, request):
        """Get list of available schemas for creating templates."""
        schemas = self.get_queryset()
        serializer = self.get_serializer(schemas, many=True, context={"request": request})
        return Response(serializer.data)

    @action(detail=False, methods=["post"], permission_classes=[IsAdminUser])
    def sync_builtin(self, request):
        """Sync builtin schemas from sdrf-pipelines package (admin only)."""
        try:
            result = Schema.sync_builtin_schemas()

            if "error" in result:
                return Response(
                    {"error": f'Error syncing schemas: {result["error"]}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

            return Response(
                {
                    "message": "Schema sync completed successfully",
                    "created": result.get("created", 0),
                    "updated": result.get("updated", 0),
                }
            )

        except Exception as e:
            return Response(
                {"error": f"Error during schema sync: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SDRFDefaultsViewSet(viewsets.ViewSet):
    """
    ViewSet for SDRF default values - provides structured default values for SDRF columns.
    """

    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=["get"])
    def columns(self, request):
        """Get all available SDRF columns with their default values."""
        from .sdrf_defaults import get_all_column_defaults

        defaults = get_all_column_defaults()
        return Response({"columns": list(defaults.keys()), "total": len(defaults)})

    @action(detail=False, methods=["get"])
    def column_values(self, request):
        """Get default values for a specific column."""
        from .sdrf_defaults import get_column_defaults

        column_name = request.query_params.get("column")
        if not column_name:
            return Response({"error": "column parameter is required"}, status=status.HTTP_400_BAD_REQUEST)

        defaults = get_column_defaults(column_name)
        if not defaults:
            return Response({"error": f"No defaults found for column: {column_name}"}, status=status.HTTP_404_NOT_FOUND)

        return Response({"column": column_name, "config": defaults})

    @action(detail=False, methods=["get"])
    def search(self, request):
        """Search for columns containing the query string."""
        from .sdrf_defaults import search_columns

        query = request.query_params.get("q", "")
        if not query:
            return Response({"error": "q parameter is required"}, status=status.HTTP_400_BAD_REQUEST)

        results = search_columns(query)
        return Response({"query": query, "matches": results, "total": len(results)})

    @action(detail=False, methods=["get"])
    def structured_options(self, request):
        """Get options for structured fields (key-value pairs)."""
        from .sdrf_defaults import get_structured_field_options

        column_name = request.query_params.get("column")
        field_type = request.query_params.get("type")

        if not column_name:
            return Response({"error": "column parameter is required"}, status=status.HTTP_400_BAD_REQUEST)

        options = get_structured_field_options(column_name, field_type)
        if not options:
            return Response(
                {"error": f"No structured options found for column: {column_name}"}, status=status.HTTP_404_NOT_FOUND
            )

        return Response({"column": column_name, "field_type": field_type, "options": options})

    @action(detail=False, methods=["get"])
    def labels(self, request):
        """Get all label options (TMT, SILAC, label-free)."""
        from .sdrf_defaults import LABEL_VALUES

        return Response(
            {
                "label_free": LABEL_VALUES["label free sample"],
                "tmt_labels": LABEL_VALUES["tmt_labels"],
                "silac_labels": LABEL_VALUES["silac_labels"],
            }
        )

    @action(detail=False, methods=["get"])
    def modifications(self, request):
        """Get protein modification options (fixed and variable)."""
        from .sdrf_defaults import PROTEIN_MODIFICATIONS

        mod_type = request.query_params.get("type")
        if mod_type and mod_type in PROTEIN_MODIFICATIONS:
            return Response({"type": mod_type, "modifications": PROTEIN_MODIFICATIONS[mod_type]})

        return Response({"fixed": PROTEIN_MODIFICATIONS["fixed"], "variable": PROTEIN_MODIFICATIONS["variable"]})

    @action(detail=False, methods=["get"])
    def instruments(self, request):
        """Get list of common instrument models."""
        from .sdrf_defaults import INSTRUMENT_MODELS

        return Response({"instruments": INSTRUMENT_MODELS})

    @action(detail=False, methods=["get"])
    def cleavage_agents(self, request):
        """Get cleavage agent options."""
        from .sdrf_defaults import CLEAVAGE_AGENTS

        agent_type = request.query_params.get("type")
        if agent_type and agent_type in CLEAVAGE_AGENTS:
            return Response({"type": agent_type, "agents": CLEAVAGE_AGENTS[agent_type]})

        return Response({"enzymatic": CLEAVAGE_AGENTS["enzymatic"], "non_enzymatic": CLEAVAGE_AGENTS["non_enzymatic"]})

    @action(detail=False, methods=["get"])
    def compound_fields(self, request):
        """Get all compound fields that require special handling."""
        from .sdrf_defaults import get_all_compound_fields

        fields = get_all_compound_fields()
        return Response({"compound_fields": list(fields.keys()), "schemas": fields})

    @action(detail=False, methods=["get"])
    def compound_schema(self, request):
        """Get schema for a specific compound field."""
        from .sdrf_defaults import get_compound_field_schema

        column_name = request.query_params.get("column")
        if not column_name:
            return Response({"error": "column parameter is required"}, status=status.HTTP_400_BAD_REQUEST)

        schema = get_compound_field_schema(column_name)
        if not schema:
            return Response(
                {"error": f"No compound schema found for column: {column_name}"}, status=status.HTTP_404_NOT_FOUND
            )

        return Response({"column": column_name, "schema": schema})

    @action(detail=False, methods=["post"])
    def validate_compound(self, request):
        """Validate a compound field value against its schema."""
        from .sdrf_defaults import validate_compound_field_value

        column_name = request.data.get("column")
        value = request.data.get("value")

        if not column_name or value is None:
            return Response({"error": "column and value parameters are required"}, status=status.HTTP_400_BAD_REQUEST)

        validation_result = validate_compound_field_value(column_name, value)

        return Response({"column": column_name, "value": value, "validation": validation_result})

    @action(detail=False, methods=["get"])
    def suggestions(self, request):
        """Get column name suggestions based on partial input."""
        from .sdrf_defaults import get_column_suggestions

        partial = request.query_params.get("partial", "")
        if not partial:
            return Response({"error": "partial parameter is required"}, status=status.HTTP_400_BAD_REQUEST)

        suggestions = get_column_suggestions(partial)
        return Response({"partial": partial, "suggestions": suggestions, "total": len(suggestions)})

    @action(detail=False, methods=["get"])
    def quick_values(self, request):
        """Get quick access to commonly used value categories."""
        from .sdrf_defaults import get_quick_values

        category = request.query_params.get("category")
        if not category:
            return Response(
                {
                    "available_categories": [
                        "labels",
                        "instruments",
                        "organisms",
                        "diseases",
                        "cell_types",
                        "data_acquisition",
                        "fragmentation",
                        "replicates",
                    ]
                }
            )

        values = get_quick_values(category)
        if values is None:
            return Response({"error": f"Unknown category: {category}"}, status=status.HTTP_404_NOT_FOUND)

        return Response({"category": category, "values": values})
