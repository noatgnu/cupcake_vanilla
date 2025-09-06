"""
Import utility functions for CUPCAKE Vanilla.

This module contains the core import logic that can be used by both
sync views and async RQ tasks.
"""

import io
import json
from collections import Counter
from typing import Any, Dict

from django.db import transaction

from openpyxl import load_workbook

from ccv.models import MetadataColumn, MetadataTableTemplate, SamplePool


def _create_modifiers(metadata_value_map: dict) -> list:
    """
    Create efficient modifiers by combining samples with the same values into ranges.

    Example:
        >>> metadata_map = {1: 'value_a', 2: 'value_a', 4: 'value_b'}
        >>> modifiers = _create_modifiers(metadata_map)
        >>> print(modifiers[0]['samples'])
        '1-2'
        >>> print(modifiers[1]['samples'])
        '4'

    Args:
        metadata_value_map: Dict mapping sample numbers to values

    Returns:
        List of modifiers with combined sample ranges for efficient storage
    """
    if not metadata_value_map:
        return []

    value_to_samples = {}
    for sample_num, value in metadata_value_map.items():
        if value not in value_to_samples:
            value_to_samples[value] = []
        value_to_samples[value].append(int(sample_num))

    modifiers = []
    for value, sample_list in value_to_samples.items():
        sample_list.sort()

        ranges = []
        i = 0

        while i < len(sample_list):
            start = sample_list[i]
            end = sample_list[i]

            while i + 1 < len(sample_list) and sample_list[i + 1] == end + 1:
                i += 1
                end = sample_list[i]

            if start == end:
                ranges.append(str(start))
            elif end - start == 1:
                ranges.append(str(start))
                ranges.append(str(end))
            else:
                ranges.append(f"{start}-{end}")

            i += 1

        modifiers.append({"samples": ",".join(ranges), "value": value})

    return modifiers


def _truncate_pool_name(pool_name: str, max_length: int = 250) -> str:
    """
    Truncate pool name to fit database constraints while preserving readability.

    Example:
        >>> long_name = "Very Long Pool Name That Exceeds Database Limits" * 10
        >>> short_name = _truncate_pool_name(long_name, 50)
        >>> len(short_name) <= 50
        True

    Args:
        pool_name: The original pool name to truncate
        max_length: Maximum allowed length (default 250 for database constraints)

    Returns:
        Truncated pool name that fits within database constraints
    """
    if not pool_name:
        return "Unknown Pool"

    pool_name = pool_name.strip()

    if len(pool_name) <= max_length:
        return pool_name

    truncated = pool_name[:max_length]

    if max_length > 20:
        last_space = truncated.rfind(" ")
        last_underscore = truncated.rfind("_")
        last_dash = truncated.rfind("-")

        boundary = max(last_space, last_underscore, last_dash)
        if boundary > max_length * 0.8:
            truncated = truncated[:boundary]

    return truncated


def _get_table_template(metadata_table):
    """
    Get the table template that existing columns originate from.

    Example:
        >>> table = MetadataTable.objects.get(id=1)
        >>> template = _get_table_template(table)
        >>> template.name if template else None
        'Proteomics Template'

    Args:
        metadata_table: MetadataTable instance to analyze

    Returns:
        Best matching MetadataTableTemplate or None if no match found
    """
    existing_columns = list(metadata_table.columns.all())
    if existing_columns:
        template_column_templates = set()
        for col in existing_columns:
            if col.template:
                template_column_templates.add(col.template)

        if template_column_templates:
            best_template = None
            best_match_count = 0

            for table_template in MetadataTableTemplate.objects.all():
                table_template_columns = set()
                for template_col in table_template.user_columns.all():
                    if hasattr(template_col, "template") and template_col.template:
                        table_template_columns.add(template_col.template)

                match_count = len(template_column_templates.intersection(table_template_columns))
                if match_count > best_match_count:
                    best_match_count = match_count
                    best_template = table_template

            if best_template and best_match_count > 0:
                return best_template

    return None


def _find_or_create_matching_column(
    clean_name, metadata_type, metadata_table, table_template, column_position, occurrence_number
):
    """
    Find existing column or create new one using template properties for consistent metadata.

    Example:
        >>> # Creates column with ontology properties from template
        >>> column = _find_or_create_matching_column(
        ...     "characteristics[organism]", "characteristic", table, template, 2, 1
        ... )
        >>> column.ontology_type
        'NCBITAXON'
        >>> column.name
        'characteristics[organism]'

    Args:
        clean_name: Column name to find or create
        metadata_type: Type of metadata column
        metadata_table: Parent metadata table
        table_template: Template to use for column properties
        column_position: Position in the table
        occurrence_number: Which occurrence of this column name

    Returns:
        MetadataColumn instance (existing or newly created)
    """
    existing_columns = list(metadata_table.columns.filter(name__iexact=clean_name).order_by("column_position"))

    if occurrence_number <= len(existing_columns):
        existing_column = existing_columns[occurrence_number - 1]
        existing_column.column_position = column_position
        existing_column.save(update_fields=["column_position"])
        return existing_column
    else:
        template_column = None
        if table_template:
            template_column = table_template.user_columns.filter(name__iexact=clean_name).first()

        if template_column:
            metadata_column = MetadataColumn.objects.create(
                name=clean_name,
                type=metadata_type or template_column.type,
                column_position=column_position,
                metadata_table=metadata_table,
                template=template_column.template,
                ontology_type=template_column.ontology_type,
                mandatory=template_column.mandatory,
                hidden=template_column.hidden,
                readonly=template_column.readonly,
                auto_generated=template_column.auto_generated,
            )
        else:
            metadata_column = MetadataColumn.objects.create(
                name=clean_name,
                type=metadata_type,
                column_position=column_position,
                metadata_table=metadata_table,
            )

        return metadata_column


def synchronize_pools_with_import_data(metadata_table, import_pools_data, metadata_columns, user):
    """
    Synchronize pools with import data for unified metadata system with intelligent merging.

    Example:
        >>> import_data = [{
        ...     'pool_name': 'Test Pool',
        ...     'pooled_only_samples': [1, 2, 3],
        ...     'is_reference': True
        ... }]
        >>> synchronize_pools_with_import_data(table, import_data, columns, user)
        >>> table.sample_pools.count()
        1

    Args:
        metadata_table: MetadataTable to synchronize pools for
        import_pools_data: List of pool data dictionaries from import
        metadata_columns: List of metadata columns for pool creation
        user: User performing the synchronization
    """
    existing_pools = SamplePool.objects.filter(metadata_table=metadata_table)
    existing_pool_names = {pool.pool_name: pool for pool in existing_pools}
    existing_pool_ids = {pool.id: pool for pool in existing_pools}

    import_pool_names = {pool_data["pool_name"] for pool_data in import_pools_data}

    for pool_data in import_pools_data:
        pool_name = pool_data["pool_name"]
        pool_id = pool_data.get("pool_id")
        existing_pool = None

        if pool_id and pool_id in existing_pool_ids:
            existing_pool = existing_pool_ids[pool_id]
        elif pool_name in existing_pool_names:
            existing_pool = existing_pool_names[pool_name]

        if existing_pool:
            existing_pool.pooled_only_samples = pool_data["pooled_only_samples"]
            existing_pool.pooled_and_independent_samples = pool_data["pooled_and_independent_samples"]
            existing_pool.is_reference = pool_data["is_reference"]
            existing_pool.save()

            existing_pool.metadata_columns.clear()
            _create_pool_metadata_from_import(existing_pool, pool_data, metadata_columns)
        else:
            new_pool = SamplePool.objects.create(
                pool_name=pool_name,
                pooled_only_samples=pool_data["pooled_only_samples"],
                pooled_and_independent_samples=pool_data["pooled_and_independent_samples"],
                is_reference=pool_data["is_reference"],
                metadata_table=metadata_table,
                created_by=user,
            )

            _create_pool_metadata_from_import(new_pool, pool_data, metadata_columns)

    pools_to_delete = [pool for pool_name, pool in existing_pool_names.items() if pool_name not in import_pool_names]

    for pool in pools_to_delete:
        pool.delete()


def _calculate_most_common_value_for_column(data_rows, pooled_sample_indices, col_index, metadata_column):
    """
    Calculate the most common value for a column among pooled samples using statistical analysis.

    Example:
        >>> # Real proteomics data with ontology terms
        >>> data = [
        ...     ['D-HEp3 #1', 'homo sapiens', 'NT=head and neck;AC=MA:0000006'],
        ...     ['D-HEp3 #2', 'homo sapiens', 'NT=head and neck;AC=MA:0000006'],
        ...     ['T-HEp3 #1', 'homo sapiens', 'NT=head and neck;AC=MA:0000006']
        ... ]
        >>> indices = [1, 2, 3]
        >>> column = MetadataColumn(not_applicable=True)
        >>> result = _calculate_most_common_value_for_column(data, indices, 1, column)
        >>> result
        'homo sapiens'

    Args:
        data_rows: All data rows from the import file
        pooled_sample_indices: List of 1-based sample indices in this pool
        col_index: Column index to analyze for common values
        metadata_column: Metadata column for fallback behavior

    Returns:
        Most frequently occurring value among pooled samples for this column
    """
    values = []

    for sample_index in pooled_sample_indices:
        row_index = sample_index - 1
        if row_index < len(data_rows):
            row = data_rows[row_index]
            if col_index < len(row):
                cell_value = row[col_index].strip()
                if cell_value and cell_value != "":
                    values.append(cell_value)

    if not values:
        if metadata_column.not_applicable:
            return "not applicable"
        else:
            return "not available"

    value_counts = Counter(values)
    most_common_value = value_counts.most_common(1)[0][0]

    return most_common_value


def create_pool_metadata_from_table_columns(pool):
    """
    Create metadata columns for a pool based on parent table columns with intelligent value derivation.

    Example:
        >>> # Creates pool with HEp-3 cell line data from fixtures
        >>> pool = SamplePool.objects.create(
        ...     pool_name='D-HEp3 Pool',
        ...     pooled_only_samples=[1, 2],
        ...     metadata_table=table,
        ...     sdrf_value='SN=D-HEp3 #1,D-HEp3 #2'
        ... )
        >>> create_pool_metadata_from_table_columns(pool)
        >>> pool.metadata_columns.filter(name__icontains='organism').first().value
        'homo sapiens'

    Args:
        pool: SamplePool instance to create metadata columns for
    """
    table_columns = list(pool.metadata_table.columns.all())
    pool.metadata_columns.clear()

    for table_column in table_columns:
        if "pooled sample" in table_column.name.lower():
            pool_metadata_column = MetadataColumn.objects.create(
                name=table_column.name,
                type=table_column.type,
                value=pool.sdrf_value,
                mandatory=table_column.mandatory,
                hidden=table_column.hidden,
            )
        elif "source name" in table_column.name.lower():
            pool_metadata_column = MetadataColumn.objects.create(
                name=table_column.name,
                type=table_column.type,
                value=pool.pool_name,
                mandatory=table_column.mandatory,
                hidden=table_column.hidden,
            )
        else:
            pool_value = _calculate_most_common_value_for_pool_column(pool, table_column)
            pool_metadata_column = MetadataColumn.objects.create(
                name=table_column.name,
                type=table_column.type,
                value=pool_value,
                mandatory=table_column.mandatory,
                hidden=table_column.hidden,
            )

        pool.metadata_columns.add(pool_metadata_column)


def _calculate_most_common_value_for_pool_column(pool, table_column):
    """
    Calculate the most common value for a column among pool samples with modifier support.

    Example:
        >>> # Pool with HEp-3 samples having organism-specific modifiers
        >>> pool = SamplePool(pooled_only_samples=[1, 2, 3])
        >>> column = MetadataColumn(
        ...     value='homo sapiens',
        ...     modifiers=[
        ...         {'samples': '1-2', 'value': 'NT=HEp-3 cell;AC=BTO:0005139'}
        ...     ]
        ... )
        >>> result = _calculate_most_common_value_for_pool_column(pool, column)
        >>> result
        'NT=HEp-3 cell;AC=BTO:0005139'

    Args:
        pool: SamplePool instance with sample indices
        table_column: MetadataColumn to analyze for values and modifiers

    Returns:
        Most common value among pool samples considering modifiers
    """
    pooled_sample_indices = pool.pooled_only_samples + pool.pooled_and_independent_samples
    values = []

    if table_column.modifiers and isinstance(table_column.modifiers, list):
        for sample_idx in pooled_sample_indices:
            sample_value = None

            for modifier in table_column.modifiers:
                if isinstance(modifier, dict):
                    sample_range = modifier.get("samples", "")
                    if sample_range:
                        sample_indices = []
                        for part in sample_range.split(","):
                            part = part.strip()
                            if "-" in part:
                                start, end = part.split("-")
                                sample_indices.extend(range(int(start), int(end) + 1))
                            else:
                                sample_indices.append(int(part))

                        if sample_idx in sample_indices:
                            sample_value = modifier.get("value", "")
                            break

            if sample_value is None:
                sample_value = table_column.value or ""

            if sample_value:
                values.append(sample_value)
    else:
        if table_column.value:
            return table_column.value

    if not values:
        if table_column.not_applicable:
            return "not applicable"
        else:
            return "not available"

    value_counts = Counter(values)
    most_common_value = value_counts.most_common(1)[0][0]

    return most_common_value


def _create_pool_metadata_from_import(pool, pool_data, metadata_columns):
    """
    Create metadata for a pool from import data with intelligent value derivation.

    Example:
        >>> # Real proteomics pool data from SDRF fixtures
        >>> pool_data = {
        ...     'pool_name': 'D-HEp3 Pool',
        ...     'sdrf_value': 'SN=D-HEp3 #1,D-HEp3 #2',
        ...     'metadata_row': [
        ...         'D-HEp3 Pool', 'homo sapiens',
        ...         'NT=head and neck;AC=MA:0000006',
        ...         'NT=HEp-3 cell;AC=BTO:0005139'
        ...     ],
        ...     'pooled_only_samples': [1, 2]
        ... }
        >>> _create_pool_metadata_from_import(pool, pool_data, columns)
        >>> pool.metadata_columns.filter(name__icontains='cell line').first().value
        'NT=HEp-3 cell;AC=BTO:0005139'

    Args:
        pool: SamplePool instance to create metadata for
        pool_data: Dictionary containing pool import data
        metadata_columns: List of metadata columns to create
    """
    row = pool_data.get("metadata_row")
    pool_name = pool_data["pool_name"]
    sdrf_value = pool_data["sdrf_value"]
    has_sn_pattern = sdrf_value.startswith("SN=")

    all_data_rows = pool_data.get("all_data_rows", [])
    pooled_sample_indices = pool_data["pooled_only_samples"] + pool_data["pooled_and_independent_samples"]

    for col_index, metadata_column in enumerate(metadata_columns):
        if "pooled sample" in metadata_column.name.lower():
            pool_metadata_column = MetadataColumn.objects.create(
                name=metadata_column.name,
                type=metadata_column.type,
                value=sdrf_value,
                mandatory=metadata_column.mandatory,
                hidden=metadata_column.hidden,
            )
        elif "source name" in metadata_column.name.lower():
            pool_metadata_column = MetadataColumn.objects.create(
                name=metadata_column.name,
                type=metadata_column.type,
                value=pool_name,
                mandatory=metadata_column.mandatory,
                hidden=metadata_column.hidden,
            )
        else:
            if has_sn_pattern and row:
                pool_value = row[col_index] if col_index < len(row) else ""
            else:
                pool_value = _calculate_most_common_value_for_column(
                    all_data_rows, pooled_sample_indices, col_index, metadata_column
                )

            pool_metadata_column = MetadataColumn.objects.create(
                name=metadata_column.name,
                type=metadata_column.type,
                value=pool_value or "not available",
                mandatory=metadata_column.mandatory,
                hidden=metadata_column.hidden,
            )

        pool.metadata_columns.add(pool_metadata_column)


def import_sdrf_data(
    file_content: str,
    metadata_table,
    user,
    replace_existing: bool = False,
    validate_ontologies: bool = True,
    create_pools: bool = True,
) -> Dict[str, Any]:
    r"""
    Import SDRF (Sample and Data Relationship Format) data with intelligent parsing and pool creation.

    Example:
        >>> # Real proteomics SDRF with ontology terms from test fixtures
        >>> content = "source name\tcharacteristics[organism]\tcharacteristics[cell line]\tcharacteristics[pooled sample]\n" + \
        ...           "D-HEp3 #1\thomo sapiens\tNT=HEp-3 cell;AC=BTO:0005139\tpooled\n" + \
        ...           "D-HEp3 #2\thomo sapiens\tNT=HEp-3 cell;AC=BTO:0005139\tpooled\n" + \
        ...           "T-HEp3 #1\thomo sapiens\tNT=HEp-3 cell;AC=BTO:0005139\tnot pooled"
        >>> result = import_sdrf_data(
        ...     file_content=content,
        ...     metadata_table=table,
        ...     user=user,
        ...     create_pools=True,
        ...     validate_ontologies=True
        ... )
        >>> result['pools_created']
        1
        >>> result['columns_created']
        4

    Args:
        file_content: Tab-separated SDRF file content as string
        metadata_table: MetadataTable instance to import data into
        user: User performing the import operation
        replace_existing: Whether to replace existing data or merge
        validate_ontologies: Whether to validate ontology terms against vocabularies
        create_pools: Whether to create sample pools from SN= patterns

    Returns:
        Dictionary containing success status, statistics, and any warnings
    """
    with transaction.atomic():
        # Parse the file content
        lines = file_content.strip().split("\n")
        if not lines:
            raise ValueError("Empty file content")

        # Parse headers and data
        headers = lines[0].split("\t")
        data_rows = [line.split("\t") for line in lines[1:]]

        # Check for pooled sample column and identify pool data
        pooled_column_index = None
        pooled_rows = []
        sn_rows = []

        for i, header in enumerate(headers):
            header_lower = header.lower()
            if "pooled sample" in header_lower or "pooled_sample" in header_lower:
                pooled_column_index = i
                break

        # If we found a pooled sample column, process the data
        if pooled_column_index is not None:
            for row_index, row in enumerate(data_rows):
                if pooled_column_index < len(row):
                    pooled_value = row[pooled_column_index].strip()
                    if pooled_value.startswith("SN="):
                        sn_rows.append(row_index)
                    elif pooled_value.lower() == "pooled":
                        pooled_rows.append(row_index)

        # Check permissions
        if not metadata_table.can_edit(user):
            raise PermissionError("Permission denied: cannot edit this metadata table")

        # Clear existing columns if replace_existing is True
        if replace_existing:
            metadata_table.columns.all().delete()
            metadata_table.sample_pools.all().delete()

        # Remove SN= rows from data before general processing
        sn_data = []
        if pooled_column_index is not None and sn_rows:
            # Store SN= rows for later pool creation
            for row_index in sorted(sn_rows, reverse=True):  # Reverse order to maintain indices
                if row_index < len(data_rows):
                    sn_data.append(data_rows[row_index])
                    del data_rows[row_index]

        # Update sample count and extend/truncate data if needed
        expected_sample_count = metadata_table.sample_count or len(data_rows)
        if len(data_rows) != expected_sample_count:
            # Extend the number of samples to match expected count and fill with empty strings
            if len(data_rows) < expected_sample_count:
                data_rows.extend(
                    [["" for i in range(len(headers))] for j in range(expected_sample_count - len(data_rows))]
                )
            else:
                data_rows = data_rows[:expected_sample_count]

                # If we removed SN= rows and truncated data, add SN= rows back for pool creation
                if pooled_column_index is not None and sn_rows and sn_data:
                    # Add SN= rows back to the end of the truncated data
                    data_rows.extend(sn_data)
                    # Update sn_rows indices to reflect their new positions at the end
                    sn_rows = list(range(expected_sample_count, len(data_rows)))

        # Update metadata table sample count
        metadata_table.sample_count = len(data_rows)
        metadata_table.save(update_fields=["sample_count"])

        # Enhanced SDRF import with intelligent column matching and schema organization
        created_columns = []
        column_name_usage = {}  # Track how many times each name is used

        # Get the metadata table template that existing columns come from
        table_template = _get_table_template(metadata_table)

        # Process headers with intelligent template matching
        for i, header in enumerate(headers):
            # Parse header format: type[name] (original CUPCAKE format)
            header_lower = header.lower()
            if "[" in header_lower and "]" in header_lower:
                metadata_type = header_lower.split("[")[0].strip()
                name = header_lower.strip()
            else:
                metadata_type = "special"
                name = header_lower.strip()
            name = name.replace("_", " ")
            # Count usage of this column name
            if name not in column_name_usage:
                column_name_usage[name] = 0
            column_name_usage[name] += 1

            metadata_column = _find_or_create_matching_column(
                name, metadata_type, metadata_table, table_template, i, column_name_usage[name]
            )

            created_columns.append(metadata_column)

        columns_updated = 0
        for i, metadata_column in enumerate(created_columns):
            metadata_value_map = {}

            # Process each data row for this column (original CUPCAKE approach)
            for j, row in enumerate(data_rows):
                if i < len(row) and row[i]:
                    cell_value = row[i].strip()

                    if cell_value == "":
                        continue
                    if cell_value == "not applicable":
                        metadata_column.not_applicable = True
                        continue

                    # Use the MetadataColumn instance method for proper SDRF conversion with ontology lookup
                    value = metadata_column.convert_sdrf_to_metadata(cell_value)

                    if value not in metadata_value_map:
                        metadata_value_map[value] = []
                    metadata_value_map[value].append(j)

            # Set the most common value as default (original CUPCAKE logic)
            max_count = 0
            max_value = None
            for value in metadata_value_map:
                if len(metadata_value_map[value]) > max_count:
                    max_count = len(metadata_value_map[value])
                    max_value = value

            if max_value:
                metadata_column.value = max_value

            # Calculate modifiers for other values (exact original CUPCAKE logic)
            modifiers = []
            for value in metadata_value_map:
                if value != max_value:
                    modifier = {"samples": [], "value": value}
                    # Sort from lowest to highest. Add samples index. For continuous samples, add range
                    samples = metadata_value_map[value]
                    samples.sort()
                    start = samples[0]
                    end = samples[0]
                    for i2 in range(1, len(samples)):
                        if samples[i2] == end + 1:
                            end = samples[i2]
                        else:
                            if start == end:
                                modifier["samples"].append(str(start + 1))
                            else:
                                modifier["samples"].append(f"{start + 1}-{end + 1}")
                            start = samples[i2]
                            end = samples[i2]
                    if start == end:
                        modifier["samples"].append(str(start + 1))
                    else:
                        modifier["samples"].append(f"{start + 1}-{end + 1}")
                    if len(modifier["samples"]) == 1:
                        modifier["samples"] = modifier["samples"][0]
                    else:
                        modifier["samples"] = ",".join(modifier["samples"])
                    modifiers.append(modifier)

            if modifiers:
                # Store modifiers as native JSON object
                metadata_column.modifiers = modifiers

            metadata_column.save()
            columns_updated += 1

        # Reorder table columns after creation (same approach as Excel import)
        if created_columns:
            try:
                # Try schema-based reordering first if possible
                schema_ids = set()
                for column in created_columns:
                    if column.template and hasattr(column.template, "schema") and column.template.schema:
                        schema_ids.add(column.template.schema.id)

                if schema_ids:
                    metadata_table.reorder_columns_by_schema(schema_ids=list(schema_ids))
                else:
                    # Fall back to basic section-based ordering if available
                    if hasattr(metadata_table, "basic_column_reordering"):
                        metadata_table.basic_column_reordering()
            except Exception:
                # Silently continue if reordering fails
                pass

        pools_created = 0
        pools_updated = 0
        if create_pools and pooled_column_index is not None:
            # Find source name column
            source_name_column_index = None
            for idx, header in enumerate(headers):
                header_lower = header.lower()
                if "source name" in header_lower or "source_name" in header_lower:
                    source_name_column_index = idx
                    break

            import_pools_data = []

            if sn_rows and sn_data:
                # Case 1: There are rows with SN= values - create pools from them
                for pool_index, row in enumerate(sn_data):
                    if pooled_column_index < len(row):
                        sdrf_value = row[pooled_column_index].strip()

                        # Extract source names from SN= value
                        if sdrf_value.startswith("SN="):
                            source_names = sdrf_value[3:].split(",")
                            source_names = [name.strip() for name in source_names]

                            # Get pool name from source name column or use default
                            pool_name = (
                                row[source_name_column_index]
                                if source_name_column_index is not None and source_name_column_index < len(row)
                                else f"Pool {pool_index + 1}"
                            )

                            # Find sample indices that match these source names
                            pooled_only_samples = []
                            pooled_and_independent_samples = []

                            for sample_index, sample_row in enumerate(data_rows):
                                if source_name_column_index is not None and source_name_column_index < len(sample_row):
                                    sample_source_name = sample_row[source_name_column_index].strip()
                                    if sample_source_name in source_names:
                                        # Check if this sample is also marked as "not pooled" or independent
                                        sample_pooled_value = ""
                                        if pooled_column_index < len(sample_row):
                                            sample_pooled_value = sample_row[pooled_column_index].strip().lower()

                                        if (
                                            sample_pooled_value == "not pooled"
                                            or sample_pooled_value == ""
                                            or sample_pooled_value == "independent"
                                        ):
                                            # Sample exists both in pool and as independent
                                            pooled_and_independent_samples.append(sample_index + 1)
                                        else:
                                            # Sample is only in pool
                                            pooled_only_samples.append(sample_index + 1)

                            # Store pool data for synchronization
                            import_pools_data.append(
                                {
                                    "pool_name": pool_name,
                                    "pooled_only_samples": pooled_only_samples,
                                    "pooled_and_independent_samples": pooled_and_independent_samples,
                                    "is_reference": True,  # SN= pools are reference pools
                                    "metadata_row": row,
                                    "sdrf_value": sdrf_value,
                                    "all_data_rows": data_rows,
                                }
                            )

            elif pooled_rows:
                # Case 2: No SN= rows but there are "pooled" rows - create a pool from them
                # Get source names of all pooled samples
                pooled_source_names = []
                pooled_only_samples = []

                for row_index in pooled_rows:
                    if (
                        source_name_column_index is not None
                        and row_index < len(data_rows)
                        and source_name_column_index < len(data_rows[row_index])
                    ):
                        source_name = data_rows[row_index][source_name_column_index].strip()
                        if source_name:
                            pooled_source_names.append(source_name)
                            pooled_only_samples.append(row_index + 1)

                if pooled_source_names:
                    # Create SN= value from source names
                    sdrf_value = "SN=" + ",".join(pooled_source_names)
                    pool_name = "Pool 1"
                    template_row = data_rows[pooled_rows[0]]

                    # Store pool data for synchronization
                    import_pools_data.append(
                        {
                            "pool_name": pool_name,
                            "pooled_only_samples": pooled_only_samples,
                            "pooled_and_independent_samples": [],
                            "is_reference": False,  # Pooled rows are not reference pools by default
                            "metadata_row": template_row,
                            "sdrf_value": sdrf_value,
                            "all_data_rows": data_rows,
                        }
                    )

            # Synchronize pools with sophisticated logic (matching original CUPCAKE)
            if import_pools_data:
                synchronize_pools_with_import_data(metadata_table, import_pools_data, created_columns, user)
                # Get the updated pools list for response
                created_pools_list = list(metadata_table.sample_pools.all())
                pools_created = len(created_pools_list)
                pools_updated = 0  # For simplicity, consider all as created

        return {
            "success": True,
            "message": "SDRF data imported successfully",
            "columns_created": len(created_columns),
            "columns_updated": columns_updated,
            "sample_count": metadata_table.sample_count,
            "pools_created": pools_created,
            "pools_updated": pools_updated,
            "warnings": [],
        }


def import_excel_data(
    file_data: bytes,
    metadata_table,
    user,
    replace_existing: bool = False,
    validate_ontologies: bool = True,
    create_pools: bool = True,
) -> Dict[str, Any]:
    """
    Import multi-sheet Excel data with comprehensive pool processing and metadata creation.
    Automatically excludes legend/note sections that are added during export.

    Example:
        >>> # Excel with proteomics data containing realistic ontology terms:
        >>> # main sheet: organism=homo sapiens, cell_line=NT=HEp-3 cell;AC=BTO:0005139
        >>> # pool_object_map: ["D-HEp3 Pool", [1,2], [], true]
        >>> # pool_main: D-HEp3 Pool, homo sapiens, NT=head and neck;AC=MA:0000006
        >>> # Automatically filters out legend sections starting with "Note:", "[*]", etc.
        >>> with open('proteomics_metadata.xlsx', 'rb') as f:
        ...     file_data = f.read()
        >>> result = import_excel_data(
        ...     file_data=file_data,
        ...     metadata_table=table,
        ...     user=user,
        ...     create_pools=True,
        ...     validate_ontologies=True
        ... )
        >>> result['has_pool_data']
        True
        >>> result['pools_created']
        2

    Args:
        file_data: Binary Excel file data (.xlsx format with multiple sheets)
        metadata_table: MetadataTable instance to import data into
        user: User performing the import operation
        replace_existing: Whether to replace existing data or merge intelligently
        validate_ontologies: Whether to validate ontology terms against vocabularies
        create_pools: Whether to create sample pools from pool_object_map sheet

    Returns:
        Dictionary containing success status, import statistics, and pool information
    """
    with transaction.atomic():
        # Check permissions
        if not metadata_table.can_edit(user):
            raise PermissionError("Permission denied: cannot edit this metadata table")

        # Read Excel workbook
        wb = load_workbook(io.BytesIO(file_data))

        # Get main worksheet
        if "main" not in wb.sheetnames:
            raise ValueError("Excel file must contain a 'main' worksheet")

        main_ws = wb["main"]
        # Read all data first, then filter out legend/note sections
        all_main_data = [list(row) for row in main_ws.iter_rows(min_row=2, values_only=True)]

        # Filter out legend/note rows by detecting the legend text patterns
        # Export creates notes with specific text patterns starting with "Note:", "[*]", "[**]", "[***]"
        main_data = []
        legend_markers = ["Note:", "[*]", "[**]", "[***]"]

        for row in all_main_data:
            # Check if this row starts a legend section
            if row and row[0] is not None:
                row_text = str(row[0]).strip()
                if any(row_text.startswith(marker) for marker in legend_markers):
                    # Found legend section, stop processing
                    break
            main_data.append(row)

        # Get hidden worksheet if exists
        hidden_data = []
        if "hidden" in wb.sheetnames:
            hidden_ws = wb["hidden"]
            if hidden_ws.max_row > 1:
                # Apply same legend filtering to hidden sheet
                all_hidden_data = [list(row) for row in hidden_ws.iter_rows(min_row=2, values_only=True)]

                for row in all_hidden_data:
                    # Check if this row starts a legend section
                    if row and row[0] is not None:
                        row_text = str(row[0]).strip()
                        if any(row_text.startswith(marker) for marker in legend_markers):
                            # Found legend section, stop processing
                            break
                    hidden_data.append(row)

        # Get ID metadata column mapping
        if "id_metadata_column_map" not in wb.sheetnames:
            raise ValueError("Excel file must contain an 'id_metadata_column_map' worksheet")

        id_metadata_column_map_ws = wb["id_metadata_column_map"]
        id_metadata_column_map_list = [
            list(row) for row in id_metadata_column_map_ws.iter_rows(min_row=2, values_only=True)
        ]
        id_metadata_column_map = {}
        for row in id_metadata_column_map_list:
            if row[0] is not None:
                id_metadata_column_map[int(row[0])] = {
                    "column": row[1],
                    "name": row[2],
                    "type": row[3],
                    "hidden": row[4],
                }

        # Check for pool sheets
        pool_id_metadata_column_map = {}
        pool_object_map_data = []
        pool_main_data = []
        pool_hidden_data = []

        if "pool_main" in wb.sheetnames:
            pool_main_ws = wb["pool_main"]
            if pool_main_ws.max_row > 1:
                pool_main_data = [list(row) for row in pool_main_ws.iter_rows(min_row=2, values_only=True)]

        if "pool_hidden" in wb.sheetnames:
            pool_hidden_ws = wb["pool_hidden"]
            if pool_hidden_ws.max_row > 1:
                pool_hidden_data = [list(row) for row in pool_hidden_ws.iter_rows(min_row=2, values_only=True)]

        if "pool_id_metadata_column_map" in wb.sheetnames:
            pool_id_metadata_column_map_ws = wb["pool_id_metadata_column_map"]
            pool_id_metadata_column_map_list = [
                list(row) for row in pool_id_metadata_column_map_ws.iter_rows(min_row=2, values_only=True)
            ]
            for row in pool_id_metadata_column_map_list:
                if row[0] is not None:
                    pool_id_metadata_column_map[int(row[0])] = {
                        "column": row[1],
                        "name": row[2],
                        "type": row[3],
                        "hidden": row[4],
                    }

        # Read pool object map data (critical for pool creation)
        if "pool_object_map" in wb.sheetnames:
            pool_object_map_ws = wb["pool_object_map"]
            if pool_object_map_ws.max_row > 1:
                pool_object_map_data = [list(row) for row in pool_object_map_ws.iter_rows(min_row=2, values_only=True)]

        if replace_existing:
            metadata_table.columns.all().delete()
            metadata_table.sample_pools.all().delete()

        all_data = []
        max_rows = max(len(main_data), len(hidden_data) if hidden_data else 0)

        for i in range(max_rows):
            main_row = main_data[i] if i < len(main_data) else []
            hidden_row = hidden_data[i] if i < len(hidden_data) and hidden_data else []
            all_data.append(main_row + hidden_row)

        metadata_table.sample_count = len(all_data)
        metadata_table.save(update_fields=["sample_count"])

        columns_created = 0
        columns_updated = 0
        created_columns = []

        for column_id, column_info in id_metadata_column_map.items():
            defaults = {
                "hidden": column_info["hidden"],
                "column_position": column_info["column"],
            }

            existing_column = MetadataColumn.objects.filter(
                name=column_info["name"],
                type=column_info["type"],
                metadata_table=metadata_table,
            ).first()

            if not existing_column:
                template_columns = metadata_table.columns.filter(template__isnull=False).select_related("template")
                template_column_info = None

                for existing_col in template_columns:
                    if existing_col.template:
                        matching_template_col = existing_col.template.column_template.user_templates.filter(
                            column_name=column_info["name"]
                        ).first()

                        if matching_template_col:
                            template_column_info = matching_template_col
                            break

                if template_column_info:
                    defaults.update(
                        {
                            "ontology_type": template_column_info.ontology_type,
                            "ontology_options": template_column_info.ontology_options or [],
                            "custom_ontology_filters": template_column_info.custom_ontology_filters or {},
                            "enable_typeahead": template_column_info.enable_typeahead,
                            "excel_validation": template_column_info.excel_validation,
                            "custom_validation_rules": template_column_info.custom_validation_rules or {},
                            "api_enhancements": template_column_info.api_enhancements or {},
                            "template": template_column_info,
                        }
                    )

            metadata_column, created = MetadataColumn.objects.get_or_create(
                name=column_info["name"],
                type=column_info["type"],
                metadata_table=metadata_table,
                defaults=defaults,
            )

            if created:
                columns_created += 1
            else:
                columns_updated += 1

            created_columns.append(metadata_column)

            metadata_value_map = {}
            column_index = column_info["column"]

            for row_index, row in enumerate(all_data):
                if column_index < len(row) and row[column_index]:
                    cell_value = str(row[column_index]).strip()
                    if cell_value:
                        if validate_ontologies:
                            cell_value = metadata_column.convert_sdrf_to_metadata(cell_value)
                        metadata_value_map[row_index + 1] = cell_value

            if metadata_value_map:
                value_counts = Counter(metadata_value_map.values())
                most_common_value, _ = value_counts.most_common(1)[0]

                metadata_column.value = most_common_value

                non_default_values = {k: v for k, v in metadata_value_map.items() if v != most_common_value}
                if non_default_values:
                    modifiers = _create_modifiers(non_default_values)
                    metadata_column.modifiers = modifiers
                else:
                    metadata_column.modifiers = []
                metadata_column.save(update_fields=["value", "modifiers"])

        # Reorder table columns after creation (same as SDRF import approach)
        if created_columns:
            try:
                # Try schema-based reordering first if possible
                schema_ids = set()
                for column in created_columns:
                    if column.template and hasattr(column.template, "schema") and column.template.schema:
                        schema_ids.add(column.template.schema.id)

                if schema_ids:
                    metadata_table.reorder_columns_by_schema(schema_ids=list(schema_ids))
                else:
                    # Fall back to basic section-based ordering if available
                    if hasattr(metadata_table, "basic_column_reordering"):
                        metadata_table.basic_column_reordering()
            except Exception:
                # Silently continue if reordering fails
                pass

        pools_created = 0
        created_pools = []
        if pool_object_map_data and create_pools:
            for pool_row in pool_object_map_data:
                pool_name = pool_row[0] if pool_row[0] else f"Pool {len(created_pools) + 1}"
                pooled_only_samples = json.loads(pool_row[1]) if pool_row[1] else []
                pooled_and_independent_samples = json.loads(pool_row[2]) if pool_row[2] else []
                is_reference = pool_row[3] if len(pool_row) > 3 else False

                existing_pool = SamplePool.objects.filter(metadata_table=metadata_table, pool_name=pool_name).first()

                if existing_pool:
                    if replace_existing:
                        existing_pool.pooled_only_samples = pooled_only_samples
                        existing_pool.pooled_and_independent_samples = pooled_and_independent_samples
                        existing_pool.is_reference = is_reference
                        existing_pool.save()
                    sample_pool = existing_pool
                else:
                    sample_pool = SamplePool.objects.create(
                        metadata_table=metadata_table,
                        pool_name=pool_name,
                        pooled_only_samples=pooled_only_samples,
                        pooled_and_independent_samples=pooled_and_independent_samples,
                        is_reference=is_reference,
                        created_by=user,
                    )
                    pools_created += 1

                if not (pool_main_data or pool_hidden_data) or not pool_id_metadata_column_map:
                    create_pool_metadata_from_table_columns(sample_pool)

                created_pools.append(sample_pool)

        if (pool_main_data or pool_hidden_data) and pool_id_metadata_column_map and created_pools:
            pool_all_data = []
            max_pool_rows = max(len(pool_main_data), len(pool_hidden_data) if pool_hidden_data else 0)

            for i in range(max_pool_rows):
                pool_main_row = pool_main_data[i] if i < len(pool_main_data) else []
                pool_hidden_row = pool_hidden_data[i] if i < len(pool_hidden_data) and pool_hidden_data else []
                pool_all_data.append(pool_main_row + pool_hidden_row)

            for pool_row_index, pool_data_row in enumerate(pool_all_data):
                if pool_row_index < len(created_pools):
                    current_pool = created_pools[pool_row_index]

                    current_pool.metadata_columns.clear()

                    for column_id, column_info in pool_id_metadata_column_map.items():
                        column_index = column_info["column"]
                        cell_value = ""

                        if column_index < len(pool_data_row) and pool_data_row[column_index] is not None:
                            cell_value = str(pool_data_row[column_index]).strip()

                        table_column = metadata_table.columns.filter(name=column_info["name"]).first()

                        if table_column:
                            final_value = cell_value or "not available"
                            if final_value != "not available" and validate_ontologies:
                                final_value = table_column.convert_sdrf_to_metadata(final_value)

                            pool_metadata_column = MetadataColumn.objects.create(
                                name=column_info["name"],
                                type=column_info["type"],
                                value=final_value,
                                hidden=column_info["hidden"],
                                column_position=column_info["column"],
                                ontology_type=table_column.ontology_type,
                                ontology_options=table_column.ontology_options or [],
                                custom_ontology_filters=table_column.custom_ontology_filters or {},
                                template=table_column.template,
                                mandatory=table_column.mandatory,
                            )

                            current_pool.metadata_columns.add(pool_metadata_column)

            for pool in created_pools:
                if pool.metadata_columns.exists():
                    try:
                        schema_ids = set()
                        for column in pool.metadata_columns.all():
                            if column.template and hasattr(column.template, "schema") and column.template.schema:
                                schema_ids.add(column.template.schema.id)

                        if schema_ids:
                            pool.reorder_pool_columns_by_schema(schema_ids=list(schema_ids))
                        else:
                            pool.basic_pool_column_reordering()
                    except Exception:
                        pool.basic_pool_column_reordering()

        return {
            "success": True,
            "message": "Excel data imported successfully",
            "columns_created": columns_created,
            "columns_updated": columns_updated,
            "sample_count": metadata_table.sample_count,
            "pools_created": pools_created,
            "has_pool_data": len(pool_object_map_data) > 0,
            "warnings": [],
        }
