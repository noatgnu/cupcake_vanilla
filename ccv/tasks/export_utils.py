"""
Export utility functions for CUPCAKE Vanilla.

This module contains the core export logic that can be used by both
sync views and async RQ tasks.
"""

import io
import zipfile
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ccv.models import FavouriteMetadataOption, MetadataColumn, MetadataTable, SamplePool

# Import helper functions from the main utils module
from ccv.utils import sort_metadata, sort_pool_metadata


def _create_safe_filename(base_name: str, extension: str, max_length: int = 200) -> str:
    """
    Create a safe filename that doesn't exceed database field limits.

    Args:
        base_name: The base name to use
        extension: File extension (including the dot)
        max_length: Maximum total filename length (default 200 to leave room for variations)

    Returns:
        Safe filename that fits within database constraints
    """
    # Clean the base name - remove invalid characters
    safe_base = "".join(c for c in base_name if c.isalnum() or c in (" ", "-", "_")).rstrip()
    safe_base = safe_base.replace(" ", "_")

    # Calculate available space for base name
    available_length = max_length - len(extension)

    # Truncate if necessary
    if len(safe_base) > available_length:
        safe_base = safe_base[:available_length]

    return f"{safe_base}{extension}"


def export_sdrf_data(
    metadata_table,
    user,
    metadata_column_ids: Optional[List[int]] = None,
    include_pools: bool = True,
    validate_sdrf: bool = False,
) -> Dict[str, Any]:
    """
    Shared utility function for exporting SDRF data.

    This function contains the core SDRF export logic that can be used by both
    sync views and async RQ tasks.

    Args:
        metadata_table: MetadataTable instance to export data from
        user: User performing the export
        metadata_column_ids: Optional list of column IDs to export
        include_pools: Whether to include sample pools
        validate_sdrf: Whether to validate SDRF format

    Returns:
        Dict containing export results and file content information
    """
    # Check permissions
    if not metadata_table.can_view(user):
        raise PermissionError("Permission denied: cannot view this metadata table")

    # Get metadata columns (user can specify specific columns or get all)
    if metadata_column_ids:
        metadata_columns = metadata_table.columns.filter(id__in=metadata_column_ids)
    else:
        metadata_columns = metadata_table.columns.all()

    # Filter out hidden columns for SDRF export (SDRF standard doesn't support hidden columns)
    visible_metadata = metadata_columns.filter(hidden=False).order_by("column_position")

    # Sort metadata and get structured output for SDRF format
    result_data, _ = sort_metadata(list(visible_metadata), metadata_table.sample_count, metadata_table)

    # Add pool data if requested and pools exist (original CUPCAKE logic)
    if include_pools:
        pools = list(metadata_table.sample_pools.all())
        if pools:
            # Add SN= rows for reference pools (original CUPCAKE approach)
            for pool in pools:
                if pool.is_reference and pool.sdrf_value:
                    # Create a pool row following original CUPCAKE format
                    pool_row = ["" for _ in result_data[0]]  # Initialize with empty values

                    # Find the pooled sample column index
                    pooled_column_index = None
                    for i, header in enumerate(result_data[0]):
                        header_lower = header.lower()
                        if "pooled sample" in header_lower or "pooled_sample" in header_lower:
                            pooled_column_index = i
                            break

                    # Find the source name column index
                    source_name_column_index = None
                    for i, header in enumerate(result_data[0]):
                        header_lower = header.lower()
                        if "source name" in header_lower or "source_name" in header_lower:
                            source_name_column_index = i
                            break

                    # Set pool data in the row
                    if pooled_column_index is not None:
                        pool_row[pooled_column_index] = pool.sdrf_value

                    if source_name_column_index is not None:
                        pool_row[source_name_column_index] = pool.pool_name

                    # Fill other columns with pool-specific default values
                    for i, metadata_column in enumerate(visible_metadata):
                        if i < len(pool_row) and not pool_row[i]:  # Only fill empty cells
                            # Use column default value or "not applicable" for required fields
                            if metadata_column.value:
                                pool_row[i] = metadata_column.value
                            elif metadata_column.name.lower() in [
                                "organism",
                                "disease",
                                "organism part",
                                "tissue",
                            ]:
                                pool_row[i] = "not applicable"
                            else:
                                pool_row[i] = "not available"

                    # Add pool row to results
                    result_data.append(pool_row)

    # Convert to tab-separated format (SDRF standard)
    sdrf_content = []
    for row in result_data:
        # Convert all values to strings and handle None values
        str_row = [str(cell) if cell is not None else "" for cell in row]
        sdrf_content.append("\t".join(str_row))

    sdrf_text = "\n".join(sdrf_content)

    # Validate SDRF format if validation was requested
    validation_results = None
    if validate_sdrf:
        try:
            validation_results = validate_sdrf(result_data)
        except Exception as e:
            validation_results = {
                "errors": [f"SDRF validation failed: {str(e)}"],
                "warnings": [],
                "suggestions": [],
            }

    # Create filename based on metadata table name
    filename = _create_safe_filename(metadata_table.name, ".sdrf.tsv")
    content_type = "text/tab-separated-values"
    file_data = sdrf_text.encode("utf-8")

    return {
        "success": True,
        "filename": filename,
        "file_size": len(file_data),
        "content_type": content_type,
        "file_data": file_data,
        "sdrf_content": sdrf_text,
        "metadata_table_name": metadata_table.name,
        "column_count": visible_metadata.count(),
        "sample_count": metadata_table.sample_count,
        "pool_count": metadata_table.sample_pools.count() if include_pools else 0,
        "validation_results": validation_results,
    }


def export_excel_template(
    metadata_columns: List[MetadataColumn],
    sample_number: int,
    pools: List[SamplePool] = None,
    favourites: Dict[str, List[str]] = None,
    field_mask_mapping: Dict[str, str] = None,
    metadata_table=None,
) -> Workbook:
    """
    Export metadata as Excel template with dropdowns and validation.

    Args:
        metadata_columns: List of metadata columns
        sample_number: Number of samples
        pools: Optional list of sample pools
        favourites: Optional dictionary of favourite options
        field_mask_mapping: Optional field name mappings
        metadata_table: Optional metadata table for pooled status determination

    Returns:
        Excel Workbook object
    """
    if favourites is None:
        favourites = {}
    if field_mask_mapping is None:
        field_mask_mapping = {}

    # Separate main and hidden metadata
    main_metadata = [m for m in metadata_columns if not m.hidden]
    hidden_metadata = [m for m in metadata_columns if m.hidden]

    # Sort metadata (original CUPCAKE approach with metadata_table for pooled status)
    result_main, id_map_main = sort_metadata(main_metadata, sample_number, metadata_table)
    result_hidden, id_map_hidden = (
        sort_metadata(hidden_metadata, sample_number, metadata_table) if hidden_metadata else ([], {})
    )

    # Handle pools (original CUPCAKE logic)
    has_pools = pools and len(pools) > 0
    pool_result_main, pool_id_map_main = ([], {})
    pool_result_hidden, pool_id_map_hidden = ([], {})

    if has_pools:
        # Get pool metadata columns - same as main metadata
        pool_main_metadata = [m for m in metadata_columns if not m.hidden]
        pool_hidden_metadata = [m for m in metadata_columns if m.hidden]

        # Generate pool data using sort_pool_metadata utility
        pool_result_main, pool_id_map_main = sort_pool_metadata(pool_main_metadata, pools)
        pool_result_hidden, pool_id_map_hidden = (
            sort_pool_metadata(pool_hidden_metadata, pools) if pool_hidden_metadata else ([], {})
        )

    # Create workbook
    wb = Workbook()
    main_ws = wb.active
    main_ws.title = "main"

    # Create worksheets
    hidden_ws = wb.create_sheet(title="hidden")
    id_metadata_column_map_ws = wb.create_sheet(title="id_metadata_column_map")

    # Create pool sheets if pools exist (original CUPCAKE structure)
    pool_main_ws = None
    pool_hidden_ws = None
    pool_id_metadata_column_map_ws = None
    pool_object_map_ws = None

    if has_pools:
        pool_main_ws = wb.create_sheet(title="pool_main")
        pool_hidden_ws = wb.create_sheet(title="pool_hidden")
        pool_id_metadata_column_map_ws = wb.create_sheet(title="pool_id_metadata_column_map")
        pool_object_map_ws = wb.create_sheet(title="pool_object_map")

    # Fill ID mapping worksheet
    id_metadata_column_map_ws.append(["id", "column", "name", "type", "hidden"])
    for k, v in id_map_main.items():
        id_metadata_column_map_ws.append([k, v["column"], v["name"], v["type"], v["hidden"]])
    for k, v in id_map_hidden.items():
        id_metadata_column_map_ws.append([k, v["column"], v["name"], v["type"], v["hidden"]])

    # Styling
    fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Fill main worksheet
    if result_main:
        # Add headers and data
        main_ws.append(result_main[0])
        main_work_area = f"A1:{get_column_letter(len(result_main[0]))}{sample_number + 1}"

        for row in result_main[1:]:
            main_ws.append(row)

        # Apply styling
        for row in main_ws[main_work_area]:
            for cell in row:
                cell.fill = fill
                cell.border = thin_border

        # Auto-adjust column widths
        for col in main_ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except (TypeError, AttributeError):
                    pass
            adjusted_width = max_length + 2
            main_ws.column_dimensions[column].width = adjusted_width

        # Add notes
        note_texts = [
            "Note: Empty cells will be filled with 'not applicable' or " "'not available' when submitted.",
            "[*] User-specific favourite options.",
            "[**] Lab group-recommended options.",
            "[***] Global recommendations.",
        ]

        start_row = sample_number + 2
        for i, note_text in enumerate(note_texts):
            main_ws.merge_cells(
                start_row=start_row + i,
                start_column=1,
                end_row=start_row + i,
                end_column=len(result_main[0]),
            )
            note_cell = main_ws.cell(row=start_row + i, column=1)
            note_cell.value = note_text
            note_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Add dropdown validation
        _add_dropdown_validation(main_ws, result_main[0], favourites, field_mask_mapping, sample_number)

    # Fill hidden worksheet
    if result_hidden:
        hidden_ws.append(result_hidden[0])
        for row in result_hidden[1:]:
            hidden_ws.append(row)

        # Apply styling and validation
        _add_dropdown_validation(hidden_ws, result_hidden[0], favourites, field_mask_mapping, sample_number)

    # Fill pool data and mapping if pools exist (original CUPCAKE logic)
    if has_pools and pool_id_metadata_column_map_ws:
        # Fill pool_main sheet
        if pool_result_main:
            for row_data in pool_result_main:
                pool_main_ws.append(row_data)

        # Fill pool_hidden sheet if there's hidden data
        if pool_result_hidden:
            for row_data in pool_result_hidden:
                pool_hidden_ws.append(row_data)

        # Fill pool ID metadata column mapping
        pool_id_metadata_column_map_ws.append(["id", "column", "name", "type", "hidden"])
        for k, v in pool_id_map_main.items():
            pool_id_metadata_column_map_ws.append([k, v["column"], v["name"], v["type"], v["hidden"]])
        for k, v in pool_id_map_hidden.items():
            pool_id_metadata_column_map_ws.append([k, v["column"], v["name"], v["type"], v["hidden"]])

        # Fill pool object mapping sheet (critical for pooled status)
        import json

        pool_object_map_ws.append(
            [
                "pool_name",
                "pooled_only_samples",
                "pooled_and_independent_samples",
                "is_reference",
                "sdrf_value",
            ]
        )
        for pool in pools:
            pool_object_map_ws.append(
                [
                    pool.pool_name,
                    (json.dumps(pool.pooled_only_samples) if pool.pooled_only_samples else "[]"),
                    (json.dumps(pool.pooled_and_independent_samples) if pool.pooled_and_independent_samples else "[]"),
                    pool.is_reference,
                    pool.sdrf_value or "",
                ]
            )

        # Apply styling and validation to pool worksheets
        if pool_result_main:
            _add_dropdown_validation(
                pool_main_ws,
                pool_result_main[0],
                favourites,
                field_mask_mapping,
                len(pools),
            )

    return wb


def _add_dropdown_validation(
    worksheet,
    headers: List[str],
    favourites: Dict[str, List[str]],
    field_mask_mapping: Dict[str, str],
    row_count: int,
) -> None:
    """
    Add dropdown validation to a worksheet.

    Args:
        worksheet: Excel worksheet object
        headers: List of column headers
        favourites: Dictionary of favourite options
        field_mask_mapping: Field name mappings
        row_count: Number of data rows
    """
    required_metadata_names = {"tissue", "organism part", "disease", "species"}

    for i, header in enumerate(headers):
        name_splitted = header.split("[")
        if len(name_splitted) > 1:
            name = name_splitted[1].replace("]", "")
        else:
            name = name_splitted[0]

        required_column = name.lower() in required_metadata_names

        # Apply field masking
        name_capitalized = name.capitalize().replace("Ms1", "MS1").replace("Ms2", "MS2")
        if name_capitalized in field_mask_mapping:
            display_name = field_mask_mapping[name_capitalized]
            if len(name_splitted) > 1:
                worksheet.cell(row=1, column=i + 1).value = header.replace(
                    name_splitted[1].rstrip("]"), display_name.lower()
                )
            else:
                worksheet.cell(row=1, column=i + 1).value = display_name.lower()

        # Build option list
        option_list = []
        if required_column:
            option_list.append("not applicable")
        else:
            option_list.append("not available")

        if name.lower() in favourites:
            option_list.extend(favourites[name.lower()])

        # Add validation
        if option_list:
            # Ensure "not applicable" and "not available" stay lowercase
            processed_options = []
            for option in option_list:
                if option.lower() in ["not applicable", "not available"]:
                    processed_options.append(option.lower())
                else:
                    processed_options.append(option)

            dv = DataValidation(type="list", formula1=f'"{",".join(processed_options)}"', showDropDown=False)
            col_letter = get_column_letter(i + 1)
            worksheet.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}{row_count + 1}")


def export_excel_template_data(
    metadata_table,
    user,
    metadata_column_ids: Optional[List[int]] = None,
    include_pools: bool = True,
    lab_group_ids: Optional[List[int]] = None,
) -> Dict[str, Any]:
    """
    Shared utility function for exporting Excel template data.

    This function contains the core Excel template export logic that can be used by both
    sync views and async RQ tasks.

    Args:
        metadata_table: MetadataTable instance to export template from
        user: User performing the export
        metadata_column_ids: Optional list of column IDs to export
        include_pools: Whether to include sample pools
        lab_group_ids: Optional list of lab group IDs for favourites

    Returns:
        Dict containing export results and file content information
    """
    import re

    # Check permissions
    if not metadata_table.can_view(user):
        raise PermissionError("Permission denied: cannot view this metadata table")

    # Get metadata columns (user can specify specific columns or get all)
    if metadata_column_ids:
        metadata_columns = metadata_table.columns.filter(id__in=metadata_column_ids)
    else:
        metadata_columns = metadata_table.columns.all()

    # Get pools
    pools = list(metadata_table.sample_pools.all()) if include_pools else []

    # Get favourites for each metadata column based on column names
    favourites = {}

    # Get column names from actual metadata columns being exported
    column_names = set(column.name.lower() for column in metadata_columns)

    # User-specific favourites - use case-insensitive name matching
    user_favourites = FavouriteMetadataOption.objects.filter(
        user=user,
        lab_group__isnull=True,
        name__iregex=r"^(" + "|".join(re.escape(name) for name in column_names) + ")$",
    )
    for fav in user_favourites:
        if fav.name.lower() not in favourites:
            favourites[fav.name.lower()] = []
        favourites[fav.name.lower()].append(f"[{fav.id}] {fav.display_value}[*]")

    # Lab group favourites
    if lab_group_ids is not None:  # Check for None vs empty list
        if lab_group_ids == []:
            # Empty list means "all lab groups"
            lab_favourites = FavouriteMetadataOption.objects.filter(
                lab_group__isnull=False,
                name__iregex=r"^(" + "|".join(re.escape(name) for name in column_names) + ")$",
            )
        else:
            # Specific lab group IDs
            lab_favourites = FavouriteMetadataOption.objects.filter(
                lab_group_id__in=lab_group_ids,
                name__iregex=r"^(" + "|".join(re.escape(name) for name in column_names) + ")$",
            )

        for fav in lab_favourites:
            if fav.name.lower() not in favourites:
                favourites[fav.name.lower()] = []
            favourites[fav.name.lower()].append(f"[{fav.id}] {fav.display_value}[**]")
            # Add "not applicable" for required metadata
            if fav.name.lower() == "tissue" or fav.name.lower() == "organism part":
                favourites[fav.name.lower()].append("not applicable")

    # Global recommendations
    global_favourites = FavouriteMetadataOption.objects.filter(
        is_global=True, name__iregex=r"^(" + "|".join(re.escape(name) for name in column_names) + ")$"
    )
    for fav in global_favourites:
        if fav.name.lower() not in favourites:
            favourites[fav.name.lower()] = []
        favourites[fav.name.lower()].append(f"[{fav.id}] {fav.display_value}[***]")

    # Create Excel workbook using existing utility
    try:
        wb = export_excel_template(
            metadata_columns=list(metadata_columns),
            sample_number=metadata_table.sample_count,
            pools=pools,
            favourites=favourites,
            metadata_table=metadata_table,
        )

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Get file data and create filename
        file_data = output.getvalue()
        print(f"Excel workbook created successfully, file size: {len(file_data)} bytes")

        filename = _create_safe_filename(metadata_table.name, "_template.xlsx")
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    except Exception as e:
        print(f"Error creating Excel workbook: {str(e)}")
        import traceback

        print(f"Full traceback: {traceback.format_exc()}")
        raise

    return {
        "success": True,
        "filename": filename,
        "file_size": len(file_data),
        "content_type": content_type,
        "file_data": file_data,
        "metadata_table_name": metadata_table.name,
        "column_count": metadata_columns.count(),
        "pool_count": len(pools),
    }


def export_multiple_sdrf_data(
    metadata_table_ids: List[int],
    user,
    include_pools: bool = True,
    validate_sdrf: bool = False,
) -> Dict[str, Any]:
    """
    Shared utility function for exporting multiple SDRF files as a ZIP archive.

    Args:
        metadata_table_ids: List of MetadataTable IDs to export
        user: User performing the export
        include_pools: Whether to include sample pools
        validate_sdrf: Whether to validate SDRF format

    Returns:
        Dict containing export results and ZIP file content information
    """
    # Validate tables and check permissions
    metadata_tables = []
    for table_id in metadata_table_ids:
        try:
            table = MetadataTable.objects.get(id=table_id)
            if not table.can_view(user):
                raise PermissionError(f"Permission denied: cannot view metadata table '{table.name}'")
            metadata_tables.append(table)
        except MetadataTable.DoesNotExist:
            raise ValueError(f"Metadata table with ID {table_id} does not exist")

    if not metadata_tables:
        raise ValueError("No valid metadata tables provided")

    # Create ZIP file in memory
    zip_buffer = io.BytesIO()
    exported_files = []
    validation_results_all = {}

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for table in metadata_tables:
            try:
                # Export SDRF data for this table
                result = export_sdrf_data(
                    metadata_table=table,
                    user=user,
                    metadata_column_ids=None,  # Export all columns for bulk export
                    include_pools=include_pools,
                    validate_sdrf=validate_sdrf,
                )

                # Add file to ZIP
                zip_file.writestr(result["filename"], result["sdrf_content"])

                exported_files.append(
                    {
                        "table_id": table.id,
                        "table_name": table.name,
                        "filename": result["filename"],
                        "file_size": result["file_size"],
                        "column_count": result["column_count"],
                        "sample_count": result["sample_count"],
                        "pool_count": result["pool_count"],
                    }
                )

                # Collect validation results if available
                if result.get("validation_results"):
                    validation_results_all[table.name] = result["validation_results"]

            except Exception as e:
                # Add error info to exported files list
                exported_files.append(
                    {
                        "table_id": table.id,
                        "table_name": table.name,
                        "error": str(e),
                    }
                )

    # Get ZIP file data
    zip_buffer.seek(0)
    zip_data = zip_buffer.getvalue()

    # Create filename for ZIP
    if len(metadata_tables) == 1:
        zip_filename = _create_safe_filename(metadata_tables[0].name, "_sdrf.zip")
    else:
        zip_filename = f"bulk_sdrf_export_{len(metadata_tables)}_tables.zip"

    # Calculate totals
    successful_exports = [f for f in exported_files if "error" not in f]
    total_columns = sum(f.get("column_count", 0) for f in successful_exports)
    total_samples = sum(f.get("sample_count", 0) for f in successful_exports)
    total_pools = sum(f.get("pool_count", 0) for f in successful_exports)

    return {
        "success": True,
        "filename": zip_filename,
        "file_size": len(zip_data),
        "content_type": "application/zip",
        "file_data": zip_data,
        "exported_files": exported_files,
        "successful_exports": len(successful_exports),
        "failed_exports": len(exported_files) - len(successful_exports),
        "total_tables": len(metadata_tables),
        "total_columns": total_columns,
        "total_samples": total_samples,
        "total_pools": total_pools,
        "validation_results": validation_results_all,
    }


def export_multiple_excel_template_data(
    metadata_table_ids: List[int],
    user,
    metadata_column_ids: Optional[List[int]] = None,
    include_pools: bool = True,
    lab_group_ids: Optional[List[int]] = None,
) -> Dict[str, Any]:
    """
    Export multiple metadata tables as Excel templates in a ZIP archive.

    Args:
        metadata_table_ids: List of MetadataTable IDs to export
        user: User performing the export
        metadata_column_ids: Optional list of column IDs to export
        include_pools: Whether to include sample pools
        lab_group_ids: Optional list of lab group IDs for favourites

    Returns:
        Dict with ZIP file data and export statistics
    """
    from django.utils import timezone

    # Create ZIP file in memory
    zip_buffer = io.BytesIO()
    exported_files = []
    failed_exports = []
    total_columns = 0
    total_pools = 0

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for table_id in metadata_table_ids:
            try:
                # Get the metadata table
                metadata_table = MetadataTable.objects.get(id=table_id)

                # Check user permissions
                if not metadata_table.can_view(user):
                    failed_exports.append(
                        {
                            "table_id": table_id,
                            "table_name": getattr(metadata_table, "name", f"Table {table_id}"),
                            "error": "Permission denied",
                        }
                    )
                    continue

                # Export Excel template for this table
                result = export_excel_template_data(
                    metadata_table=metadata_table,
                    user=user,
                    metadata_column_ids=metadata_column_ids,
                    include_pools=include_pools,
                    lab_group_ids=lab_group_ids,
                )

                # Add file to ZIP
                zip_file.writestr(result["filename"], result["file_data"])

                # Track successful export
                exported_files.append(
                    {
                        "table_id": table_id,
                        "table_name": metadata_table.name,
                        "filename": result["filename"],
                        "file_size": result["file_size"],
                        "column_count": result.get("column_count", 0),
                        "pool_count": result.get("pool_count", 0),
                    }
                )

                # Update totals
                total_columns += result.get("column_count", 0)
                total_pools += result.get("pool_count", 0)

            except MetadataTable.DoesNotExist:
                failed_exports.append(
                    {"table_id": table_id, "table_name": f"Table {table_id}", "error": "Table not found"}
                )
            except Exception as e:
                table_name = f"Table {table_id}"
                try:
                    metadata_table = MetadataTable.objects.get(id=table_id)
                    table_name = metadata_table.name
                except Exception:
                    pass
                failed_exports.append({"table_id": table_id, "table_name": table_name, "error": str(e)})

    # Generate filename with timestamp
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f"bulk_excel_templates_{timestamp}.zip"

    return {
        "filename": filename,
        "file_data": zip_buffer.getvalue(),
        "content_type": "application/zip",
        "file_size": zip_buffer.tell(),
        "exported_files": exported_files,
        "successful_exports": len(exported_files),
        "failed_exports": len(failed_exports),
        "total_tables": len(metadata_table_ids),
        "total_columns": total_columns,
        "total_pools": total_pools,
    }
