"""
Tests for Excel export favourite options appearing in dropdown validation and
round-trip import of exported data containing favourite option markers.
"""

import io

from django.contrib.auth.models import User
from django.test import TestCase

from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from ccc.models import LabGroup
from ccv.models import FavouriteMetadataOption, MetadataColumn, MetadataColumnTemplate, MetadataTable
from ccv.tasks.export_utils import export_excel_template_data
from ccv.tasks.import_utils import _resolve_favourite_cell_value, import_excel_data


class ExcelExportFavouritesTest(TestCase):
    """Verify that saved favourite options appear in the Excel dropdown for export."""

    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(username="exportuser", password="testpass")
        self.lab_group = LabGroup.objects.create(name="Test Lab", creator=self.user)
        self.lab_group.members.add(self.user)

        self.table = MetadataTable.objects.create(
            name="Test Table",
            owner=self.user,
            lab_group=self.lab_group,
            sample_count=3,
        )

        self.organism_template = MetadataColumnTemplate.objects.create(
            column_name="characteristics[organism]",
            column_type="characteristics",
            ontology_type="species",
            is_active=True,
            owner=self.user,
            lab_group=self.lab_group,
            not_available=True,
            possible_default_values=["homo sapiens", "mus musculus"],
        )

        self.column = MetadataColumn.objects.create(
            metadata_table=self.table,
            name="characteristics[organism]",
            type="characteristics",
            column_position=0,
            template=self.organism_template,
        )

        # Stored as inner name (legacy path)
        FavouriteMetadataOption.objects.create(
            user=self.user,
            lab_group=None,
            name="organism",
            type="characteristics",
            value="homo sapiens",
            display_value="Homo sapiens",
            is_global=False,
        )
        # Stored as full SDRF name (frontend typeahead path)
        FavouriteMetadataOption.objects.create(
            user=self.user,
            lab_group=None,
            name="characteristics[organism]",
            type="characteristics",
            value="mus musculus",
            display_value="Mus musculus",
            is_global=False,
        )

        self.client.force_authenticate(user=self.user)

    def _export_and_get_dropdowns(self):
        response = self.client.post(
            "/api/v1/metadata-management/export_excel_template/",
            {
                "metadata_table_id": self.table.id,
                "metadata_column_ids": [self.column.id],
                "sample_number": self.table.sample_count,
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        wb = load_workbook(
            io.BytesIO(
                b"".join(response.streaming_content if hasattr(response, "streaming_content") else [response.content])
            )
        )
        ws = wb["main"]
        options_by_col = {}
        for dv in ws.data_validations.dataValidation:
            if dv.type == "list" and dv.formula1:
                raw = dv.formula1.strip('"')
                options = [o.strip() for o in raw.split(",")]
                for cell_range in dv.sqref.ranges:
                    col = cell_range.min_col
                    options_by_col[col] = options
        return options_by_col

    def test_favourite_options_appear_in_dropdown(self):
        """Homo sapiens and mus musculus favourites must appear in the organism dropdown."""
        options_by_col = self._export_and_get_dropdowns()
        self.assertTrue(options_by_col, "No data validation found in the exported Excel file")
        all_options = []
        for opts in options_by_col.values():
            all_options.extend(opts)

        has_homo = any("homo sapiens" in o.lower() for o in all_options)
        has_mus = any("mus musculus" in o.lower() for o in all_options)
        self.assertTrue(has_homo, f"'homo sapiens' not found in dropdown options: {all_options}")
        self.assertTrue(has_mus, f"'mus musculus' not found in dropdown options: {all_options}")

    def test_not_available_from_template(self):
        """not_available must appear when the linked template allows it (even if column flag is False)."""
        self.column.not_available = False
        self.column.save()

        options_by_col = self._export_and_get_dropdowns()
        all_options = []
        for opts in options_by_col.values():
            all_options.extend(opts)

        has_not_available = any("not available" in o.lower() for o in all_options)
        self.assertTrue(has_not_available, f"'not available' not found despite template allowing it: {all_options}")

    def test_sdrf_imported_column_name_format(self):
        """Column stored as 'characteristics[organism]' must still match favourites stored as 'organism'."""
        self.assertEqual(self.column.name, "characteristics[organism]")
        options_by_col = self._export_and_get_dropdowns()
        all_options = []
        for opts in options_by_col.values():
            all_options.extend(opts)
        has_homo = any("homo sapiens" in o.lower() for o in all_options)
        self.assertTrue(has_homo, f"SDRF-format column name did not match favourites. Options: {all_options}")


def _get_dropdowns_from_wb(wb):
    ws = wb["main"]
    options_by_col = {}
    for dv in ws.data_validations.dataValidation:
        if dv.type == "list" and dv.formula1:
            raw = dv.formula1.strip('"')
            options = [o.strip() for o in raw.split(",")]
            for cell_range in dv.sqref.ranges:
                col = cell_range.min_col
                options_by_col[col] = options
    return options_by_col


class AsyncExportFavouritesTest(TestCase):
    """Verify favourites appear via the export_excel_template_data path (used by the GUI async export)."""

    def setUp(self):
        self.user = User.objects.create_user(username="asyncuser", password="testpass")
        self.lab_group = LabGroup.objects.create(name="Test Lab", creator=self.user)
        self.lab_group.members.add(self.user)

        self.table = MetadataTable.objects.create(
            name="Async Test Table",
            owner=self.user,
            lab_group=self.lab_group,
            sample_count=3,
        )

        self.organism_template = MetadataColumnTemplate.objects.create(
            column_name="characteristics[organism]",
            column_type="characteristics",
            ontology_type="species",
            is_active=True,
            owner=self.user,
            lab_group=self.lab_group,
            not_available=True,
            possible_default_values=["homo sapiens", "mus musculus"],
        )

        self.column = MetadataColumn.objects.create(
            metadata_table=self.table,
            name="characteristics[organism]",
            type="characteristics",
            column_position=0,
            template=self.organism_template,
        )

        # Stored as inner name (legacy path)
        FavouriteMetadataOption.objects.create(
            user=self.user,
            lab_group=None,
            name="organism",
            type="characteristics",
            value="homo sapiens",
            display_value="Homo sapiens",
            is_global=False,
        )
        # Stored as full SDRF name (frontend typeahead path)
        FavouriteMetadataOption.objects.create(
            user=self.user,
            lab_group=None,
            name="characteristics[organism]",
            type="characteristics",
            value="mus musculus",
            display_value="Mus musculus",
            is_global=False,
        )

    def test_async_path_favourite_options(self):
        """export_excel_template_data (GUI async path) must include favourites in dropdowns."""
        result = export_excel_template_data(self.table, self.user)
        wb = load_workbook(io.BytesIO(result["file_data"]))
        options_by_col = _get_dropdowns_from_wb(wb)
        all_options = []
        for opts in options_by_col.values():
            all_options.extend(opts)
        has_homo = any("homo sapiens" in o.lower() for o in all_options)
        has_mus = any("mus musculus" in o.lower() for o in all_options)
        self.assertTrue(has_homo, f"'homo sapiens' not in async export dropdowns: {all_options}")
        self.assertTrue(has_mus, f"'mus musculus' not in async export dropdowns: {all_options}")

    def test_async_path_not_available_from_template(self):
        """export_excel_template_data must include 'not available' when template allows it even if column flag is False."""
        self.column.not_available = False
        self.column.save()
        result = export_excel_template_data(self.table, self.user)
        wb = load_workbook(io.BytesIO(result["file_data"]))
        options_by_col = _get_dropdowns_from_wb(wb)
        all_options = []
        for opts in options_by_col.values():
            all_options.extend(opts)
        has_not_available = any("not available" in o.lower() for o in all_options)
        self.assertTrue(has_not_available, f"'not available' missing from async export despite template: {all_options}")


class ResolveFavouriteCellValueTest(TestCase):
    """Unit tests for _resolve_favourite_cell_value helper."""

    def setUp(self):
        self.user = User.objects.create_user(username="favresolveuser", password="x")
        self.fav = FavouriteMetadataOption.objects.create(
            user=self.user,
            lab_group=None,
            name="organism",
            type="characteristics",
            value="homo sapiens",
            display_value="Homo sapiens",
            is_global=False,
        )

    def test_personal_marker_resolves_to_value(self):
        """[id] display[*] returns the stored FavouriteMetadataOption.value."""
        result = _resolve_favourite_cell_value(f"[{self.fav.id}] Homo sapiens[*]")
        self.assertEqual(result, "homo sapiens")

    def test_lab_group_marker_resolves(self):
        """[id] display[**] (lab group marker) is also resolved."""
        result = _resolve_favourite_cell_value(f"[{self.fav.id}] Homo sapiens[**]")
        self.assertEqual(result, "homo sapiens")

    def test_global_marker_resolves(self):
        """[id] display[***] (global marker) is also resolved."""
        result = _resolve_favourite_cell_value(f"[{self.fav.id}] Homo sapiens[***]")
        self.assertEqual(result, "homo sapiens")

    def test_nonexistent_id_falls_back_to_display_value(self):
        """When the favourite ID no longer exists, the display value is used."""
        result = _resolve_favourite_cell_value("[99999] Fallback value[*]")
        self.assertEqual(result, "Fallback value")

    def test_plain_value_is_unchanged(self):
        """Values without the marker format pass through unmodified."""
        result = _resolve_favourite_cell_value("homo sapiens")
        self.assertEqual(result, "homo sapiens")

    def test_ontology_formatted_value_is_unchanged(self):
        """SDRF ontology format (NT=...;AC=...) passes through unmodified."""
        result = _resolve_favourite_cell_value("NT=Homo sapiens;AC=NCBITaxon:9606")
        self.assertEqual(result, "NT=Homo sapiens;AC=NCBITaxon:9606")


def _build_excel_bytes(column_id: int, column_position: int, column_name: str, column_type: str, rows: list) -> bytes:
    """Build a minimal Excel workbook with main and id_metadata_column_map sheets."""
    wb = Workbook()
    main_ws = wb.active
    main_ws.title = "main"
    main_ws.append([column_name])
    for row in rows:
        main_ws.append([row])

    id_map_ws = wb.create_sheet(title="id_metadata_column_map")
    id_map_ws.append(["id", "column", "name", "type", "hidden"])
    id_map_ws.append([column_id, column_position, column_name, column_type, False])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class ExcelImportRoundTripTest(TestCase):
    """Verify that re-importing an exported Excel with favourite markers strips them correctly."""

    def setUp(self):
        self.user = User.objects.create_user(username="roundtripuser", password="x")
        self.lab_group = LabGroup.objects.create(name="Lab", creator=self.user)
        self.lab_group.members.add(self.user)

        self.table = MetadataTable.objects.create(
            name="RT Table",
            owner=self.user,
            lab_group=self.lab_group,
            sample_count=2,
        )

        self.template = MetadataColumnTemplate.objects.create(
            column_name="characteristics[organism]",
            column_type="characteristics",
            ontology_type="species",
            is_active=True,
            owner=self.user,
            lab_group=self.lab_group,
            not_available=True,
            possible_default_values=["homo sapiens", "mus musculus"],
        )

        self.column = MetadataColumn.objects.create(
            metadata_table=self.table,
            name="characteristics[organism]",
            type="characteristics",
            column_position=0,
            template=self.template,
        )

        self.fav_homo = FavouriteMetadataOption.objects.create(
            user=self.user,
            lab_group=None,
            name="characteristics[organism]",
            type="characteristics",
            value="homo sapiens",
            display_value="Homo sapiens",
            is_global=False,
        )
        self.fav_mus = FavouriteMetadataOption.objects.create(
            user=self.user,
            lab_group=None,
            name="characteristics[organism]",
            type="characteristics",
            value="mus musculus",
            display_value="Mus musculus",
            is_global=False,
        )

    def test_round_trip_favourite_marker_stripped_on_import(self):
        """Cells filled from the dropdown (e.g. '[1] Homo sapiens[*]') are stored as clean values."""
        cell_1 = f"[{self.fav_homo.id}] Homo sapiens[*]"
        cell_2 = f"[{self.fav_mus.id}] Mus musculus[*]"

        excel_bytes = _build_excel_bytes(
            self.column.id, 0, "characteristics[organism]", "characteristics", [cell_1, cell_2]
        )

        result = import_excel_data(
            file_data=excel_bytes,
            metadata_table=self.table,
            user=self.user,
            validate_ontologies=False,
        )

        self.assertTrue(result.get("success"), result.get("error"))
        self.column.refresh_from_db()
        self.assertNotIn("[*]", self.column.value or "")
        self.assertNotIn("[*]", str(self.column.modifiers or ""))

    def test_round_trip_full_export_then_import(self):
        """Export produces a valid Excel; re-importing that file without cell data succeeds."""
        export_result = export_excel_template_data(self.table, self.user)
        self.assertTrue(export_result.get("success"), export_result.get("error"))

        result = import_excel_data(
            file_data=export_result["file_data"],
            metadata_table=self.table,
            user=self.user,
            validate_ontologies=False,
        )
        self.assertTrue(result.get("success"), result.get("error"))

    def test_plain_text_value_unchanged_on_import(self):
        """Plain text values (not favourite-formatted) are stored as-is."""
        excel_bytes = _build_excel_bytes(
            self.column.id, 0, "characteristics[organism]", "characteristics", ["homo sapiens", "mus musculus"]
        )

        result = import_excel_data(
            file_data=excel_bytes,
            metadata_table=self.table,
            user=self.user,
            validate_ontologies=False,
        )

        self.assertTrue(result.get("success"), result.get("error"))
        self.column.refresh_from_db()
        self.assertIn(self.column.value, ["homo sapiens", "mus musculus"])
